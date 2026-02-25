"""
自定义报表服务
支持报表模板 CRUD、按模板生成报表、定时订阅管理
"""
import uuid
import io
import csv
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, timedelta, date

from sqlalchemy import select, and_, or_, func
from sqlalchemy.ext.asyncio import AsyncSession
import structlog

from src.core.database import get_db_session
from src.models.report_template import ReportTemplate, ScheduledReport, ReportFormat, ScheduleFrequency
from src.models.finance import FinancialTransaction
from src.models.inventory import InventoryItem
from src.models.order import Order
from src.models.kpi import KPIRecord

logger = structlog.get_logger()

# 支持的数据源及其可用字段
DATA_SOURCE_FIELDS: Dict[str, List[Dict[str, str]]] = {
    "transactions": [
        {"field": "transaction_date", "label": "交易日期", "type": "date"},
        {"field": "transaction_type", "label": "交易类型", "type": "string"},
        {"field": "category", "label": "分类", "type": "string"},
        {"field": "subcategory", "label": "子分类", "type": "string"},
        {"field": "amount", "label": "金额（元）", "type": "currency"},
        {"field": "description", "label": "描述", "type": "string"},
        {"field": "payment_method", "label": "支付方式", "type": "string"},
        {"field": "store_id", "label": "门店ID", "type": "string"},
    ],
    "inventory": [
        {"field": "name", "label": "物品名称", "type": "string"},
        {"field": "category", "label": "分类", "type": "string"},
        {"field": "current_quantity", "label": "当前库存", "type": "number"},
        {"field": "min_quantity", "label": "最低库存", "type": "number"},
        {"field": "unit", "label": "单位", "type": "string"},
        {"field": "unit_cost", "label": "单价（元）", "type": "currency"},
        {"field": "status", "label": "状态", "type": "string"},
        {"field": "store_id", "label": "门店ID", "type": "string"},
    ],
    "orders": [
        {"field": "order_number", "label": "订单号", "type": "string"},
        {"field": "status", "label": "状态", "type": "string"},
        {"field": "total_amount", "label": "总金额（元）", "type": "currency"},
        {"field": "table_number", "label": "桌号", "type": "string"},
        {"field": "created_at", "label": "下单时间", "type": "datetime"},
        {"field": "store_id", "label": "门店ID", "type": "string"},
    ],
    "kpi": [
        {"field": "record_date", "label": "记录日期", "type": "date"},
        {"field": "value", "label": "实际值", "type": "number"},
        {"field": "target_value", "label": "目标值", "type": "number"},
        {"field": "achievement_rate", "label": "达成率", "type": "percent"},
        {"field": "status", "label": "状态", "type": "string"},
        {"field": "store_id", "label": "门店ID", "type": "string"},
    ],
}


class CustomReportService:
    """自定义报表服务"""

    # ------------------------------------------------------------------ #
    # 模板 CRUD                                                            #
    # ------------------------------------------------------------------ #

    async def list_templates(
        self,
        user_id: str,
        store_id: Optional[str] = None,
        skip: int = 0,
        limit: int = 50,
    ) -> Tuple[List[ReportTemplate], int]:
        """获取用户可见的模板列表（自己创建的 + 公开的）"""
        async with get_db_session() as session:
            conditions = [
                or_(
                    ReportTemplate.created_by == user_id,
                    ReportTemplate.is_public == True,
                )
            ]
            if store_id:
                conditions.append(
                    or_(ReportTemplate.store_id == store_id, ReportTemplate.store_id.is_(None))
                )

            count_stmt = select(func.count(ReportTemplate.id)).where(and_(*conditions))
            total = (await session.execute(count_stmt)).scalar() or 0

            stmt = (
                select(ReportTemplate)
                .where(and_(*conditions))
                .order_by(ReportTemplate.created_at.desc())
                .offset(skip)
                .limit(limit)
            )
            result = await session.execute(stmt)
            return list(result.scalars().all()), total

    async def get_template(self, template_id: str, user_id: str) -> Optional[ReportTemplate]:
        """获取单个模板（需有权限）"""
        async with get_db_session() as session:
            stmt = select(ReportTemplate).where(
                and_(
                    ReportTemplate.id == template_id,
                    or_(
                        ReportTemplate.created_by == user_id,
                        ReportTemplate.is_public == True,
                    ),
                )
            )
            result = await session.execute(stmt)
            return result.scalar_one_or_none()

    async def create_template(
        self,
        name: str,
        data_source: str,
        columns: List[Dict],
        user_id: str,
        description: Optional[str] = None,
        filters: Optional[Dict] = None,
        sort_by: Optional[List[Dict]] = None,
        default_format: str = ReportFormat.XLSX,
        is_public: bool = False,
        store_id: Optional[str] = None,
    ) -> ReportTemplate:
        """创建报表模板"""
        if data_source not in DATA_SOURCE_FIELDS:
            raise ValueError(f"不支持的数据源: {data_source}，可选: {list(DATA_SOURCE_FIELDS.keys())}")

        async with get_db_session() as session:
            template = ReportTemplate(
                id=uuid.uuid4(),
                name=name,
                description=description,
                data_source=data_source,
                columns=columns,
                filters=filters or {},
                sort_by=sort_by or [],
                default_format=default_format,
                is_public=is_public,
                created_by=user_id,
                store_id=store_id,
            )
            session.add(template)
            await session.commit()
            await session.refresh(template)
            logger.info("报表模板已创建", template_id=str(template.id), name=name)
            return template

    async def update_template(
        self,
        template_id: str,
        user_id: str,
        **kwargs,
    ) -> Optional[ReportTemplate]:
        """更新报表模板（只有创建者可以修改）"""
        async with get_db_session() as session:
            stmt = select(ReportTemplate).where(
                and_(ReportTemplate.id == template_id, ReportTemplate.created_by == user_id)
            )
            result = await session.execute(stmt)
            template = result.scalar_one_or_none()
            if not template:
                return None

            allowed_fields = {"name", "description", "columns", "filters", "sort_by", "default_format", "is_public"}
            for key, value in kwargs.items():
                if key in allowed_fields and value is not None:
                    setattr(template, key, value)

            await session.commit()
            await session.refresh(template)
            return template

    async def delete_template(self, template_id: str, user_id: str) -> bool:
        """删除报表模板（只有创建者可以删除）"""
        async with get_db_session() as session:
            stmt = select(ReportTemplate).where(
                and_(ReportTemplate.id == template_id, ReportTemplate.created_by == user_id)
            )
            result = await session.execute(stmt)
            template = result.scalar_one_or_none()
            if not template:
                return False
            await session.delete(template)
            await session.commit()
            return True

    # ------------------------------------------------------------------ #
    # 报表生成                                                              #
    # ------------------------------------------------------------------ #

    async def generate_report(
        self,
        template_id: str,
        user_id: str,
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        store_id: Optional[str] = None,
        fmt: Optional[str] = None,
    ) -> Tuple[bytes, str, str]:
        """
        按模板生成报表

        Returns:
            (文件内容, 文件名, media_type)
        """
        template = await self.get_template(template_id, user_id)
        if not template:
            raise ValueError("模板不存在或无权限访问")

        # 合并过滤条件（模板默认 + 请求参数）
        filters = dict(template.filters or {})
        if store_id:
            filters["store_id"] = store_id
        if start_date:
            filters["start_date"] = start_date.isoformat()
        if end_date:
            filters["end_date"] = end_date.isoformat()

        rows = await self._fetch_data(template.data_source, filters, template.sort_by or [])
        export_fmt = fmt or template.default_format

        if export_fmt == ReportFormat.CSV:
            content = self._to_csv(template.columns, rows)
            filename = f"{template.name}_{datetime.now().strftime('%Y%m%d')}.csv"
            media_type = "text/csv; charset=utf-8"
        elif export_fmt == ReportFormat.XLSX:
            content = self._to_xlsx(template.name, template.columns, rows)
            filename = f"{template.name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            raise ValueError(f"不支持的导出格式: {export_fmt}")

        return content, filename, media_type

    async def _fetch_data(
        self,
        data_source: str,
        filters: Dict[str, Any],
        sort_by: List[Dict],
    ) -> List[Dict[str, Any]]:
        """从数据库获取数据"""
        async with get_db_session() as session:
            if data_source == "transactions":
                return await self._fetch_transactions(session, filters, sort_by)
            elif data_source == "inventory":
                return await self._fetch_inventory(session, filters, sort_by)
            elif data_source == "orders":
                return await self._fetch_orders(session, filters, sort_by)
            elif data_source == "kpi":
                return await self._fetch_kpi(session, filters, sort_by)
            else:
                raise ValueError(f"不支持的数据源: {data_source}")

    async def _fetch_transactions(self, session, filters, sort_by):
        conditions = []
        if "store_id" in filters:
            conditions.append(FinancialTransaction.store_id == filters["store_id"])
        if "transaction_type" in filters:
            conditions.append(FinancialTransaction.transaction_type == filters["transaction_type"])
        if "category" in filters:
            conditions.append(FinancialTransaction.category == filters["category"])
        if "start_date" in filters:
            conditions.append(FinancialTransaction.transaction_date >= date.fromisoformat(filters["start_date"]))
        if "end_date" in filters:
            conditions.append(FinancialTransaction.transaction_date <= date.fromisoformat(filters["end_date"]))

        stmt = select(FinancialTransaction)
        if conditions:
            stmt = stmt.where(and_(*conditions))
        stmt = stmt.order_by(FinancialTransaction.transaction_date.desc()).limit(10000)

        result = await session.execute(stmt)
        rows = []
        for t in result.scalars().all():
            rows.append({
                "transaction_date": t.transaction_date.isoformat() if t.transaction_date else "",
                "transaction_type": t.transaction_type or "",
                "category": t.category or "",
                "subcategory": t.subcategory or "",
                "amount": round(t.amount / 100, 2) if t.amount else 0,
                "description": t.description or "",
                "payment_method": t.payment_method or "",
                "store_id": t.store_id or "",
            })
        return rows

    async def _fetch_inventory(self, session, filters, sort_by):
        conditions = []
        if "store_id" in filters:
            conditions.append(InventoryItem.store_id == filters["store_id"])
        if "category" in filters:
            conditions.append(InventoryItem.category == filters["category"])
        if "status" in filters:
            conditions.append(InventoryItem.status == filters["status"])

        stmt = select(InventoryItem)
        if conditions:
            stmt = stmt.where(and_(*conditions))
        stmt = stmt.limit(10000)

        result = await session.execute(stmt)
        rows = []
        for item in result.scalars().all():
            rows.append({
                "name": item.name or "",
                "category": item.category or "",
                "current_quantity": float(item.current_quantity or 0),
                "min_quantity": float(item.min_quantity or 0),
                "unit": item.unit or "",
                "unit_cost": round((item.unit_cost or 0) / 100, 2),
                "status": item.status.value if hasattr(item.status, "value") else str(item.status or ""),
                "store_id": item.store_id or "",
            })
        return rows

    async def _fetch_orders(self, session, filters, sort_by):
        conditions = []
        if "store_id" in filters:
            conditions.append(Order.store_id == filters["store_id"])
        if "status" in filters:
            conditions.append(Order.status == filters["status"])
        if "start_date" in filters:
            conditions.append(Order.created_at >= datetime.fromisoformat(filters["start_date"]))
        if "end_date" in filters:
            conditions.append(Order.created_at <= datetime.fromisoformat(filters["end_date"]))

        stmt = select(Order)
        if conditions:
            stmt = stmt.where(and_(*conditions))
        stmt = stmt.order_by(Order.created_at.desc()).limit(10000)

        result = await session.execute(stmt)
        rows = []
        for o in result.scalars().all():
            rows.append({
                "order_number": o.order_number or str(o.id),
                "status": o.status.value if hasattr(o.status, "value") else str(o.status or ""),
                "total_amount": round((o.total_amount or 0) / 100, 2),
                "table_number": o.table_number or "",
                "created_at": o.created_at.isoformat() if o.created_at else "",
                "store_id": o.store_id or "",
            })
        return rows

    async def _fetch_kpi(self, session, filters, sort_by):
        conditions = []
        if "store_id" in filters:
            conditions.append(KPIRecord.store_id == filters["store_id"])
        if "start_date" in filters:
            conditions.append(KPIRecord.record_date >= date.fromisoformat(filters["start_date"]))
        if "end_date" in filters:
            conditions.append(KPIRecord.record_date <= date.fromisoformat(filters["end_date"]))

        stmt = select(KPIRecord)
        if conditions:
            stmt = stmt.where(and_(*conditions))
        stmt = stmt.order_by(KPIRecord.record_date.desc()).limit(10000)

        result = await session.execute(stmt)
        rows = []
        for r in result.scalars().all():
            rows.append({
                "record_date": r.record_date.isoformat() if r.record_date else "",
                "value": float(r.value or 0),
                "target_value": float(r.target_value or 0),
                "achievement_rate": round(float(r.achievement_rate or 0) * 100, 2),
                "status": r.status or "",
                "store_id": r.store_id or "",
            })
        return rows

    def _to_csv(self, columns: List[Dict], rows: List[Dict]) -> bytes:
        """生成 CSV 字节流"""
        output = io.StringIO()
        writer = csv.writer(output)
        headers = [col.get("label", col["field"]) for col in columns]
        writer.writerow(headers)
        for row in rows:
            writer.writerow([row.get(col["field"], "") for col in columns])
        return output.getvalue().encode("utf-8-sig")

    def _to_xlsx(self, title: str, columns: List[Dict], rows: List[Dict]) -> bytes:
        """生成 Excel 字节流"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel 工作表名最长 31 字符

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

        for col_idx, col_def in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_def.get("label", col_def["field"]))
            cell.font = header_font
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_def in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_def["field"], ""))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # ------------------------------------------------------------------ #
    # 定时订阅 CRUD                                                        #
    # ------------------------------------------------------------------ #

    async def list_scheduled_reports(self, user_id: str) -> List[ScheduledReport]:
        """获取用户的定时报表订阅列表"""
        async with get_db_session() as session:
            stmt = select(ScheduledReport).where(ScheduledReport.user_id == user_id)
            result = await session.execute(stmt)
            return list(result.scalars().all())

    async def create_scheduled_report(
        self,
        template_id: str,
        user_id: str,
        frequency: str,
        run_at: str,
        channels: List[str],
        fmt: str = ReportFormat.XLSX,
        recipients: Optional[List[str]] = None,
        day_of_week: Optional[int] = None,
        day_of_month: Optional[int] = None,
    ) -> ScheduledReport:
        """创建定时报表订阅"""
        # 验证模板存在且有权限
        template = await self.get_template(template_id, user_id)
        if not template:
            raise ValueError("模板不存在或无权限访问")

        next_run = self._calc_next_run(frequency, run_at, day_of_week, day_of_month)

        async with get_db_session() as session:
            sr = ScheduledReport(
                id=uuid.uuid4(),
                template_id=template_id,
                user_id=user_id,
                frequency=frequency,
                run_at=run_at,
                day_of_week=day_of_week,
                day_of_month=day_of_month,
                channels=channels,
                recipients=recipients or [],
                format=fmt,
                is_active=True,
                next_run_at=next_run,
            )
            session.add(sr)
            await session.commit()
            await session.refresh(sr)
            logger.info("定时报表已创建", scheduled_id=str(sr.id), template_id=template_id)
            return sr

    async def update_scheduled_report(
        self,
        scheduled_id: str,
        user_id: str,
        **kwargs,
    ) -> Optional[ScheduledReport]:
        """更新定时报表订阅"""
        async with get_db_session() as session:
            stmt = select(ScheduledReport).where(
                and_(ScheduledReport.id == scheduled_id, ScheduledReport.user_id == user_id)
            )
            result = await session.execute(stmt)
            sr = result.scalar_one_or_none()
            if not sr:
                return None

            allowed = {"frequency", "run_at", "channels", "recipients", "format", "is_active",
                       "day_of_week", "day_of_month"}
            for key, value in kwargs.items():
                if key in allowed and value is not None:
                    setattr(sr, key, value)

            # 重新计算下次执行时间
            sr.next_run_at = self._calc_next_run(
                sr.frequency, sr.run_at, sr.day_of_week, sr.day_of_month
            )
            await session.commit()
            await session.refresh(sr)
            return sr

    async def delete_scheduled_report(self, scheduled_id: str, user_id: str) -> bool:
        """删除定时报表订阅"""
        async with get_db_session() as session:
            stmt = select(ScheduledReport).where(
                and_(ScheduledReport.id == scheduled_id, ScheduledReport.user_id == user_id)
            )
            result = await session.execute(stmt)
            sr = result.scalar_one_or_none()
            if not sr:
                return False
            await session.delete(sr)
            await session.commit()
            return True

    def _calc_next_run(
        self,
        frequency: str,
        run_at: str,
        day_of_week: Optional[int],
        day_of_month: Optional[int],
    ) -> str:
        """计算下次执行时间（UTC）"""
        now = datetime.utcnow()
        hour, minute = map(int, run_at.split(":"))

        if frequency == ScheduleFrequency.DAILY:
            next_run = now.replace(hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(days=1)
        elif frequency == ScheduleFrequency.WEEKLY:
            dow = day_of_week if day_of_week is not None else 0
            days_ahead = (dow - now.weekday()) % 7
            next_run = (now + timedelta(days=days_ahead)).replace(
                hour=hour, minute=minute, second=0, microsecond=0
            )
            if next_run <= now:
                next_run += timedelta(weeks=1)
        elif frequency == ScheduleFrequency.MONTHLY:
            dom = day_of_month if day_of_month is not None else 1
            next_run = now.replace(day=dom, hour=hour, minute=minute, second=0, microsecond=0)
            if next_run <= now:
                if now.month == 12:
                    next_run = next_run.replace(year=now.year + 1, month=1)
                else:
                    next_run = next_run.replace(month=now.month + 1)
        else:
            next_run = now + timedelta(days=1)

        return next_run.isoformat()


# 全局实例
custom_report_service = CustomReportService()

