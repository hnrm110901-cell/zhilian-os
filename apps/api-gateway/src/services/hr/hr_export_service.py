"""HRExportService — HR数据Excel导出"""
from datetime import date
from io import BytesIO
import uuid
import structlog
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from ...models.hr.payroll_batch import PayrollBatch
from ...models.hr.payroll_item import PayrollItem

logger = structlog.get_logger()


class HRExportService:

    async def export_payroll_batch(
        self,
        batch_id: uuid.UUID,
        session: AsyncSession,
    ) -> BytesIO:
        """导出薪资批次为Excel（3个sheet）"""
        # Get batch
        batch_result = await session.execute(
            select(PayrollBatch).where(PayrollBatch.id == batch_id)
        )
        batch = batch_result.scalar_one_or_none()
        if batch is None:
            raise ValueError(f"PayrollBatch {batch_id} not found")

        # Get items
        items_result = await session.execute(
            select(PayrollItem).where(PayrollItem.batch_id == batch_id)
        )
        items = list(items_result.scalars().all())

        wb = Workbook()

        # Sheet 1: 月度汇总
        ws_summary = wb.active
        ws_summary.title = "月度汇总"
        ws_summary.append(["薪资核算期间", f"{batch.period_year}年{batch.period_month}月"])
        ws_summary.append(["员工人数", len(items)])
        ws_summary.append(["税前总额（元）", batch.total_gross_fen / 100])
        ws_summary.append(["税后总额（元）", batch.total_net_fen / 100])
        ws_summary.append([])

        # Sheet 2: 个人工资条
        ws_detail = wb.create_sheet("个人工资条")
        ws_detail.append([
            "员工ID", "基本工资", "加班费", "迟到扣款", "缺勤扣款",
            "税前合计", "社保", "个税", "实发工资",
        ])
        for item in items:
            ws_detail.append([
                str(item.assignment_id)[:8],
                item.base_salary_fen / 100,
                item.overtime_fen / 100,
                item.deduction_late_fen / 100,
                item.deduction_absent_fen / 100,
                item.gross_fen / 100,
                item.social_insurance_fen / 100,
                item.tax_fen / 100,
                item.net_fen / 100,
            ])

        # Sheet 3: 部门成本
        ws_cost = wb.create_sheet("部门成本")
        ws_cost.append(["部门", "总成本（元）"])
        ws_cost.append([batch.org_node_id, batch.total_gross_fen / 100])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    async def export_attendance_monthly(
        self,
        org_node_id: str,
        year: int,
        month: int,
        session: AsyncSession,
    ) -> BytesIO:
        """导出月度考勤报表Excel"""
        from ...models.hr.daily_attendance import DailyAttendance
        from ...models.hr.employment_assignment import EmploymentAssignment
        from ...models.hr.person import Person

        first_day = date(year, month, 1)
        last_day = (
            date(year, month + 1, 1) if month < 12 else date(year + 1, 1, 1)
        )

        # 获取该组织节点下所有在岗员工
        result = await session.execute(
            select(EmploymentAssignment, Person)
            .join(Person, Person.id == EmploymentAssignment.person_id)
            .where(
                EmploymentAssignment.org_node_id == org_node_id,
                EmploymentAssignment.status == "active",
            )
        )
        rows = result.all()

        wb = Workbook()
        ws = wb.active
        ws.title = f"考勤报表{year}年{month}月"
        ws.append([
            "姓名", "部门", "出勤天数", "正常天数", "迟到次数",
            "早退次数", "缺勤天数", "加班小时", "总工时",
        ])

        for assignment, person in rows:
            att_result = await session.execute(
                select(DailyAttendance).where(
                    DailyAttendance.assignment_id == assignment.id,
                    DailyAttendance.date >= first_day,
                    DailyAttendance.date < last_day,
                )
            )
            att = list(att_result.scalars().all())

            total_days = len(att)
            normal = sum(1 for a in att if a.status == "normal")
            late = sum(1 for a in att if a.status == "late")
            early = sum(1 for a in att if a.status == "early_leave")
            absent = sum(1 for a in att if a.status == "absent")
            overtime_h = round(
                sum(a.overtime_minutes for a in att) / 60, 1,
            )
            work_h = round(sum(a.work_minutes for a in att) / 60, 1)

            ws.append([
                person.name, org_node_id, total_days, normal,
                late, early, absent, overtime_h, work_h,
            ])

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
