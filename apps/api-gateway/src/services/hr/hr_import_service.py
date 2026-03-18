"""HRImportService — 员工Excel批量导入"""
import uuid
from datetime import date, datetime
from io import BytesIO
from typing import Optional

import structlog
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import load_workbook

from ...models.hr.person import Person
from ...models.hr.employment_assignment import EmploymentAssignment
from ...models.hr.employment_contract import EmploymentContract

logger = structlog.get_logger()

# 必须列
_REQUIRED_COLUMNS = ["姓名", "手机", "岗位", "入职日期", "薪酬类型", "基本工资"]

# 薪酬类型映射
_PAY_TYPE_MAP = {
    "固定月薪": "fixed_monthly",
    "时薪": "hourly",
    "底薪+提成": "base_plus_commission",
    "计件": "piecework",
}


class HRImportService:

    async def import_employee_roster(
        self,
        file_content: bytes,
        org_node_id: str,
        created_by: str,
        session: AsyncSession,
    ) -> dict:
        """从Excel导入员工花名册"""
        wb = load_workbook(BytesIO(file_content), read_only=True)
        ws = wb.active

        # 读表头
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

        # 校验必须列
        missing = [c for c in _REQUIRED_COLUMNS if c not in headers]
        if missing:
            wb.close()
            return {"imported": 0, "skipped": 0, "errors": [f"缺少必须列: {', '.join(missing)}"]}

        col_map = {h: i for i, h in enumerate(headers)}
        imported = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name = row[col_map["姓名"]]
                phone = str(row[col_map["手机"]])
                position = row[col_map["岗位"]]
                start_str = row[col_map["入职日期"]]
                pay_type_cn = row[col_map["薪酬类型"]]
                base_salary = row[col_map["基本工资"]]

                if not name or not phone:
                    skipped += 1
                    continue

                # 检查手机号重复
                existing = await session.execute(
                    select(Person.id).where(Person.phone == phone)
                )
                if existing.scalar_one_or_none():
                    skipped += 1
                    continue

                # 解析日期
                if isinstance(start_str, datetime):
                    start_date = start_str.date()
                elif isinstance(start_str, date):
                    start_date = start_str
                else:
                    start_date = date.fromisoformat(str(start_str)[:10])

                # 解析薪酬
                pay_type = _PAY_TYPE_MAP.get(str(pay_type_cn), "fixed_monthly")
                base_fen = int(float(base_salary) * 100)
                pay_scheme = {"type": pay_type, "base_salary_fen": base_fen}

                id_number = row[col_map["身份证号"]] if "身份证号" in col_map else None
                gender = row[col_map["性别"]] if "性别" in col_map else None

                # 创建Person
                person = Person(
                    name=name,
                    phone=phone,
                    id_number=str(id_number) if id_number else None,
                )
                session.add(person)
                await session.flush()

                # 创建Assignment
                assignment = EmploymentAssignment(
                    person_id=person.id,
                    org_node_id=org_node_id,
                    employment_type="full_time",
                    start_date=start_date,
                    status="active",
                )
                session.add(assignment)
                await session.flush()

                # 创建Contract
                contract = EmploymentContract(
                    assignment_id=assignment.id,
                    contract_type="labor",
                    pay_scheme=pay_scheme,
                    valid_from=start_date,
                )
                session.add(contract)

                imported += 1

            except Exception as exc:
                errors.append(f"第{row_idx}行: {str(exc)}")

        await session.flush()
        wb.close()

        logger.info("hr_import.completed", imported=imported, skipped=skipped, error_count=len(errors))
        return {"imported": imported, "skipped": skipped, "errors": errors}
