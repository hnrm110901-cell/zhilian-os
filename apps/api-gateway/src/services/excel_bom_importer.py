"""
Excel BOM 导入器（徐记海鲜 POC 配方批量录入）

支持的 Excel 格式：
  Sheet: "配方总表"（或第一个 Sheet）

  必须列（不区分大小写，支持别名）：
    菜品编码  dish_code
    菜品名称  dish_name
    版本      version       （缺省填 v1）
    食材名称  ingredient_name
    标准用量  standard_qty
    单位      unit
    出成率    yield_rate    （缺省 1.0，可选列）
    食材分类  ingredient_category  （可选列）
    核心食材  is_key             （Y/N，可选列）

导入流程：
  1. 读取 Excel → 按 (dish_code, version) 分组
  2. 查找/创建 InventoryItem（食材主档）
  3. 查找/创建 Dish 主档
  4. 调用 BOMService 创建 BOMTemplate + BOMItems
  5. 触发 Neo4j 同步
  6. 返回导入报告

依赖：
  pip install openpyxl  （已在 requirements.txt）
"""

import io
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import structlog
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.dish import Dish
from src.models.inventory import InventoryItem, InventoryStatus
from src.services.bom_service import BOMService

logger = structlog.get_logger()


# ── 列名别名映射 ──────────────────────────────────────────────────────────────

COLUMN_ALIASES: Dict[str, str] = {
    # 菜品
    "菜品编码": "dish_code", "dish_code": "dish_code",
    "菜品名称": "dish_name", "菜品": "dish_name", "dish_name": "dish_name",
    "版本": "version", "配方版本": "version", "version": "version",
    "出成率": "yield_rate", "yield_rate": "yield_rate",
    "标准份重": "standard_portion", "份重": "standard_portion",
    "制作工时": "prep_time_minutes", "工时": "prep_time_minutes",
    "备注": "notes", "notes": "notes",
    # 食材
    "食材名称": "ingredient_name", "食材": "ingredient_name",
    "食材编码": "ingredient_code",
    "食材分类": "ingredient_category", "分类": "ingredient_category",
    "标准用量": "standard_qty", "用量": "standard_qty", "qty": "standard_qty",
    "单位": "unit",
    "毛料用量": "raw_qty",
    "核心食材": "is_key", "关键食材": "is_key",
    "可选": "is_optional",
    "加工说明": "prep_notes",
}


@dataclass
class ImportRow:
    dish_code: str
    dish_name: str
    version: str
    ingredient_name: str
    ingredient_code: str
    ingredient_category: str
    standard_qty: float
    unit: str
    raw_qty: Optional[float]
    yield_rate: float
    is_key: bool
    is_optional: bool
    prep_notes: Optional[str]
    standard_portion: Optional[float]
    prep_time_minutes: Optional[int]
    notes: Optional[str]
    row_number: int


@dataclass
class ImportReport:
    total_rows: int = 0
    total_dishes: int = 0
    total_boms_created: int = 0
    total_items_created: int = 0
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    dish_results: List[dict] = field(default_factory=list)


class ExcelBOMImporter:
    """
    Excel BOM 批量导入器

    用法::

        async with get_db() as db:
            importer = ExcelBOMImporter(db, store_id="XJ-CHANGSHA-001")
            report = await importer.import_from_bytes(excel_bytes)
    """

    def __init__(self, db: AsyncSession, store_id: str, created_by: str = "excel_import"):
        self.db = db
        self.store_id = store_id
        self.created_by = created_by
        self.bom_svc = BOMService(db)

    async def import_from_bytes(self, excel_bytes: bytes) -> ImportReport:
        """从 Excel 字节流导入 BOM"""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("缺少依赖：pip install openpyxl")

        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)

        # 优先使用名为 "配方总表" 的 Sheet，否则取第一个
        sheet_name = "配方总表" if "配方总表" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet_name]

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Excel 内容为空")

        # 解析表头
        header_row = rows[0]
        col_map = self._parse_header(header_row)

        # 校验必需列
        required = {"dish_code", "dish_name", "ingredient_name", "standard_qty", "unit"}
        missing = required - set(col_map.keys())
        if missing:
            raise ValueError(f"Excel 缺少必需列：{', '.join(missing)}")

        # 解析数据行
        import_rows = []
        for i, row in enumerate(rows[1:], start=2):
            parsed = self._parse_row(row, col_map, i)
            if parsed:
                import_rows.append(parsed)

        report = ImportReport(total_rows=len(import_rows))
        await self._process_rows(import_rows, report)
        return report

    # ── 内部方法 ──────────────────────────────────────────────────────────────

    def _parse_header(self, header_row: tuple) -> Dict[str, int]:
        """解析表头行，返回 {规范列名: 列索引}"""
        col_map: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if cell is None:
                continue
            key = str(cell).strip()
            canonical = COLUMN_ALIASES.get(key, key.lower())
            col_map[canonical] = idx
        return col_map

    def _parse_row(self, row: tuple, col_map: Dict[str, int], row_num: int) -> Optional[ImportRow]:
        """解析单行数据，跳过空行"""
        def get(col: str):
            idx = col_map.get(col)
            return row[idx] if idx is not None and idx < len(row) else None

        dish_code = str(get("dish_code") or "").strip()
        dish_name = str(get("dish_name") or "").strip()
        ingredient_name = str(get("ingredient_name") or "").strip()

        if not dish_code or not ingredient_name:
            return None

        try:
            standard_qty = float(get("standard_qty") or 0)
        except (TypeError, ValueError):
            standard_qty = 0.0

        unit = str(get("unit") or "克").strip()
        version = str(get("version") or "v1").strip()

        try:
            yield_rate = float(get("yield_rate") or 1.0)
        except (TypeError, ValueError):
            yield_rate = 1.0

        raw_qty_val = get("raw_qty")
        raw_qty = float(raw_qty_val) if raw_qty_val is not None else None

        is_key_val = str(get("is_key") or "").strip().upper()
        is_key = is_key_val in ("Y", "YES", "是", "√", "1", "TRUE")

        is_opt_val = str(get("is_optional") or "").strip().upper()
        is_optional = is_opt_val in ("Y", "YES", "是", "√", "1", "TRUE")

        std_portion = get("standard_portion")
        prep_time = get("prep_time_minutes")

        return ImportRow(
            dish_code=dish_code,
            dish_name=dish_name,
            version=version,
            ingredient_name=ingredient_name,
            ingredient_code=str(get("ingredient_code") or ingredient_name).strip(),
            ingredient_category=str(get("ingredient_category") or "食材").strip(),
            standard_qty=standard_qty,
            unit=unit,
            raw_qty=raw_qty,
            yield_rate=yield_rate,
            is_key=is_key,
            is_optional=is_optional,
            prep_notes=str(get("prep_notes") or "").strip() or None,
            standard_portion=float(std_portion) if std_portion else None,
            prep_time_minutes=int(prep_time) if prep_time else None,
            notes=str(get("notes") or "").strip() or None,
            row_number=row_num,
        )

    async def _process_rows(self, rows: List[ImportRow], report: ImportReport) -> None:
        """按 (dish_code, version) 分组，创建 BOM"""
        # 按 (dish_code, version) 分组
        groups: Dict[Tuple[str, str], List[ImportRow]] = defaultdict(list)
        for r in rows:
            groups[(r.dish_code, r.version)].append(r)

        report.total_dishes = len({r.dish_code for r in rows})

        for (dish_code, version), group_rows in groups.items():
            first = group_rows[0]
            try:
                # 1. 确保 Dish 主档存在
                dish = await self._ensure_dish(dish_code, first.dish_name)

                # 2. 创建 BOM 版本
                bom = await self.bom_svc.create_bom(
                    store_id=self.store_id,
                    dish_id=str(dish.id),
                    version=version,
                    yield_rate=first.yield_rate,
                    standard_portion=first.standard_portion,
                    prep_time_minutes=first.prep_time_minutes,
                    notes=first.notes,
                    created_by=self.created_by,
                    activate=True,
                )
                report.total_boms_created += 1

                # 3. 逐行添加食材
                for ir in group_rows:
                    try:
                        ing = await self._ensure_ingredient(
                            code=ir.ingredient_code,
                            name=ir.ingredient_name,
                            category=ir.ingredient_category,
                            unit=ir.unit,
                        )
                        await self.bom_svc.add_bom_item(
                            bom_id=str(bom.id),
                            ingredient_id=ing.id,
                            standard_qty=ir.standard_qty,
                            unit=ir.unit,
                            raw_qty=ir.raw_qty,
                            waste_factor=0.0,
                            is_key_ingredient=ir.is_key,
                            is_optional=ir.is_optional,
                            prep_notes=ir.prep_notes,
                        )
                        report.total_items_created += 1
                    except Exception as e:
                        report.warnings.append(
                            f"行 {ir.row_number} 食材添加失败({ir.ingredient_name}): {e}"
                        )

                # 4. Neo4j 同步
                await self.bom_svc.sync_to_neo4j(bom)

                report.dish_results.append({
                    "dish_code": dish_code,
                    "dish_name": first.dish_name,
                    "version": version,
                    "bom_id": str(bom.id),
                    "items_count": report.total_items_created,
                    "status": "success",
                })

            except Exception as e:
                report.errors.append(f"菜品 {dish_code}({version}) 导入失败: {e}")
                report.dish_results.append({
                    "dish_code": dish_code,
                    "dish_name": first.dish_name,
                    "version": version,
                    "status": "error",
                    "error": str(e),
                })

        await self.db.commit()

    async def _ensure_dish(self, code: str, name: str) -> Dish:
        """查找或创建菜品主档"""
        stmt = select(Dish).where(Dish.code == code)
        result = await self.db.execute(stmt)
        dish = result.scalar_one_or_none()
        if dish:
            return dish

        import uuid as _uuid
        from decimal import Decimal
        dish = Dish(
            id=_uuid.uuid4(),
            store_id=self.store_id,
            code=code,
            name=name,
            price=Decimal("0.00"),  # 占位价格，后续由 POS 同步更新
        )
        self.db.add(dish)
        await self.db.flush()
        logger.info("创建菜品主档（Excel 导入）", code=code, name=name)
        return dish

    async def _ensure_ingredient(
        self,
        code: str,
        name: str,
        category: str,
        unit: str,
    ) -> InventoryItem:
        """查找或创建食材主档（InventoryItem）"""
        stmt = select(InventoryItem).where(InventoryItem.id == code)
        result = await self.db.execute(stmt)
        ing = result.scalar_one_or_none()
        if ing:
            return ing

        ing = InventoryItem(
            id=code,
            store_id=self.store_id,
            name=name,
            category=category,
            unit=unit,
            current_quantity=0.0,
            min_quantity=0.0,
            status=InventoryStatus.NORMAL,
        )
        self.db.add(ing)
        await self.db.flush()
        logger.info("创建食材主档（Excel 导入）", code=code, name=name)
        return ing
