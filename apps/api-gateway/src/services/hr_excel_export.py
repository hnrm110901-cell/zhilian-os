"""
HR Excel 导出服务 — 生成月度人事报表 Excel 文件
支持7张月报表 + 工资明细 + 考勤报表 + 花名册
所有金额：数据库存分（fen），Excel 展示元（÷100）
"""
from io import BytesIO
from typing import Any, Dict, List, Optional
from datetime import date, timedelta
import calendar
import structlog

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

from sqlalchemy import select, and_, func
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.employee import Employee
from src.models.payroll import PayrollRecord, SalaryStructure
from src.models.employee_lifecycle import EmployeeChange, ChangeType
from src.models.exit_interview import ExitInterview
from src.models.mentorship import Mentorship
from src.models.social_insurance import EmployeeSocialInsurance
from src.models.attendance import AttendanceLog
from src.models.store import Store

logger = structlog.get_logger()

# ── 样式常量 ───────────────────────────────────────────────
HEADER_FONT = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="0AAF9A", end_color="0AAF9A", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14)
NORMAL_FONT = Font(name="微软雅黑", size=10)
MONEY_FORMAT = '#,##0.00'
PCT_FORMAT = '0.0%'
DATE_FORMAT = 'YYYY-MM-DD'
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _fen_to_yuan(fen: Optional[int]) -> float:
    """分 → 元"""
    return (fen or 0) / 100


def _apply_header_style(ws, row: int, col_count: int):
    """给表头行应用样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_data_style(ws, row: int, col_count: int):
    """给数据行应用样式"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER


def _auto_width(ws, min_width: int = 10, max_width: int = 40):
    """自动列宽"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                # 中文字符按2个宽度计算
                val = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        adjusted = max(min(max_len + 2, max_width), min_width)
        ws.column_dimensions[col_letter].width = adjusted


def _write_title_row(ws, title: str, col_count: int):
    """写标题行并合并单元格"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")


async def _get_store_name(db: AsyncSession, store_id: str) -> str:
    """获取门店名称"""
    result = await db.execute(
        select(Store.name).where(Store.id == store_id)
    )
    name = result.scalar()
    return name or store_id


class HRExcelExporter:
    """生成 HR 月度报表 Excel 文件"""

    async def export_monthly_report(
        self, db: AsyncSession, store_id: str, month: str, brand_id: str
    ) -> bytes:
        """
        生成完整月度报表 Excel（7张 Sheet）
        month 格式: YYYY-MM
        返回 bytes 供 StreamingResponse 下载
        """
        store_name = await _get_store_name(db, store_id)
        year, mon = int(month[:4]), int(month[5:7])
        month_start = date(year, mon, 1)
        month_end = date(year, mon, calendar.monthrange(year, mon)[1])
        title_prefix = f"{store_name} {month}"

        wb = Workbook()
        # 删除默认 Sheet
        wb.remove(wb.active)

        # Sheet 1: 工资异动表
        await self._sheet_salary_changes(
            wb, db, store_id, month, month_start, month_end, title_prefix
        )
        # Sheet 2: 月末编制盘存
        await self._sheet_headcount_inventory(
            wb, db, store_id, month_end, title_prefix
        )
        # Sheet 3: 核心岗位培养统计
        await self._sheet_mentorship(
            wb, db, store_id, month_start, month_end, title_prefix
        )
        # Sheet 4: 小时工/灵活用工考勤
        await self._sheet_hourly_attendance(
            wb, db, store_id, month_start, month_end, title_prefix
        )
        # Sheet 5: 离职回访汇总
        await self._sheet_exit_interview(
            wb, db, store_id, month_start, month_end, title_prefix
        )
        # Sheet 6: 人事工作总结与计划
        await self._sheet_hr_summary(
            wb, db, store_id, brand_id, month, title_prefix
        )
        # Sheet 7: 社保/保险变动
        await self._sheet_insurance_changes(
            wb, db, store_id, year, month_start, month_end, title_prefix
        )

        return self._save_to_bytes(wb)

    async def export_payroll_detail(
        self, db: AsyncSession, store_id: str, month: str
    ) -> bytes:
        """导出工资明细表"""
        store_name = await _get_store_name(db, store_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "工资明细"

        headers = [
            "姓名", "工号", "岗位", "基本工资", "岗位津贴", "餐补",
            "交通补贴", "绩效奖金", "加班费", "提成", "奖励",
            "迟到扣款", "缺勤扣款", "社保个人", "公积金个人", "个税",
            "应发合计", "实发合计",
        ]
        money_cols = set(range(4, 19))  # 列4~18是金额列（1-based）

        _write_title_row(ws, f"{store_name} {month} 工资明细表", len(headers))

        # 表头（第2行）
        for col, h in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=h)
        _apply_header_style(ws, 2, len(headers))

        # 数据
        result = await db.execute(
            select(PayrollRecord, Employee.name, Employee.position).join(
                Employee, PayrollRecord.employee_id == Employee.id
            ).where(
                and_(
                    PayrollRecord.store_id == store_id,
                    PayrollRecord.pay_month == month,
                )
            ).order_by(Employee.name)
        )
        rows = result.all()

        for i, (pr, emp_name, position) in enumerate(rows, start=3):
            data = [
                emp_name,
                pr.employee_id,
                position or "",
                _fen_to_yuan(pr.base_salary_fen),
                _fen_to_yuan(pr.position_allowance_fen),
                _fen_to_yuan(pr.meal_allowance_fen),
                _fen_to_yuan(pr.transport_allowance_fen),
                _fen_to_yuan(pr.performance_bonus_fen),
                _fen_to_yuan(pr.overtime_pay_fen),
                _fen_to_yuan(pr.commission_fen),
                _fen_to_yuan(pr.reward_fen),
                _fen_to_yuan(pr.late_deduction_fen),
                _fen_to_yuan(pr.absence_deduction_fen),
                _fen_to_yuan(pr.social_insurance_fen),
                _fen_to_yuan(pr.housing_fund_fen),
                _fen_to_yuan(pr.tax_fen),
                _fen_to_yuan(pr.gross_salary_fen),
                _fen_to_yuan(pr.net_salary_fen),
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=i, column=col, value=val)
                if col in money_cols:
                    cell.number_format = MONEY_FORMAT
            _apply_data_style(ws, i, len(headers))

        # 合计行
        if rows:
            total_row = len(rows) + 3
            ws.cell(row=total_row, column=1, value="合计")
            ws.cell(row=total_row, column=1).font = Font(name="微软雅黑", bold=True, size=10)
            for col in money_cols:
                col_letter = get_column_letter(col)
                formula = f"=SUM({col_letter}3:{col_letter}{total_row - 1})"
                cell = ws.cell(row=total_row, column=col, value=formula)
                cell.number_format = MONEY_FORMAT
                cell.font = Font(name="微软雅黑", bold=True, size=10)

        ws.freeze_panes = "A3"
        _auto_width(ws)
        return self._save_to_bytes(wb)

    async def export_attendance_report(
        self, db: AsyncSession, store_id: str, month: str
    ) -> bytes:
        """导出月度考勤报表"""
        store_name = await _get_store_name(db, store_id)
        year, mon = int(month[:4]), int(month[5:7])
        month_start = date(year, mon, 1)
        month_end = date(year, mon, calendar.monthrange(year, mon)[1])
        workdays = sum(
            1 for d in range(calendar.monthrange(year, mon)[1])
            if date(year, mon, d + 1).weekday() < 5
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "考勤报表"

        headers = [
            "姓名", "工号", "应出勤天数", "实际出勤", "迟到次数",
            "早退次数", "请假天数", "缺勤天数", "加班时数", "出勤率",
        ]

        _write_title_row(ws, f"{store_name} {month} 考勤报表", len(headers))

        for col, h in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=h)
        _apply_header_style(ws, 2, len(headers))

        # 查询在职员工
        emp_result = await db.execute(
            select(Employee).where(
                and_(
                    Employee.store_id == store_id,
                    Employee.is_active.is_(True),
                )
            ).order_by(Employee.name)
        )
        employees = emp_result.scalars().all()

        row_idx = 3
        for emp in employees:
            # 查询考勤汇总
            att_result = await db.execute(
                select(
                    func.count(AttendanceLog.id).label("total"),
                    func.sum(
                        func.cast(
                            AttendanceLog.status == "normal", Integer
                        ) + func.cast(
                            AttendanceLog.status == "late", Integer
                        )
                    ).label("attended"),
                    func.sum(func.cast(AttendanceLog.status == "late", Integer)).label("late_count"),
                    func.sum(func.cast(AttendanceLog.status == "early_leave", Integer)).label("early_count"),
                    func.sum(func.cast(AttendanceLog.status == "leave", Integer)).label("leave_count"),
                    func.sum(func.cast(AttendanceLog.status == "absent", Integer)).label("absent_count"),
                    func.coalesce(func.sum(AttendanceLog.overtime_hours), 0).label("ot_hours"),
                ).where(
                    and_(
                        AttendanceLog.employee_id == emp.id,
                        AttendanceLog.store_id == store_id,
                        AttendanceLog.work_date >= month_start,
                        AttendanceLog.work_date <= month_end,
                    )
                )
            )
            stats = att_result.one()

            attended = int(stats.attended or 0)
            late_count = int(stats.late_count or 0)
            early_count = int(stats.early_count or 0)
            leave_count = int(stats.leave_count or 0)
            absent_count = int(stats.absent_count or 0)
            ot_hours = float(stats.ot_hours or 0)
            attendance_rate = attended / workdays if workdays > 0 else 0

            data = [
                emp.name,
                emp.id,
                workdays,
                attended,
                late_count,
                early_count,
                leave_count,
                absent_count,
                round(ot_hours, 1),
                attendance_rate,
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if col == 10:
                    cell.number_format = PCT_FORMAT
            _apply_data_style(ws, row_idx, len(headers))
            row_idx += 1

        ws.freeze_panes = "A3"
        _auto_width(ws)
        return self._save_to_bytes(wb)

    async def export_roster(
        self, db: AsyncSession, store_id: str
    ) -> bytes:
        """导出员工花名册"""
        store_name = await _get_store_name(db, store_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "花名册"

        headers = [
            "工号", "姓名", "性别", "岗位", "职级", "用工类型", "员工状态",
            "入职日期", "转正日期", "司龄(月)", "手机号", "邮箱",
            "出生日期", "民族", "学历", "毕业院校", "专业",
            "户籍类型", "户籍地", "政治面貌",
            "紧急联系人", "紧急联系电话", "紧急联系关系",
            "开户行", "银行支行",
            "健康证到期", "身份证到期", "背调状态",
            "工时制度", "宿舍", "工会会员", "专业证书",
        ]

        _write_title_row(ws, f"{store_name} 员工花名册", len(headers))

        for col, h in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=h)
        _apply_header_style(ws, 2, len(headers))

        result = await db.execute(
            select(Employee).where(
                Employee.store_id == store_id,
            ).order_by(Employee.is_active.desc(), Employee.name)
        )
        employees = result.scalars().all()

        emp_type_map = {
            "regular": "正式", "part_time": "兼职", "intern": "实习",
            "trainee": "培训", "rehire": "返聘", "temp": "临时",
            "outsource": "外包", "outsource_flex": "灵活用工",
        }
        status_map = {
            "trial": "试岗", "probation": "试用", "regular": "正式", "resigned": "离职",
        }

        for i, emp in enumerate(employees, start=3):
            data = [
                emp.id,
                emp.name,
                emp.gender or "",
                emp.position or "",
                emp.grade_level or "",
                emp_type_map.get(emp.employment_type, emp.employment_type or ""),
                status_map.get(emp.employment_status, emp.employment_status or ""),
                str(emp.hire_date) if emp.hire_date else "",
                str(emp.regular_date) if emp.regular_date else "",
                emp.seniority_months or "",
                emp.phone or "",
                emp.email or "",
                str(emp.birth_date) if emp.birth_date else "",
                emp.ethnicity or "",
                emp.education or "",
                emp.graduation_school or "",
                emp.major or "",
                emp.hukou_type or "",
                emp.hukou_location or "",
                emp.political_status or "",
                emp.emergency_contact or "",
                emp.emergency_phone or "",
                emp.emergency_relation or "",
                emp.bank_name or "",
                emp.bank_branch or "",
                str(emp.health_cert_expiry) if emp.health_cert_expiry else "",
                str(emp.id_card_expiry) if emp.id_card_expiry else "",
                emp.background_check or "",
                emp.work_hour_type or "",
                emp.accommodation or "",
                "是" if emp.union_member else "否",
                emp.professional_cert or "",
            ]
            for col, val in enumerate(data, 1):
                ws.cell(row=i, column=col, value=val)
            _apply_data_style(ws, i, len(headers))

        ws.freeze_panes = "A3"
        _auto_width(ws)
        return self._save_to_bytes(wb)

    # ── 内部方法：7 张月报 Sheet ────────────────────────────────

    async def _sheet_salary_changes(
        self, wb: Workbook, db: AsyncSession, store_id: str,
        month: str, month_start: date, month_end: date, title_prefix: str,
    ):
        """Sheet 1: 工资异动表"""
        ws = wb.create_sheet("工资异动表")

        # 新进员工
        new_result = await db.execute(
            select(Employee).where(
                and_(
                    Employee.store_id == store_id,
                    Employee.hire_date >= month_start,
                    Employee.hire_date <= month_end,
                )
            )
        )
        new_employees = new_result.scalars().all()

        # 离职
        resign_result = await db.execute(
            select(EmployeeChange, Employee.name, Employee.position).join(
                Employee, EmployeeChange.employee_id == Employee.id
            ).where(
                and_(
                    EmployeeChange.store_id == store_id,
                    EmployeeChange.change_type == ChangeType.RESIGN,
                    EmployeeChange.effective_date >= month_start,
                    EmployeeChange.effective_date <= month_end,
                )
            )
        )
        resignations = resign_result.all()

        # 调薪
        adj_result = await db.execute(
            select(SalaryStructure, Employee.name, Employee.position).join(
                Employee, SalaryStructure.employee_id == Employee.id
            ).where(
                and_(
                    SalaryStructure.store_id == store_id,
                    SalaryStructure.effective_date >= month_start,
                    SalaryStructure.effective_date <= month_end,
                    SalaryStructure.is_active.is_(True),
                )
            )
        )
        adjustments = adj_result.all()

        # 标题
        _write_title_row(ws, f"{title_prefix} 工资异动表", 5)

        # 新进
        row = 3
        ws.cell(row=row, column=1, value="一、新入职员工")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
        sub_headers = ["姓名", "工号", "岗位", "入职日期"]
        for col, h in enumerate(sub_headers, 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, len(sub_headers))
        row += 1
        for emp in new_employees:
            data = [emp.name, emp.id, emp.position or "", str(emp.hire_date) if emp.hire_date else ""]
            for col, val in enumerate(data, 1):
                ws.cell(row=row, column=col, value=val)
            _apply_data_style(ws, row, len(sub_headers))
            row += 1
        if not new_employees:
            ws.cell(row=row, column=1, value="本月无新入职")
            row += 1

        # 离职
        row += 1
        ws.cell(row=row, column=1, value="二、离职员工")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
        sub_headers = ["姓名", "工号", "岗位", "离职日期"]
        for col, h in enumerate(sub_headers, 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, len(sub_headers))
        row += 1
        for change, name, position in resignations:
            data = [name, change.employee_id, position or "", str(change.effective_date)]
            for col, val in enumerate(data, 1):
                ws.cell(row=row, column=col, value=val)
            _apply_data_style(ws, row, len(sub_headers))
            row += 1
        if not resignations:
            ws.cell(row=row, column=1, value="本月无离职")
            row += 1

        # 调薪
        row += 1
        ws.cell(row=row, column=1, value="三、薪资调整")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
        sub_headers = ["姓名", "工号", "岗位", "调整后基本工资(元)", "生效日期"]
        for col, h in enumerate(sub_headers, 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, len(sub_headers))
        row += 1
        for ss, name, position in adjustments:
            data = [
                name, ss.employee_id, position or "",
                _fen_to_yuan(ss.base_salary_fen), str(ss.effective_date),
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if col == 4:
                    cell.number_format = MONEY_FORMAT
            _apply_data_style(ws, row, len(sub_headers))
            row += 1
        if not adjustments:
            ws.cell(row=row, column=1, value="本月无调薪")
            row += 1

        # 汇总
        row += 1
        ws.cell(row=row, column=1, value="汇总")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value=f"新入职: {len(new_employees)}人")
        ws.cell(row=row, column=3, value=f"离职: {len(resignations)}人")
        ws.cell(row=row, column=5, value=f"调薪: {len(adjustments)}人")

        _auto_width(ws)

    async def _sheet_headcount_inventory(
        self, wb: Workbook, db: AsyncSession, store_id: str,
        month_end: date, title_prefix: str,
    ):
        """Sheet 2: 月末编制盘存"""
        ws = wb.create_sheet("月末编制盘存")

        result = await db.execute(
            select(
                Employee.position,
                Employee.employment_type,
                func.count(Employee.id).label("count"),
            ).where(
                and_(
                    Employee.store_id == store_id,
                    Employee.is_active.is_(True),
                )
            ).group_by(Employee.position, Employee.employment_type)
        )
        rows = result.all()

        _write_title_row(ws, f"{title_prefix} 月末编制盘存", 4)

        headers = ["岗位", "用工类型", "在编人数", "统计日期"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=h)
        _apply_header_style(ws, 2, len(headers))

        emp_type_map = {
            "regular": "正式", "part_time": "兼职", "intern": "实习",
            "temp": "临时", "outsource_flex": "灵活用工",
        }

        total = 0
        for i, (position, emp_type, count) in enumerate(rows, start=3):
            data = [
                position or "未设置",
                emp_type_map.get(emp_type, emp_type or "正式"),
                count,
                str(month_end),
            ]
            for col, val in enumerate(data, 1):
                ws.cell(row=i, column=col, value=val)
            _apply_data_style(ws, i, len(headers))
            total += count

        # 合计行
        total_row = len(rows) + 3
        ws.cell(row=total_row, column=1, value="合计")
        ws.cell(row=total_row, column=1).font = Font(name="微软雅黑", bold=True, size=10)
        ws.cell(row=total_row, column=3, value=total)
        ws.cell(row=total_row, column=3).font = Font(name="微软雅黑", bold=True, size=10)

        ws.freeze_panes = "A3"
        _auto_width(ws)

    async def _sheet_mentorship(
        self, wb: Workbook, db: AsyncSession, store_id: str,
        month_start: date, month_end: date, title_prefix: str,
    ):
        """Sheet 3: 核心岗位培养统计"""
        ws = wb.create_sheet("核心岗位培养统计")

        result = await db.execute(
            select(Mentorship).where(
                and_(
                    Mentorship.store_id == store_id,
                    Mentorship.status.in_(["active", "completed"]),
                )
            )
        )
        mentorships = result.scalars().all()

        active = [m for m in mentorships if m.status == "active"]
        completed = [
            m for m in mentorships
            if m.status == "completed" and m.actual_review_date
            and month_start <= m.actual_review_date <= month_end
        ]

        _write_title_row(ws, f"{title_prefix} 核心岗位培养统计", 6)

        # 进行中
        row = 3
        ws.cell(row=row, column=1, value="一、进行中的师徒培养")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
        headers = ["师傅", "徒弟", "目标岗位", "预计考核日期", "状态"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, len(headers))
        row += 1
        for m in active:
            data = [
                m.mentor_name, m.apprentice_name, m.target_position or "",
                str(m.expected_review_date) if m.expected_review_date else "",
                "进行中",
            ]
            for col, val in enumerate(data, 1):
                ws.cell(row=row, column=col, value=val)
            _apply_data_style(ws, row, len(headers))
            row += 1
        if not active:
            ws.cell(row=row, column=1, value="无进行中的培养")
            row += 1

        # 本月完成
        row += 1
        ws.cell(row=row, column=1, value="二、本月完成的培养")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
        headers = ["师傅", "徒弟", "目标岗位", "考核日期", "奖金(元)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h)
        _apply_header_style(ws, row, len(headers))
        row += 1
        for m in completed:
            data = [
                m.mentor_name, m.apprentice_name, m.target_position or "",
                str(m.actual_review_date),
                _fen_to_yuan(m.reward_fen),
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if col == 5:
                    cell.number_format = MONEY_FORMAT
            _apply_data_style(ws, row, len(headers))
            row += 1
        if not completed:
            ws.cell(row=row, column=1, value="本月无完成的培养")
            row += 1

        # 汇总
        row += 1
        total_reward = sum(_fen_to_yuan(m.reward_fen) for m in completed)
        ws.cell(row=row, column=1, value=f"进行中: {len(active)}对")
        ws.cell(row=row, column=3, value=f"本月完成: {len(completed)}对")
        ws.cell(row=row, column=5, value=f"奖金合计: ¥{total_reward:.2f}")

        _auto_width(ws)

    async def _sheet_hourly_attendance(
        self, wb: Workbook, db: AsyncSession, store_id: str,
        month_start: date, month_end: date, title_prefix: str,
    ):
        """Sheet 4: 小时工/灵活用工考勤"""
        ws = wb.create_sheet("灵活用工考勤")

        result = await db.execute(
            select(Employee).where(
                and_(
                    Employee.store_id == store_id,
                    Employee.is_active.is_(True),
                    Employee.employment_type.in_(["part_time", "temp", "outsource_flex"]),
                )
            )
        )
        workers = result.scalars().all()

        _write_title_row(ws, f"{title_prefix} 小时工/灵活用工考勤", 6)

        headers = ["姓名", "工号", "用工类型", "出勤天数", "日薪(元)", "应发合计(元)"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=h)
        _apply_header_style(ws, 2, len(headers))

        emp_type_map = {
            "part_time": "兼职", "temp": "临时", "outsource_flex": "灵活用工",
        }

        row = 3
        total_days = 0
        total_pay = 0.0
        for w in workers:
            att_result = await db.execute(
                select(func.count(AttendanceLog.id)).where(
                    and_(
                        AttendanceLog.employee_id == w.id,
                        AttendanceLog.work_date >= month_start,
                        AttendanceLog.work_date <= month_end,
                        AttendanceLog.status.in_(["normal", "late"]),
                    )
                )
            )
            days = att_result.scalar() or 0
            daily_wage = _fen_to_yuan(w.daily_wage_standard_fen)
            pay = days * daily_wage

            data = [
                w.name, w.id, emp_type_map.get(w.employment_type, w.employment_type or ""),
                days, daily_wage, pay,
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                if col in (5, 6):
                    cell.number_format = MONEY_FORMAT
            _apply_data_style(ws, row, len(headers))
            total_days += days
            total_pay += pay
            row += 1

        if not workers:
            ws.cell(row=row, column=1, value="本月无灵活用工人员")
            row += 1
        else:
            # 合计行
            ws.cell(row=row, column=1, value="合计")
            ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=10)
            ws.cell(row=row, column=4, value=total_days)
            cell = ws.cell(row=row, column=6, value=total_pay)
            cell.number_format = MONEY_FORMAT
            cell.font = Font(name="微软雅黑", bold=True, size=10)

        ws.freeze_panes = "A3"
        _auto_width(ws)

    async def _sheet_exit_interview(
        self, wb: Workbook, db: AsyncSession, store_id: str,
        month_start: date, month_end: date, title_prefix: str,
    ):
        """Sheet 5: 离职回访汇总"""
        ws = wb.create_sheet("离职回访汇总")

        result = await db.execute(
            select(ExitInterview).where(
                and_(
                    ExitInterview.store_id == store_id,
                    ExitInterview.resign_date >= month_start,
                    ExitInterview.resign_date <= month_end,
                )
            )
        )
        interviews = result.scalars().all()

        _write_title_row(ws, f"{title_prefix} 离职回访汇总", 7)

        headers = ["姓名", "离职日期", "离职原因", "回访日期", "目前状况", "是否愿意回来", "备注"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=h)
        _apply_header_style(ws, 2, len(headers))

        reason_map = {
            "personal": "个人原因", "salary": "薪资原因", "development": "发展原因",
            "management": "管理原因", "relocation": "搬迁", "other": "其他",
        }
        return_map = {"yes": "愿意", "no": "不愿意", "maybe": "考虑中"}

        row = 3
        for iv in interviews:
            data = [
                iv.employee_name or iv.employee_id,
                str(iv.resign_date),
                reason_map.get(iv.resign_reason, iv.resign_reason or ""),
                str(iv.interview_date) if iv.interview_date else "未回访",
                iv.current_status or "",
                return_map.get(iv.willing_to_return, iv.willing_to_return or ""),
                iv.remark or "",
            ]
            for col, val in enumerate(data, 1):
                ws.cell(row=row, column=col, value=val)
            _apply_data_style(ws, row, len(headers))
            row += 1

        if not interviews:
            ws.cell(row=row, column=1, value="本月无离职记录")
            row += 1
        else:
            # 汇总
            row += 1
            interviewed = sum(1 for iv in interviews if iv.interview_date)
            willing = sum(1 for iv in interviews if iv.willing_to_return == "yes")
            ws.cell(row=row, column=1, value=f"离职总计: {len(interviews)}人")
            ws.cell(row=row, column=3, value=f"已回访: {interviewed}人")
            rate = round(interviewed / max(len(interviews), 1) * 100, 1)
            ws.cell(row=row, column=5, value=f"回访率: {rate}%")
            ws.cell(row=row, column=7, value=f"愿意回来: {willing}人")

        ws.freeze_panes = "A3"
        _auto_width(ws)

    async def _sheet_hr_summary(
        self, wb: Workbook, db: AsyncSession, store_id: str,
        brand_id: str, month: str, title_prefix: str,
    ):
        """Sheet 6: 人事工作总结与计划"""
        from src.services.hr_report_engine import HRReportEngine

        ws = wb.create_sheet("人事工作总结")

        engine = HRReportEngine(store_id, brand_id)
        report = await engine.generate_monthly_report(db, month)
        summary = report.get("hr_summary", {})

        _write_title_row(ws, f"{title_prefix} 人事工作总结与计划", 3)

        row = 3
        # 本月要点
        ws.cell(row=row, column=1, value="一、本月工作要点")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
        for item in summary.get("highlights", []):
            ws.cell(row=row, column=1, value=f"• {item}")
            ws.cell(row=row, column=1).font = NORMAL_FONT
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="二、关注事项")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
        concerns = summary.get("concerns", [])
        if concerns:
            for item in concerns:
                ws.cell(row=row, column=1, value=f"• {item}")
                ws.cell(row=row, column=1).font = NORMAL_FONT
                row += 1
        else:
            ws.cell(row=row, column=1, value="• 无重大关注事项")
            ws.cell(row=row, column=1).font = NORMAL_FONT
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="三、下月工作计划")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
        for item in summary.get("next_month_plans", []):
            ws.cell(row=row, column=1, value=f"• {item}")
            ws.cell(row=row, column=1).font = NORMAL_FONT
            row += 1

        row += 1
        turnover = summary.get("turnover_rate_pct", 0)
        ws.cell(row=row, column=1, value=f"本月离职率: {turnover}%")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=11)

        _auto_width(ws, min_width=15)

    async def _sheet_insurance_changes(
        self, wb: Workbook, db: AsyncSession, store_id: str,
        year: int, month_start: date, month_end: date, title_prefix: str,
    ):
        """Sheet 7: 社保/保险变动"""
        ws = wb.create_sheet("社保保险变动")

        # 本月新增参保
        new_result = await db.execute(
            select(EmployeeSocialInsurance, Employee.name, Employee.position).join(
                Employee, EmployeeSocialInsurance.employee_id == Employee.id
            ).where(
                and_(
                    EmployeeSocialInsurance.effective_year == year,
                    EmployeeSocialInsurance.created_at >= month_start,
                )
            )
        )
        new_enrollments = new_result.all()

        _write_title_row(ws, f"{title_prefix} 社保/保险变动", 5)

        headers = ["姓名", "工号", "岗位", "变动类型", "生效年度"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=2, column=col, value=h)
        _apply_header_style(ws, 2, len(headers))

        row = 3
        for si, name, position in new_enrollments:
            data = [name, si.employee_id, position or "", "新增参保", str(year)]
            for col, val in enumerate(data, 1):
                ws.cell(row=row, column=col, value=val)
            _apply_data_style(ws, row, len(headers))
            row += 1

        if not new_enrollments:
            ws.cell(row=row, column=1, value="本月无社保变动")
            row += 1

        # 汇总
        row += 1
        ws.cell(row=row, column=1, value=f"本月新增参保: {len(new_enrollments)}人")
        ws.cell(row=row, column=1).font = Font(name="微软雅黑", bold=True, size=10)

        ws.freeze_panes = "A3"
        _auto_width(ws)

    # ── 工具方法 ────────────────────────────────────────────────

    @staticmethod
    def _save_to_bytes(wb: Workbook) -> bytes:
        """将 Workbook 保存到 BytesIO 并返回 bytes"""
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
