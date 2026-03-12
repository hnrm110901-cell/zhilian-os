"""
月度经营报告服务（Monthly Report Service）

基于 CaseStoryGenerator 聚合月度数据，生成完整报告结构：
  - 经营摘要（成本率/营业额/损耗/决策采纳）
  - Top3 节省案例（含叙述文字）
  - 周趋势（成本率折线）
  - 场景分析（当月主要经营场景）
  - HTML 版本（供浏览器打印为 PDF）

Rule 6 兼容：所有金额字段含 _yuan
Rule 8 兼容：仅涉及 MVP-9 月度经营报告功能
"""

from __future__ import annotations

import os
from calendar import monthrange
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional

import structlog
from sqlalchemy.ext.asyncio import AsyncSession

from src.services.case_story_generator import CaseStoryGenerator

logger = structlog.get_logger()

# 报告徽标文字（可通过环境变量覆盖）
BRAND_NAME = os.getenv("BRAND_NAME", "屯象经营助手")


# ════════════════════════════════════════════════════════════════════════════════
# 纯函数：报告内容构建
# ════════════════════════════════════════════════════════════════════════════════

def build_executive_summary(monthly_story: Dict[str, Any]) -> Dict[str, Any]:
    """
    从月度故事数据提取高管摘要（一屏可读）。

    Returns:
        {
          headline, cost_rate_pct, cost_rate_status, revenue_yuan,
          waste_cost_yuan, decision_adoption_pct, total_saving_yuan,
          narrative
        }
    """
    cm = monthly_story["cost_metrics"]
    dm = monthly_story["decision_summary"]
    period = monthly_story["period_label"]

    status_map = {"ok": "正常", "warning": "偏高", "critical": "超标"}
    status_label = status_map.get(cm["cost_rate_status"], "")

    headline_parts = []
    if cm["cost_rate_status"] == "critical":
        headline_parts.append(f"⚠️ 成本率超标，需重点关注")
    elif dm["total_saving_yuan"] > 0:
        headline_parts.append(f"✅ 本月通过决策执行累计节省 ¥{dm['total_saving_yuan']:,.0f}")
    else:
        headline_parts.append(f"本月经营整体{status_label}")

    return {
        "period":                period,
        "headline":              " | ".join(headline_parts),
        "revenue_yuan":          cm["revenue_yuan"],
        "actual_cost_pct":       cm["actual_cost_pct"],
        "cost_rate_status":      cm["cost_rate_status"],
        "cost_rate_label":       status_label,
        "waste_cost_yuan":       cm["waste_cost_yuan"],
        "waste_pct":             cm["waste_pct"],
        "decision_adoption_pct": dm["adoption_rate_pct"],
        "total_saving_yuan":     dm["total_saving_yuan"],
        "decisions_approved":    dm["approved"],
        "decisions_total":       dm["total"],
        "narrative":             monthly_story["narrative"],
    }


def build_weekly_trend_chart(weekly_trend: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    构建周趋势折线图所需数据（ECharts 格式）。

    Returns:
        {x_axis: [...], cost_rate_series: [...], revenue_series: [...]}
    """
    x_axis         = [w["week_start"] for w in weekly_trend]
    cost_rate_data = [w["actual_cost_pct"] for w in weekly_trend]
    revenue_data   = [w["revenue_yuan"] for w in weekly_trend]
    status_colors  = {
        "ok":       "#52c41a",
        "warning":  "#faad14",
        "critical": "#f5222d",
    }
    point_colors = [
        status_colors.get(w.get("cost_rate_status", "ok"), "#1890ff")
        for w in weekly_trend
    ]

    return {
        "x_axis":         x_axis,
        "cost_rate_data": cost_rate_data,
        "revenue_data":   revenue_data,
        "point_colors":   point_colors,
    }


def render_html_report(
    store_id:         str,
    year:             int,
    month:            int,
    executive_summary: Dict[str, Any],
    top3_decisions:   List[Dict[str, Any]],
    weekly_chart:     Dict[str, Any],
) -> str:
    """
    将报告数据渲染为 HTML 字符串（可在浏览器打印为 PDF）。

    样式：极简打印友好，无外部依赖，Ant Design 色系。
    """
    period = executive_summary["period"]
    rev    = executive_summary["revenue_yuan"]
    cost   = executive_summary["actual_cost_pct"]
    waste  = executive_summary["waste_cost_yuan"]
    save   = executive_summary["total_saving_yuan"]
    adopt  = executive_summary["decision_adoption_pct"]

    # Top3 决策节
    top3_rows = ""
    for i, d in enumerate(top3_decisions[:3], 1):
        saving = d.get("expected_saving_yuan", 0)
        action = d.get("action", "")
        outcome = d.get("outcome") or "待统计"
        top3_rows += f"""
        <tr>
          <td style="padding:8px;border-bottom:1px solid #f0f0f0;font-weight:bold;">#{i}</td>
          <td style="padding:8px;border-bottom:1px solid #f0f0f0;">{action}</td>
          <td style="padding:8px;border-bottom:1px solid #f0f0f0;color:#1890ff;">¥{saving:,.0f}</td>
          <td style="padding:8px;border-bottom:1px solid #f0f0f0;">{outcome}</td>
        </tr>"""

    # 周趋势表格（文字版，无需图表库）
    trend_rows = ""
    x_axis     = weekly_chart.get("x_axis", [])
    cost_data  = weekly_chart.get("cost_rate_data", [])
    rev_data   = weekly_chart.get("revenue_data", [])
    colors     = weekly_chart.get("point_colors", [])

    for i, week in enumerate(x_axis):
        c = cost_data[i] if i < len(cost_data) else 0
        r = rev_data[i]  if i < len(rev_data)  else 0
        color = colors[i] if i < len(colors) else "#1890ff"
        trend_rows += f"""
        <tr>
          <td style="padding:6px 8px;border-bottom:1px solid #f5f5f5;">{week}</td>
          <td style="padding:6px 8px;border-bottom:1px solid #f5f5f5;color:{color};font-weight:bold;">{c:.1f}%</td>
          <td style="padding:6px 8px;border-bottom:1px solid #f5f5f5;">¥{r:,.0f}</td>
        </tr>"""

    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M")

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1.0" />
  <title>{period}经营报告 — {store_id}</title>
  <style>
    @media print {{ @page {{ size: A4; margin: 15mm; }} }}
    body {{ font-family: "PingFang SC","Helvetica Neue",Arial,sans-serif;
           color:#333; margin:0; padding:24px; font-size:14px; line-height:1.6; }}
    h1 {{ font-size:22px; color:#1890ff; margin-bottom:4px; }}
    h2 {{ font-size:16px; color:#595959; border-bottom:2px solid #1890ff;
          padding-bottom:4px; margin-top:28px; }}
    .kpi-grid {{ display:grid; grid-template-columns:repeat(3,1fr); gap:16px; margin:16px 0; }}
    .kpi-card {{ background:#fafafa; border:1px solid #e8e8e8; border-radius:6px;
                 padding:12px 16px; }}
    .kpi-label {{ font-size:12px; color:#8c8c8c; }}
    .kpi-value {{ font-size:22px; font-weight:700; color:#1890ff; }}
    .kpi-value.critical {{ color:#f5222d; }}
    .kpi-value.warning  {{ color:#faad14; }}
    .kpi-value.ok       {{ color:#52c41a; }}
    table {{ width:100%; border-collapse:collapse; }}
    thead th {{ background:#e6f7ff; padding:8px; text-align:left; font-size:13px; }}
    .headline {{ background:#e6f7ff; border-left:4px solid #1890ff;
                 padding:10px 16px; border-radius:0 4px 4px 0; margin:12px 0; }}
    .footer {{ margin-top:36px; text-align:center; color:#bfbfbf; font-size:12px; }}
  </style>
</head>
<body>
  <h1>{BRAND_NAME} · {period}经营报告</h1>
  <p style="color:#8c8c8c;margin:0;">门店 {store_id} &nbsp;·&nbsp; 生成于 {now}</p>

  <div class="headline">{executive_summary["headline"]}</div>

  <h2>核心经营指标</h2>
  <div class="kpi-grid">
    <div class="kpi-card">
      <div class="kpi-label">月营业额</div>
      <div class="kpi-value">¥{rev:,.0f}</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">食材成本率</div>
      <div class="kpi-value {executive_summary['cost_rate_status']}">{cost:.1f}%</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">本月损耗金额</div>
      <div class="kpi-value">¥{waste:,.0f}</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">决策采纳率</div>
      <div class="kpi-value">{adopt:.0f}%</div>
    </div>
    <div class="kpi-card">
      <div class="kpi-label">决策节省金额</div>
      <div class="kpi-value">¥{save:,.0f}</div>
    </div>
  </div>

  <p style="color:#595959;">{executive_summary["narrative"]}</p>

  <h2>Top3 关键决策案例</h2>
  <table>
    <thead>
      <tr>
        <th>排名</th><th>执行动作</th><th>预期节省</th><th>执行结果</th>
      </tr>
    </thead>
    <tbody>{top3_rows}</tbody>
  </table>

  <h2>成本率周趋势</h2>
  <table>
    <thead>
      <tr><th>周起始</th><th>食材成本率</th><th>营业额</th></tr>
    </thead>
    <tbody>{trend_rows}</tbody>
  </table>

  <div class="footer">
    {BRAND_NAME} · v2.0 · 本报告由系统自动生成，数据来源于 POS 实时同步
  </div>
</body>
</html>"""


# ════════════════════════════════════════════════════════════════════════════════
# MonthlyReportService
# ════════════════════════════════════════════════════════════════════════════════

class MonthlyReportService:
    """月度经营报告服务：聚合数据 → 构建报告 → 输出 JSON / HTML"""

    @staticmethod
    async def generate(
        store_id: str,
        year:     int,
        month:    int,
        db:       AsyncSession,
    ) -> Dict[str, Any]:
        """
        生成月度报告 JSON 数据结构。

        Returns:
            {
              store_id, year, month, period_label,
              executive_summary, weekly_trend_chart,
              top3_decisions, cost_metrics, decision_summary,
              generated_at
            }
        """
        from src.services.behavior_score_engine import BehaviorScoreEngine

        monthly_story = await CaseStoryGenerator.generate_monthly_story(
            store_id=store_id, year=year, month=month, db=db
        )

        # 用 BehaviorScoreEngine 替代 CaseStoryGenerator 的手工采纳率汇总（更准确）
        try:
            days = monthrange(year, month)[1]
            behavior = await BehaviorScoreEngine.get_store_report(
                store_id=store_id,
                start_date=date(year, month, 1),
                end_date=date(year, month, days),
                db=db,
            )
            # 覆盖 decision_summary 中的采纳率字段
            monthly_story["decision_summary"]["adoption_rate_pct"] = behavior["adoption_rate_pct"]
            monthly_story["decision_summary"]["total_saving_yuan"]  = behavior["total_saving_yuan"]
            monthly_story["decision_summary"]["total"]              = behavior["total_sent"]
            monthly_story["decision_summary"]["approved"]           = behavior["total_adopted"]
        except Exception as exc:
            logger.warning(
                "monthly_report.behavior_engine_failed",
                store_id=store_id,
                error=str(exc),
            )

        executive_summary = build_executive_summary(monthly_story)
        weekly_chart      = build_weekly_trend_chart(monthly_story["weekly_trend"])

        logger.info(
            "monthly_report_generated",
            store_id=store_id, year=year, month=month,
            cost_pct=monthly_story["cost_metrics"]["actual_cost_pct"],
        )

        return {
            "store_id":              store_id,
            "year":                  year,
            "month":                 month,
            "period_label":          monthly_story["period_label"],
            "executive_summary":     executive_summary,
            "weekly_trend_chart":    weekly_chart,
            "top3_decisions":        monthly_story["top3_decisions"],
            "cost_metrics":          monthly_story["cost_metrics"],
            "decision_summary":      monthly_story["decision_summary"],
            "generated_at":          monthly_story["generated_at"],
        }

    @staticmethod
    async def generate_html(
        store_id: str,
        year:     int,
        month:    int,
        db:       AsyncSession,
    ) -> str:
        """生成月度报告 HTML（供浏览器打印为 PDF）"""
        report = await MonthlyReportService.generate(store_id, year, month, db)
        return render_html_report(
            store_id          = store_id,
            year              = year,
            month             = month,
            executive_summary = report["executive_summary"],
            top3_decisions    = report["top3_decisions"],
            weekly_chart      = report["weekly_trend_chart"],
        )

    @staticmethod
    async def generate_excel(
        store_id: str,
        year:     int,
        month:    int,
        db:       AsyncSession,
    ) -> bytes:
        """
        生成月度报告 Excel（.xlsx）字节流，供下载。

        工作表：
          Sheet1 经营摘要  — KPI 核心指标
          Sheet2 周趋势    — 成本率 + 营业额折线数据
          Sheet3 Top3决策  — 本月最高价值决策明细
        """
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        report = await MonthlyReportService.generate(store_id, year, month, db)

        wb = openpyxl.Workbook()

        # ── 通用样式 ──────────────────────────────────────────────
        HDR_FILL  = PatternFill("solid", fgColor="FF6B2C")  # 品牌橙
        HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
        TITLE_FONT = Font(bold=True, size=13)
        BORDER    = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )
        CENTER    = Alignment(horizontal="center", vertical="center")
        RIGHT     = Alignment(horizontal="right")

        def _set_header(ws, row: int, cols: list[str]) -> None:
            for c, label in enumerate(cols, start=1):
                cell = ws.cell(row=row, column=c, value=label)
                cell.fill   = HDR_FILL
                cell.font   = HDR_FONT
                cell.border = BORDER
                cell.alignment = CENTER

        def _set_cell(ws, row: int, col: int, value, num_fmt: str = None) -> None:
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = BORDER
            if num_fmt:
                cell.number_format = num_fmt
            return cell

        # ════════════════════════════════════════════════════════
        # Sheet1: 经营摘要
        # ════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "经营摘要"
        period = report["period_label"]

        # 标题行
        ws1.merge_cells("A1:C1")
        title_cell = ws1["A1"]
        title_cell.value = f"{BRAND_NAME} — {period} 月度经营报告"
        title_cell.font  = TITLE_FONT
        title_cell.alignment = CENTER

        es = report["executive_summary"]
        cm = report["cost_metrics"]
        dm = report["decision_summary"]

        kpis = [
            ("指标",          "数值",          "单位"),
            ("营业额",         es.get("revenue_yuan", "—"),    "元"),
            ("食材成本率",      f"{cm.get('actual_cost_pct', 0):.2f}%",  "%"),
            ("成本率状态",      {"ok": "正常", "warning": "偏高", "critical": "超标"}.get(
                cm.get("cost_rate_status", ""), "—"), ""),
            ("损耗成本",       es.get("waste_cost_yuan", "—"),  "元"),
            ("决策采纳率",     f"{dm.get('adoption_rate_pct', 0):.1f}%", "%"),
            ("累计节省",       dm.get("total_saving_yuan", 0),   "元"),
            ("发出决策数",     dm.get("total", 0),               "条"),
            ("已执行决策数",   dm.get("approved", 0),            "条"),
        ]
        _set_header(ws1, 2, kpis[0])
        for r_offset, (label, val, unit) in enumerate(kpis[1:], start=3):
            _set_cell(ws1, r_offset, 1, label)
            _set_cell(ws1, r_offset, 2, val)
            _set_cell(ws1, r_offset, 3, unit)

        for col_idx in range(1, 4):
            ws1.column_dimensions[get_column_letter(col_idx)].width = 20

        # ════════════════════════════════════════════════════════
        # Sheet2: 周趋势
        # ════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("周趋势")
        ws2.merge_cells("A1:D1")
        t2 = ws2["A1"]
        t2.value = f"{period} 周趋势数据"
        t2.font  = TITLE_FONT
        t2.alignment = CENTER

        wc = report.get("weekly_trend_chart", {})
        weeks   = wc.get("weeks", [])
        cost_rates = wc.get("cost_rate_pcts", [])
        revenues   = wc.get("revenues", [])

        _set_header(ws2, 2, ["周次", "成本率(%)", "营业额(元)", "目标成本率(%)"])
        target_pct = cm.get("target_cost_pct", "—")
        for i, week in enumerate(weeks):
            r = i + 3
            _set_cell(ws2, r, 1, week)
            _set_cell(ws2, r, 2, cost_rates[i] if i < len(cost_rates) else None, "0.00")
            _set_cell(ws2, r, 3, revenues[i]   if i < len(revenues)   else None, "#,##0.00")
            _set_cell(ws2, r, 4, target_pct)

        for col_idx in range(1, 5):
            ws2.column_dimensions[get_column_letter(col_idx)].width = 18

        # ════════════════════════════════════════════════════════
        # Sheet3: Top3 决策
        # ════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Top3决策")
        ws3.merge_cells("A1:E1")
        t3 = ws3["A1"]
        t3.value = f"{period} 高价值决策明细"
        t3.font  = TITLE_FONT
        t3.alignment = CENTER

        _set_header(ws3, 2, ["排名", "决策标题", "决策叙述", "节省金额(元)", "置信度(%)"])
        for rank, decision in enumerate(report.get("top3_decisions", []), start=1):
            r = rank + 2
            _set_cell(ws3, r, 1, rank)
            _set_cell(ws3, r, 2, decision.get("title", ""))
            narr_cell = _set_cell(ws3, r, 3, decision.get("narrative", ""))
            narr_cell.alignment = Alignment(wrap_text=True)
            _set_cell(ws3, r, 4, decision.get("saving_yuan", 0), "#,##0.00")
            _set_cell(ws3, r, 5, decision.get("confidence_pct", 0), "0.0")

        ws3.column_dimensions["A"].width = 6
        ws3.column_dimensions["B"].width = 28
        ws3.column_dimensions["C"].width = 50
        ws3.column_dimensions["D"].width = 16
        ws3.column_dimensions["E"].width = 12

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

