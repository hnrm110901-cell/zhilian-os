"""
花名册Excel导入服务 — 支持从乐才/钉钉/企微等HR系统导入员工数据
"""
from typing import Any, Dict, List, Optional, Tuple
from datetime import date, datetime
import structlog
import io

from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.employee import Employee
from src.models.organization import Organization

logger = structlog.get_logger()

# 标准列映射（乐才花名册格式）
LECAI_COLUMN_MAP = {
    "工号": "id",
    "姓名": "name",
    "手机号码": "phone",
    "邮箱": "email",
    "岗位": "position",
    "用工类型": "employment_type",
    "职级": "grade_level",
    "健康证到期日": "health_cert_expiry",
    "身份证号码": "id_card_no",
    "身份证到期日": "id_card_expiry",
    "背调状态": "background_check",
    "性别": "gender",
    "出生日期": "birth_date",
    "学历": "education",
    "婚姻状况": "marital_status",
    "民族": "ethnicity",
    "户籍类型": "hukou_type",
    "户籍地址": "hukou_location",
    "身高": "height_cm",
    "体重": "weight_kg",
    "政治面貌": "political_status",
    "紧急联系人": "emergency_contact",
    "紧急联系人电话": "emergency_phone",
    "紧急联系人关系": "emergency_relation",
    "开户行": "bank_name",
    "银行卡号": "bank_account",
    "支行": "bank_branch",
    "日薪标准": "daily_wage_standard_fen",
    "工时类型": "work_hour_type",
    "首次工作日期": "first_work_date",
    "转正日期": "regular_date",
    "司龄月数": "seniority_months",
    "宿舍": "accommodation",
    "工会会员": "union_member",
    "工会干部": "union_cadre",
    "专业": "major",
    "毕业院校": "graduation_school",
    "专业证书": "professional_cert",
    "入职日期": "hire_date",
    "在职状态": "employment_status",
    "门店": "store_id",
    "部门": "_department",
    "区域": "_region",
    "品牌": "_brand",
}

# 日期类字段
DATE_FIELDS = {
    "health_cert_expiry", "id_card_expiry", "birth_date",
    "first_work_date", "regular_date", "hire_date",
}

# 布尔类字段
BOOL_FIELDS = {"union_member", "union_cadre"}

# 金额类字段（元→分）
FEN_FIELDS = {"daily_wage_standard_fen"}


class HRRosterImportService:
    """花名册导入服务"""

    def __init__(self, brand_id: str, store_id: Optional[str] = None):
        self.brand_id = brand_id
        self.store_id = store_id

    async def preview_import(
        self,
        db: AsyncSession,
        file_bytes: bytes,
        column_map: Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        """
        预览导入：解析Excel → 返回列映射 + 前10行预览 + 统计
        """
        try:
            import openpyxl
        except ImportError:
            return {"error": "请安装 openpyxl: pip install openpyxl"}

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        if not ws:
            return {"error": "Excel文件无有效工作表"}

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return {"error": "Excel文件无数据行"}

        headers = [str(h).strip() if h else "" for h in rows[0]]
        mapping = column_map or LECAI_COLUMN_MAP

        # 自动匹配列
        matched_columns = {}
        unmatched_columns = []
        for i, header in enumerate(headers):
            if header in mapping:
                matched_columns[header] = mapping[header]
            else:
                unmatched_columns.append(header)

        # 预览前10行
        preview_rows = []
        for row in rows[1:11]:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    field = matched_columns.get(headers[i])
                    if field:
                        row_dict[field] = str(val) if val is not None else None
            preview_rows.append(row_dict)

        wb.close()

        return {
            "total_rows": len(rows) - 1,
            "total_columns": len(headers),
            "matched_columns": len(matched_columns),
            "unmatched_columns": unmatched_columns,
            "column_mapping": matched_columns,
            "preview": preview_rows,
        }

    async def confirm_import(
        self,
        db: AsyncSession,
        file_bytes: bytes,
        column_map: Optional[Dict[str, str]] = None,
        default_store_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        确认导入：解析Excel → 批量upsert员工数据
        """
        try:
            import openpyxl
        except ImportError:
            return {"error": "请安装 openpyxl: pip install openpyxl"}

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in rows[0]]
        mapping = column_map or LECAI_COLUMN_MAP

        # 构建列索引
        col_index = {}
        for i, header in enumerate(headers):
            if header in mapping:
                col_index[i] = mapping[header]

        stats = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
        store_id = default_store_id or self.store_id

        for row_num, row in enumerate(rows[1:], start=2):
            try:
                data = {}
                for i, val in enumerate(row):
                    field = col_index.get(i)
                    if not field or field.startswith("_"):
                        continue
                    data[field] = self._convert_value(field, val)

                if not data.get("id") or not data.get("name"):
                    stats["skipped"] += 1
                    continue

                emp_id = str(data["id"]).strip()
                data["store_id"] = data.get("store_id") or store_id

                # upsert
                existing = await db.execute(
                    select(Employee).where(Employee.id == emp_id)
                )
                emp = existing.scalar_one_or_none()

                if emp:
                    for key, val in data.items():
                        if key != "id" and val is not None:
                            setattr(emp, key, val)
                    stats["updated"] += 1
                else:
                    emp = Employee(**data)
                    db.add(emp)
                    stats["created"] += 1

                # 自动创建组织节点（如果有部门/区域信息）
                dept = None
                for i, val in enumerate(row):
                    field = col_index.get(i)
                    if field == "_department" and val:
                        dept = str(val).strip()

                if dept:
                    await self._ensure_org_node(db, dept, data.get("store_id"))

            except Exception as e:
                stats["errors"].append({"row": row_num, "error": str(e)})
                logger.warning("roster_import_row_error", row=row_num, error=str(e))

        await db.flush()
        wb.close()

        logger.info(
            "roster_import_completed",
            brand_id=self.brand_id,
            created=stats["created"],
            updated=stats["updated"],
        )
        return stats

    def _convert_value(self, field: str, val: Any) -> Any:
        """根据字段类型转换值"""
        if val is None:
            return None

        if field in DATE_FIELDS:
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, date):
                return val
            try:
                return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
            except (ValueError, TypeError):
                try:
                    return datetime.strptime(str(val).strip(), "%Y/%m/%d").date()
                except (ValueError, TypeError):
                    return None

        if field in BOOL_FIELDS:
            return str(val).strip().lower() in ("是", "true", "1", "yes")

        if field in FEN_FIELDS:
            try:
                return int(float(str(val)) * 100)
            except (ValueError, TypeError):
                return None

        if field in ("height_cm", "weight_kg", "seniority_months"):
            try:
                return int(float(str(val)))
            except (ValueError, TypeError):
                return None

        return str(val).strip() if val else None

    async def _ensure_org_node(
        self, db: AsyncSession, dept_name: str, store_id: Optional[str]
    ) -> None:
        """确保组织节点存在，不存在则创建"""
        existing = await db.execute(
            select(Organization).where(
                and_(
                    Organization.brand_id == self.brand_id,
                    Organization.name == dept_name,
                )
            )
        )
        if existing.scalar_one_or_none():
            return

        org = Organization(
            brand_id=self.brand_id,
            name=dept_name,
            code=f"{self.brand_id}_{dept_name}",
            level=6,  # 默认部门组
            org_type="department",
            store_id=store_id,
        )
        db.add(org)
