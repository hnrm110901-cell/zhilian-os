"""
Celery异步任务
用于Neural System的事件处理和向量数据库索引
"""

import asyncio
import inspect
import os
import re
import sys
from typing import Any, Dict

import structlog

# 确保项目根目录在 sys.path 上，使 `packages.api_adapters.*` 可直接 import
_PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "../../../.."))
if _PROJECT_ROOT not in sys.path:
    sys.path.insert(0, _PROJECT_ROOT)

# 表名只允许小写字母/数字/下划线，防止通过 backup_jobs.tables 字段注入任意 SQL
_SAFE_TABLE_RE = re.compile(r"^[a-z][a-z0-9_]*$")
from celery import Task

from .celery_app import celery_app

logger = structlog.get_logger()


async def _maybe_await(value):
    if inspect.isawaitable(value):
        return await value
    return value


class CallbackTask(Task):
    """带回调的任务基类"""

    def on_success(self, retval, task_id, args, kwargs):
        """任务成功回调"""
        logger.info(
            "Celery任务成功",
            task_id=task_id,
            task_name=self.name,
            result=retval,
        )

    def on_failure(self, exc, task_id, args, kwargs, einfo):
        """任务失败回调"""
        logger.error(
            "Celery任务失败",
            task_id=task_id,
            task_name=self.name,
            error=str(exc),
            traceback=str(einfo),
        )

    def on_retry(self, exc, task_id, args, kwargs, einfo):
        """任务重试回调"""
        logger.warning(
            "Celery任务重试",
            task_id=task_id,
            task_name=self.name,
            error=str(exc),
            retry_count=self.request.retries,
        )


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY", "60")),
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=int(os.getenv("CELERY_RETRY_BACKOFF_MAX", "600")),
    retry_jitter=True,
)
def process_neural_event(
    self,
    event_id: str,
    event_type: str,
    event_source: str,
    store_id: str,
    data: Dict[str, Any],
    priority: int = 0,
) -> Dict[str, Any]:
    """
    处理神经系统事件（异步任务）

    Args:
        event_id: 事件ID
        event_type: 事件类型
        event_source: 事件来源
        store_id: 门店ID
        data: 事件数据
        priority: 优先级

    Returns:
        处理结果
    """

    async def _run():
        from datetime import datetime

        from ..core.database import AsyncSessionLocal
        from ..models.neural_event_log import EventProcessingStatus, NeuralEventLog
        from ..services.vector_db_service import vector_db_service

        # 1. 写入 DB — 标记为 processing
        async with AsyncSessionLocal() as session:
            log = NeuralEventLog(
                event_id=event_id,
                celery_task_id=self.request.id,
                event_type=event_type,
                event_source=event_source,
                store_id=store_id,
                priority=priority,
                data=data,
                processing_status=EventProcessingStatus.PROCESSING,
                queued_at=datetime.utcnow(),
                started_at=datetime.utcnow(),
            )
            session.add(log)
            await session.commit()

        logger.info(
            "开始处理神经系统事件",
            event_id=event_id,
            event_type=event_type,
            store_id=store_id,
        )

        actions_taken = []
        downstream_tasks = []
        vector_indexed = False
        wechat_sent = False

        try:
            # 2. 向量化存储（全局索引 + 领域分割索引）
            event_payload = {
                "event_id": event_id,
                "event_type": event_type,
                "event_source": event_source,
                "timestamp": datetime.utcnow(),
                "store_id": store_id,
                "data": data,
                "priority": priority,
            }
            await vector_db_service.index_event(event_payload)
            from ..services.domain_vector_service import domain_vector_service

            await domain_vector_service.index_neural_event(store_id, event_payload)
            vector_indexed = True
            actions_taken.append("vector_indexed")

            # 3. 触发企微推送
            from ..services.wechat_trigger_service import wechat_trigger_service

            try:
                await wechat_trigger_service.trigger_push(
                    event_type=event_type,
                    event_data=data,
                    store_id=store_id,
                )
                wechat_sent = True
                actions_taken.append("wechat_sent")
            except Exception as e:
                logger.warning("企微推送触发失败", event_type=event_type, error=str(e))

            # 4. 根据事件类型触发下游任务
            if event_type.startswith("order."):
                t = index_order_to_vector_db.delay(data)
                downstream_tasks.append({"task_name": "index_order_to_vector_db", "task_id": t.id})
                actions_taken.append("dispatched:index_order_to_vector_db")
            elif event_type.startswith("dish."):
                t = index_dish_to_vector_db.delay(data)
                downstream_tasks.append({"task_name": "index_dish_to_vector_db", "task_id": t.id})
                actions_taken.append("dispatched:index_dish_to_vector_db")

            processed_at = datetime.utcnow()

            # 5. 写回 DB — 标记为 completed
            async with AsyncSessionLocal() as session:
                db_log = await session.get(NeuralEventLog, event_id)
                if db_log:
                    db_log.processing_status = EventProcessingStatus.COMPLETED
                    db_log.vector_indexed = vector_indexed
                    db_log.wechat_sent = wechat_sent
                    db_log.downstream_tasks = downstream_tasks
                    db_log.actions_taken = actions_taken
                    db_log.processed_at = processed_at
                    await session.commit()

            logger.info("神经系统事件处理完成", event_id=event_id, event_type=event_type)
            return {
                "success": True,
                "event_id": event_id,
                "processed_at": processed_at.isoformat(),
                "actions_taken": actions_taken,
            }

        except Exception as e:
            logger.error("神经系统事件处理失败", event_id=event_id, error=str(e), exc_info=e)
            # 写回 DB — 标记为 failed / retrying
            try:
                is_last_retry = self.request.retries >= self.max_retries
                async with AsyncSessionLocal() as session:
                    db_log = await session.get(NeuralEventLog, event_id)
                    if db_log:
                        db_log.processing_status = (
                            EventProcessingStatus.FAILED if is_last_retry else EventProcessingStatus.RETRYING
                        )
                        db_log.error_message = str(e)
                        db_log.retry_count = self.request.retries + 1
                        await session.commit()
            except Exception as db_err:
                logger.warning("celery_tasks.status_update_failed", error=str(db_err))
            raise self.retry(exc=e)

    return asyncio.run(_run())


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY_SHORT", "30")),
)
def index_to_vector_db(
    self,
    collection_name: str,
    data: Dict[str, Any],
) -> Dict[str, Any]:
    """
    索引数据到向量数据库（通用任务）

    Args:
        collection_name: 集合名称
        data: 要索引的数据

    Returns:
        索引结果
    """

    async def _run():
        try:
            from ..services.vector_db_service import vector_db_service

            logger.info(
                "开始索引到向量数据库",
                collection=collection_name,
                data_id=data.get("id"),
            )

            # 根据集合类型调用相应的索引方法
            if collection_name == "orders":
                await vector_db_service.index_order(data)
            elif collection_name == "dishes":
                await vector_db_service.index_dish(data)
            elif collection_name == "events":
                await vector_db_service.index_event(data)
            else:
                raise ValueError(f"不支持的集合类型: {collection_name}")

            logger.info(
                "向量数据库索引完成",
                collection=collection_name,
                data_id=data.get("id"),
            )

            return {
                "success": True,
                "collection": collection_name,
                "data_id": data.get("id"),
            }

        except Exception as e:
            logger.error(
                "向量数据库索引失败",
                collection=collection_name,
                error=str(e),
                exc_info=e,
            )
            raise self.retry(exc=e)

    return asyncio.run(_run())


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
)
def index_order_to_vector_db(
    self,
    order_data: Dict[str, Any],
) -> Dict[str, Any]:
    """
    索引订单到向量数据库

    Args:
        order_data: 订单数据

    Returns:
        索引结果
    """

    async def _run():
        from ..services.domain_vector_service import domain_vector_service
        from ..services.vector_db_service import vector_db_service

        store_id = order_data.get("store_id", "")
        logger.info("开始索引到向量数据库", collection="orders", data_id=order_data.get("id"))
        await vector_db_service.index_order(order_data)
        await domain_vector_service.index_revenue_event(store_id, order_data)
        logger.info("向量数据库索引完成", collection="orders/revenue", data_id=order_data.get("id"))
        return {"success": True, "collection": "orders", "data_id": order_data.get("id")}

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
)
def index_dish_to_vector_db(
    self,
    dish_data: Dict[str, Any],
) -> Dict[str, Any]:
    """
    索引菜品到向量数据库

    Args:
        dish_data: 菜品数据

    Returns:
        索引结果
    """

    async def _run():
        from ..services.domain_vector_service import domain_vector_service
        from ..services.vector_db_service import vector_db_service

        store_id = dish_data.get("store_id", "")
        logger.info("开始索引到向量数据库", collection="dishes", data_id=dish_data.get("id"))
        await vector_db_service.index_dish(dish_data)
        await domain_vector_service.index_menu_item(store_id, dish_data)
        logger.info("向量数据库索引完成", collection="dishes/menu", data_id=dish_data.get("id"))
        return {"success": True, "collection": "dishes", "data_id": dish_data.get("id")}

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


@celery_app.task(
    base=CallbackTask,
    bind=True,
)
def batch_index_orders(
    self,
    orders: list[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    批量索引订单到向量数据库

    Args:
        orders: 订单列表

    Returns:
        批量索引结果
    """
    try:
        logger.info("开始批量索引订单", count=len(orders))

        # 为每个订单创建异步任务
        tasks = [index_order_to_vector_db.delay(order) for order in orders]

        # 等待所有任务完成
        results = [task.get(timeout=int(os.getenv("CELERY_TASK_GET_TIMEOUT", "300"))) for task in tasks]

        success_count = sum(1 for r in results if r.get("success"))

        logger.info(
            "批量索引订单完成",
            total=len(orders),
            success=success_count,
            failed=len(orders) - success_count,
        )

        return {
            "success": True,
            "total": len(orders),
            "success_count": success_count,
            "failed_count": len(orders) - success_count,
        }

    except Exception as e:
        logger.error("批量索引订单失败", error=str(e), exc_info=e)
        raise self.retry(exc=e)


@celery_app.task(
    base=CallbackTask,
    bind=True,
)
def batch_index_dishes(
    self,
    dishes: list[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    批量索引菜品到向量数据库

    Args:
        dishes: 菜品列表

    Returns:
        批量索引结果
    """
    try:
        logger.info("开始批量索引菜品", count=len(dishes))

        # 为每个菜品创建异步任务
        tasks = [index_dish_to_vector_db.delay(dish) for dish in dishes]

        # 等待所有任务完成
        results = [task.get(timeout=int(os.getenv("CELERY_TASK_GET_TIMEOUT", "300"))) for task in tasks]

        success_count = sum(1 for r in results if r.get("success"))

        logger.info(
            "批量索引菜品完成",
            total=len(dishes),
            success=success_count,
            failed=len(dishes) - success_count,
        )

        return {
            "success": True,
            "total": len(dishes),
            "success_count": success_count,
            "failed_count": len(dishes) - success_count,
        }

    except Exception as e:
        logger.error("批量索引菜品失败", error=str(e), exc_info=e)
        raise self.retry(exc=e)


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY_LONG", "300")),  # 5分钟
)
def generate_and_send_daily_report(
    self,
    store_id: str = None,
    report_date: str = None,
) -> Dict[str, Any]:
    """
    生成并发送营业日报

    Args:
        store_id: 门店ID (None表示为所有门店生成，Beat调度时使用)
        report_date: 报告日期（YYYY-MM-DD格式，默认为昨天）

    Returns:
        生成和发送结果
    """

    async def _run():
        try:
            from datetime import date, datetime, timedelta

            from sqlalchemy import select

            from ..core.database import get_db_session
            from ..models.store import Store
            from ..models.user import User, UserRole
            from ..services.daily_report_service import daily_report_service
            from ..services.wechat_work_message_service import wechat_work_message_service

            # 解析日期
            target_date = (
                datetime.strptime(report_date, "%Y-%m-%d").date() if report_date else date.today() - timedelta(days=1)
            )

            logger.info("开始生成营业日报", store_id=store_id, report_date=str(target_date))

            # 获取要生成报告的门店列表
            async with get_db_session() as session:
                if store_id:
                    result = await session.execute(select(Store).where(Store.id == store_id, Store.is_active == True))
                else:
                    result = await session.execute(select(Store).where(Store.is_active == True))
                stores = result.scalars().all()

            total_sent = 0
            for store in stores:
                try:
                    # 1. 生成日报
                    report = await daily_report_service.generate_daily_report(store_id=str(store.id), report_date=target_date)

                    # 2. 构建推送消息
                    message = f"""【营业日报】{target_date.strftime('%Y年%m月%d日')}
门店：{store.name}（{store.id}）

{report.summary}

📊 详细数据：
• 订单数：{report.order_count}笔
• 客流量：{report.customer_count}人
• 客单价：¥{report.avg_order_value / 100:.2f}

📈 运营指标：
• 任务完成率：{report.task_completion_rate:.1f}%
• 库存预警：{report.inventory_alert_count}个
"""

                    if report.highlights:
                        message += "\n✨ 今日亮点：\n"
                        for highlight in report.highlights:
                            message += f"• {highlight}\n"

                    if report.alerts:
                        message += "\n⚠️ 需要关注：\n"
                        for alert in report.alerts:
                            message += f"• {alert}\n"

                    # 3. 查询店长和管理员，发送推送
                    async with get_db_session() as session:
                        mgr_result = await session.execute(
                            select(User).where(
                                User.store_id == store.id,
                                User.is_active == True,
                                User.role.in_([UserRole.STORE_MANAGER, UserRole.ADMIN]),
                                User.wechat_user_id.isnot(None),
                            )
                        )
                        managers = mgr_result.scalars().all()

                    sent_count = 0
                    for manager in managers:
                        try:
                            send_result = await wechat_work_message_service.send_text_message(
                                user_id=manager.wechat_user_id, content=message
                            )
                            if send_result.get("success"):
                                sent_count += 1
                        except Exception as send_err:
                            logger.error("发送日报失败", user_id=str(manager.id), error=str(send_err))

                    # 4. 标记为已发送
                    if sent_count > 0:
                        await daily_report_service.mark_as_sent(report.id)

                    logger.info(
                        "营业日报生成并发送完成", store_id=str(store.id), report_date=str(target_date), sent_count=sent_count
                    )
                    total_sent += sent_count

                except Exception as store_err:
                    logger.error("门店日报生成失败", store_id=str(store.id), error=str(store_err))
                    continue

            logger.info(
                "所有门店营业日报生成完成", stores_processed=len(stores), total_sent=total_sent, report_date=str(target_date)
            )

            return {
                "success": True,
                "stores_processed": len(stores),
                "total_sent": total_sent,
                "report_date": str(target_date),
            }

        except Exception as e:
            logger.error("生成营业日报失败", store_id=store_id, error=str(e), exc_info=e)
            raise self.retry(exc=e)

    return asyncio.run(_run())


# ── 周报生成（每周五 10:00 UTC / 北京时间 18:00）──────────────────────────────


@celery_app.task(bind=True, max_retries=2, default_retry_delay=300)
def generate_and_send_weekly_report(self) -> Dict[str, Any]:
    """
    为所有活跃门店生成周报（汇总本周 7 天 DailyReport）。

    周报内容：本周营收/订单/客流汇总 + 周环比 + 每日趋势。
    通过企微推送给店长和管理员。

    调度：每周五 10:00 UTC（beat_schedule 配置）
    """
    import asyncio

    async def _run():
        from src.services.weekly_report_service import WeeklyReportService

        svc = WeeklyReportService()
        result = await svc.generate_all_stores()

        # 推送周报摘要到企微
        sent_count = 0
        try:
            from src.core.database import get_db_session
            from src.models.user import User, UserRole
            from src.services.wechat_work_message_service import wechat_work_message_service

            for report in result.get("reports", []):
                store_id = report["store_id"]
                summary = report.get("summary", "")
                if not summary:
                    continue

                async with get_db_session() as session:
                    from sqlalchemy import select

                    mgr_result = await session.execute(
                        select(User).where(
                            User.store_id == store_id,
                            User.is_active == True,
                            User.role.in_([UserRole.STORE_MANAGER, UserRole.ADMIN]),
                            User.wechat_user_id.isnot(None),
                        )
                    )
                    managers = mgr_result.scalars().all()

                for mgr in managers:
                    try:
                        await wechat_work_message_service.send_text_message(
                            user_id=mgr.wechat_user_id, content=summary,
                        )
                        sent_count += 1
                    except Exception:
                        pass  # 发送失败不阻塞
        except Exception as e:
            logger.warning("weekly_report.push_failed", error=str(e))

        logger.info(
            "weekly_report.all_done",
            stores=result.get("reports_generated", 0),
            errors=result.get("errors", 0),
            sent=sent_count,
        )
        return {
            "success": True,
            "stores_processed": result.get("reports_generated", 0),
            "errors": result.get("errors", 0),
            "sent_count": sent_count,
        }

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("weekly_report.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── 定时报表执行器（每 10 分钟扫描 ScheduledReport）─────────────────────────


@celery_app.task(bind=True, max_retries=2, default_retry_delay=60)
def execute_scheduled_reports(self) -> Dict[str, Any]:
    """
    扫描 scheduled_reports 表中 next_run_at <= NOW() 且 is_active=True 的记录，
    按模板生成报表文件并通过配置的渠道（email/system）投递。

    完成后更新 last_run_at 并计算 next_run_at。
    每次最多处理 50 条，防止单次执行过长。
    """
    import asyncio

    async def _run() -> Dict[str, Any]:
        from datetime import datetime

        from sqlalchemy import select, text as sa_text

        from src.core.database import get_db_session
        from src.models.report_template import ScheduledReport
        from src.services.custom_report_service import custom_report_service

        executed = 0
        emailed = 0
        errors = []

        async with get_db_session() as session:
            now_iso = datetime.utcnow().isoformat()
            stmt = (
                select(ScheduledReport)
                .where(
                    ScheduledReport.is_active.is_(True),
                    ScheduledReport.next_run_at.isnot(None),
                    ScheduledReport.next_run_at <= now_iso,
                )
                .limit(50)
            )
            result = await session.execute(stmt)
            due_reports = list(result.scalars().all())

            if not due_reports:
                return {"executed": 0, "emailed": 0, "total_due": 0}

            for sr in due_reports:
                try:
                    # 1. 生成报表文件
                    file_bytes, filename, media_type = await custom_report_service.generate_report(
                        template_id=str(sr.template_id),
                        user_id=str(sr.user_id),
                        fmt=sr.format,
                    )

                    # 2. 按渠道投递
                    channels = sr.channels or []
                    recipients = sr.recipients or []

                    if "email" in channels and recipients:
                        from src.services.multi_channel_notification import (
                            EmailNotificationHandler,
                        )

                        handler = EmailNotificationHandler()
                        for email_addr in recipients:
                            try:
                                ok = await handler.send(
                                    recipient=email_addr,
                                    title=f"定时报表: {filename}",
                                    content=f"您订阅的定时报表已生成，请查看附件。\n文件: {filename}",
                                    extra_data={
                                        "attachments": [
                                            {"filename": filename, "data": file_bytes},
                                        ],
                                    },
                                )
                                if ok:
                                    emailed += 1
                            except Exception as e:
                                logger.warning(
                                    "scheduled_report.email_failed",
                                    recipient=email_addr,
                                    error=str(e)[:100],
                                )

                    # 3. 更新 last_run_at + 计算 next_run_at
                    sr.last_run_at = datetime.utcnow().isoformat()
                    sr.next_run_at = custom_report_service._calc_next_run(
                        sr.frequency, sr.run_at, sr.day_of_week, sr.day_of_month,
                    )
                    executed += 1

                except Exception as e:
                    errors.append({"scheduled_id": str(sr.id), "error": str(e)[:100]})
                    logger.warning(
                        "scheduled_report.exec_failed",
                        scheduled_id=str(sr.id),
                        error=str(e),
                    )

            await session.commit()

        logger.info(
            "execute_scheduled_reports.done",
            total_due=len(due_reports),
            executed=executed,
            emailed=emailed,
            errors=len(errors),
        )
        return {
            "total_due": len(due_reports),
            "executed": executed,
            "emailed": emailed,
            "errors": errors[:5],
        }

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("execute_scheduled_reports.failed", error=str(exc))
        raise self.retry(exc=exc)


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY_LONG", "300")),  # 5分钟
)
def perform_daily_reconciliation(
    self,
    store_id: str,
    reconciliation_date: str = None,
) -> Dict[str, Any]:
    """
    执行每日对账

    Args:
        store_id: 门店ID
        reconciliation_date: 对账日期（YYYY-MM-DD格式，默认为昨天）

    Returns:
        对账结果
    """

    async def _run():
        try:
            from datetime import date, datetime

            from ..services.reconcile_service import reconcile_service

            logger.info("开始执行每日对账", store_id=store_id, reconciliation_date=reconciliation_date)

            # 解析日期
            if reconciliation_date:
                target_date = datetime.strptime(reconciliation_date, "%Y-%m-%d").date()
            else:
                from datetime import timedelta

                target_date = date.today() - timedelta(days=1)

            # 执行对账
            record = await reconcile_service.perform_reconciliation(store_id=store_id, reconciliation_date=target_date)

            logger.info(
                "每日对账完成",
                store_id=store_id,
                reconciliation_date=str(target_date),
                status=record.status.value,
                diff_ratio=record.diff_ratio,
            )

            return {
                "success": True,
                "store_id": store_id,
                "reconciliation_date": str(target_date),
                "record_id": str(record.id),
                "status": record.status.value,
                "diff_ratio": record.diff_ratio,
                "alert_sent": record.alert_sent,
            }

        except Exception as e:
            logger.error("执行每日对账失败", store_id=store_id, error=str(e), exc_info=e)
            raise self.retry(exc=e)

    return asyncio.run(_run())


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY", "60")),
)
def detect_revenue_anomaly(
    self,
    store_id: str = None,
) -> Dict[str, Any]:
    """
    检测营收异常 (每15分钟执行)

    Args:
        store_id: 门店ID (None表示检测所有门店)

    Returns:
        检测结果
    """

    async def _run():
        try:
            from datetime import date, datetime, timedelta

            from sqlalchemy import func, select

            from ..agents.decision_agent import DecisionAgent
            from ..core.database import get_db_session
            from ..models.order import Order, OrderStatus
            from ..models.store import Store
            from ..models.user import User, UserRole
            from ..services.wechat_alert_service import wechat_alert_service

            logger.info("开始检测营收异常", store_id=store_id)

            decision_agent = DecisionAgent()
            alerts_sent = 0

            # 获取要检测的门店列表
            async with get_db_session() as session:
                if store_id:
                    result = await session.execute(select(Store).where(Store.id == store_id, Store.is_active == True))
                    stores = result.scalars().all()
                else:
                    result = await session.execute(select(Store).where(Store.is_active == True))
                    stores = result.scalars().all()

                for store in stores:
                    try:
                        now = datetime.now()
                        today_start = datetime.combine(date.today(), datetime.min.time())

                        # 当前营收：今天到目前为止已完成/已上菜的订单
                        rev_result = await session.execute(
                            select(func.coalesce(func.sum(Order.final_amount), 0)).where(
                                Order.store_id == store.id,
                                Order.order_time >= today_start,
                                Order.order_time <= now,
                                Order.status.in_([OrderStatus.COMPLETED, OrderStatus.SERVED]),
                            )
                        )
                        current_revenue = float(rev_result.scalar() or 0) / 100

                        # 预期营收：过去4周同星期同时段的平均值
                        current_elapsed = timedelta(hours=now.hour, minutes=now.minute)
                        expected_samples = []
                        for weeks_ago in range(1, 5):
                            past_date = date.today() - timedelta(weeks=weeks_ago)
                            past_start = datetime.combine(past_date, datetime.min.time())
                            past_end = past_start + current_elapsed
                            past_rev = await session.execute(
                                select(func.coalesce(func.sum(Order.final_amount), 0)).where(
                                    Order.store_id == store.id,
                                    Order.order_time >= past_start,
                                    Order.order_time <= past_end,
                                    Order.status.in_([OrderStatus.COMPLETED, OrderStatus.SERVED]),
                                )
                            )
                            val = float(past_rev.scalar() or 0) / 100
                            if val > 0:
                                expected_samples.append(val)

                        if not expected_samples:
                            # 无历史数据，跳过本门店
                            logger.debug("无历史营收数据，跳过", store_id=str(store.id))
                            continue

                        expected_revenue = sum(expected_samples) / len(expected_samples)

                        # 计算偏差
                        deviation = ((current_revenue - expected_revenue) / expected_revenue) * 100

                        # 只有偏差超过阈值才告警
                        if abs(deviation) > float(os.getenv("REVENUE_ANOMALY_THRESHOLD_PERCENT", "15")):
                            # 使用DecisionAgent分析
                            analysis = await decision_agent.analyze_revenue_anomaly(
                                store_id=str(store.id),
                                current_revenue=current_revenue,
                                expected_revenue=expected_revenue,
                                time_period="today",
                            )

                            if analysis["success"]:
                                # 查询店长和管理员的企微ID
                                user_result = await session.execute(
                                    select(User).where(
                                        User.store_id == store.id,
                                        User.is_active == True,
                                        User.role.in_([UserRole.STORE_MANAGER, UserRole.ADMIN]),
                                        User.wechat_user_id.isnot(None),
                                    )
                                )
                                managers = user_result.scalars().all()
                                recipient_ids = [m.wechat_user_id for m in managers]

                                if recipient_ids:
                                    # 使用WeChatAlertService发送告警
                                    alert_result = await wechat_alert_service.send_revenue_alert(
                                        store_id=str(store.id),
                                        store_name=store.name,
                                        current_revenue=current_revenue,
                                        expected_revenue=expected_revenue,
                                        deviation=deviation,
                                        analysis=analysis["data"]["analysis"],
                                        recipient_ids=recipient_ids,
                                    )

                                    if alert_result.get("success"):
                                        alerts_sent += alert_result.get("sent_count", 0)
                                        logger.info(
                                            "营收异常告警已发送",
                                            store_id=str(store.id),
                                            deviation=deviation,
                                            sent_count=alert_result.get("sent_count"),
                                        )
                                else:
                                    logger.warning("无可用接收人", store_id=str(store.id))

                    except Exception as e:
                        logger.error("门店营收异常检测失败", store_id=str(store.id), error=str(e))
                        continue

            logger.info("营收异常检测完成", stores_checked=len(stores), alerts_sent=alerts_sent)

            return {
                "success": True,
                "stores_checked": len(stores),
                "alerts_sent": alerts_sent,
                "timestamp": datetime.now().isoformat(),
            }

        except Exception as e:
            logger.error("营收异常检测失败", error=str(e), exc_info=e)
            raise self.retry(exc=e)

    return asyncio.run(_run())


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY_LONG", "300")),
)
def generate_daily_report_with_rag(
    self,
    store_id: str = None,
) -> Dict[str, Any]:
    """
    生成并发送昨日简报 (RAG增强版，每天6AM执行)

    Args:
        store_id: 门店ID (None表示为所有门店生成)

    Returns:
        生成结果
    """

    async def _run():
        try:
            from datetime import date, datetime, timedelta

            from sqlalchemy import select

            from ..agents.decision_agent import DecisionAgent
            from ..core.database import get_db_session
            from ..models.store import Store
            from ..services.wechat_work_message_service import wechat_work_message_service

            logger.info("开始生成昨日简报(RAG增强)", store_id=store_id)

            decision_agent = DecisionAgent()
            reports_sent = 0
            yesterday = date.today() - timedelta(days=1)

            # 获取要生成报告的门店列表
            async with get_db_session() as session:
                if store_id:
                    result = await session.execute(select(Store).where(Store.id == store_id, Store.is_active == True))
                    stores = result.scalars().all()
                else:
                    result = await session.execute(select(Store).where(Store.is_active == True))
                    stores = result.scalars().all()

                for store in stores:
                    try:
                        # 使用DecisionAgent生成经营建议
                        recommendations = await decision_agent.generate_business_recommendations(
                            store_id=str(store.id), focus_area=None  # 全面分析
                        )

                        if recommendations["success"]:
                            # 构建简报消息
                            message = f"""📊 昨日简报 {yesterday.strftime('%Y年%m月%d日')}

    门店: {store.name}

    AI经营分析:
    {recommendations['data']['recommendations']}

    ---
    基于{recommendations['data']['context_used']}条历史数据分析
    生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}
    """

                            # 查询店长和管理员的企微ID并发送
                            from ..models.user import User, UserRole

                            user_result = await session.execute(
                                select(User).where(
                                    User.store_id == store.id,
                                    User.is_active == True,
                                    User.role.in_([UserRole.STORE_MANAGER, UserRole.ADMIN]),
                                    User.wechat_user_id.isnot(None),
                                )
                            )
                            managers = user_result.scalars().all()
                            sent_count = 0
                            for manager in managers:
                                try:
                                    send_result = await wechat_work_message_service.send_text_message(
                                        user_id=manager.wechat_user_id, content=message
                                    )
                                    if send_result.get("success"):
                                        sent_count += 1
                                except Exception as send_err:
                                    logger.error("发送简报失败", user_id=str(manager.id), error=str(send_err))

                            logger.info("昨日简报已生成并发送", store_id=str(store.id), sent_count=sent_count)
                            reports_sent += sent_count

                    except Exception as e:
                        logger.error("门店简报生成失败", store_id=str(store.id), error=str(e))
                        continue

            logger.info("昨日简报生成完成", stores_processed=len(stores), reports_sent=reports_sent)

            return {
                "success": True,
                "stores_processed": len(stores),
                "reports_sent": reports_sent,
                "report_date": str(yesterday),
                "timestamp": datetime.now().isoformat(),
            }

        except Exception as e:
            logger.error("昨日简报生成失败", error=str(e), exc_info=e)
            raise self.retry(exc=e)

    return asyncio.run(_run())


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY", "60")),
)
def check_inventory_alert(
    self,
    store_id: str = None,
) -> Dict[str, Any]:
    """
    检查库存预警 (午高峰前1小时，每天10AM执行)

    Args:
        store_id: 门店ID (None表示检查所有门店)

    Returns:
        检查结果
    """

    async def _run():
        try:
            from datetime import datetime

            from sqlalchemy import select

            from ..agents.inventory_agent import InventoryAgent
            from ..core.database import get_db_session
            from ..models.inventory import InventoryItem, InventoryStatus
            from ..models.store import Store
            from ..models.user import User, UserRole
            from ..services.wechat_alert_service import wechat_alert_service

            logger.info("开始检查库存预警", store_id=store_id)

            inventory_agent = InventoryAgent()
            alerts_sent = 0

            # 获取要检查的门店列表
            async with get_db_session() as session:
                if store_id:
                    result = await session.execute(select(Store).where(Store.id == store_id, Store.is_active == True))
                    stores = result.scalars().all()
                else:
                    result = await session.execute(select(Store).where(Store.is_active == True))
                    stores = result.scalars().all()

                for store in stores:
                    try:
                        # 从数据库查询低库存/缺货库存项
                        inv_result = await session.execute(
                            select(InventoryItem).where(
                                InventoryItem.store_id == store.id,
                                InventoryItem.status.in_(
                                    [
                                        InventoryStatus.LOW,
                                        InventoryStatus.CRITICAL,
                                        InventoryStatus.OUT_OF_STOCK,
                                    ]
                                ),
                            )
                        )
                        low_stock_items = inv_result.scalars().all()

                        if not low_stock_items:
                            logger.debug("无库存预警项", store_id=str(store.id))
                            continue

                        # 构建 InventoryAgent 所需的 current_inventory 字典
                        current_inventory = {item.id: item.current_quantity for item in low_stock_items}

                        # 使用InventoryAgent检查低库存
                        alert_result = await inventory_agent.check_low_stock_alert(
                            store_id=str(store.id),
                            current_inventory=current_inventory,
                            threshold_hours=int(os.getenv("INVENTORY_ALERT_THRESHOLD_HOURS", "4")),  # 午高峰前N小时预警
                        )

                        if alert_result["success"]:
                            # 构建预警项目列表（来自真实数据）
                            alert_items = [
                                {
                                    "dish_name": item.name,
                                    "quantity": item.current_quantity,
                                    "unit": item.unit or "",
                                    "min_quantity": item.min_quantity,
                                    "risk": (
                                        "high"
                                        if item.status in (InventoryStatus.CRITICAL, InventoryStatus.OUT_OF_STOCK)
                                        else "medium"
                                    ),
                                }
                                for item in low_stock_items
                            ]

                            # 查询店长和管理员的企微ID
                            user_result = await session.execute(
                                select(User).where(
                                    User.store_id == store.id,
                                    User.is_active == True,
                                    User.role.in_([UserRole.STORE_MANAGER, UserRole.ADMIN]),
                                    User.wechat_user_id.isnot(None),
                                )
                            )
                            managers = user_result.scalars().all()
                            recipient_ids = [m.wechat_user_id for m in managers]

                            if recipient_ids:
                                # 使用WeChatAlertService发送预警
                                send_result = await wechat_alert_service.send_inventory_alert(
                                    store_id=str(store.id),
                                    store_name=store.name,
                                    alert_items=alert_items,
                                    analysis=alert_result["data"]["alert"],
                                    recipient_ids=recipient_ids,
                                )

                                if send_result.get("success"):
                                    alerts_sent += send_result.get("sent_count", 0)
                                    logger.info(
                                        "库存预警已发送", store_id=str(store.id), sent_count=send_result.get("sent_count")
                                    )
                            else:
                                logger.warning("无可用接收人", store_id=str(store.id))

                    except Exception as e:
                        logger.error("门店库存检查失败", store_id=str(store.id), error=str(e))
                        continue

            logger.info("库存预警检查完成", stores_checked=len(stores), alerts_sent=alerts_sent)

            return {
                "success": True,
                "stores_checked": len(stores),
                "alerts_sent": alerts_sent,
                "timestamp": datetime.now().isoformat(),
            }

        except Exception as e:
            logger.error("库存预警检查失败", error=str(e), exc_info=e)
            raise self.retry(exc=e)

    return asyncio.run(_run())


# ------------------------------------------------------------------ #
# 大数据异步导出任务                                                    #
# ------------------------------------------------------------------ #


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="async_export_data",
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY", "60")),
)
def async_export_data(self, job_id: str) -> Dict[str, Any]:
    """
    异步大数据导出任务

    从数据库分批读取数据，生成 CSV/Excel 文件，
    并将结果写入临时目录，更新 ExportJob 状态。
    """
    import csv
    import tempfile
    from datetime import date, datetime

    async def _run():
        from sqlalchemy import and_, select
        from src.core.database import AsyncSessionLocal
        from src.models.export_job import ExportJob, ExportStatus

        BATCH_SIZE = int(os.getenv("EXPORT_BATCH_SIZE", "1000"))

        async with AsyncSessionLocal() as session:
            job = await session.get(ExportJob, job_id)
            if not job:
                logger.error("导出任务不存在", job_id=job_id)
                return {"success": False, "error": "job not found"}
            job.status = ExportStatus.RUNNING
            job.celery_task_id = self.request.id
            await session.commit()
            job_type = job.job_type
            fmt = job.format
            params = job.params or {}

        try:
            rows, headers = await _fetch_export_data(job_type, params)

            total = len(rows)
            tmp_dir = os.getenv("EXPORT_TMP_DIR", tempfile.gettempdir())
            os.makedirs(tmp_dir, exist_ok=True)
            filename = f"export_{job_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{fmt}"
            file_path = os.path.join(tmp_dir, filename)

            if fmt == "csv":
                file_size = _write_csv(file_path, headers, rows)
            elif fmt == "xlsx":
                file_size = _write_xlsx(file_path, headers, rows)
            else:
                raise ValueError(f"不支持的格式: {fmt}")

            async with AsyncSessionLocal() as session:
                job = await session.get(ExportJob, job_id)
                if job:
                    job.status = ExportStatus.COMPLETED
                    job.progress = 100
                    job.total_rows = total
                    job.processed_rows = total
                    job.file_path = file_path
                    job.file_size_bytes = file_size
                    job.completed_at = datetime.utcnow().isoformat()
                    await session.commit()

            logger.info("导出任务完成", job_id=job_id, total_rows=total)
            return {"success": True, "job_id": job_id, "total_rows": total}

        except Exception as e:
            logger.error("导出任务失败", job_id=job_id, error=str(e))
            async with AsyncSessionLocal() as session:
                job = await session.get(ExportJob, job_id)
                if job:
                    job.status = ExportStatus.FAILED
                    job.error_message = str(e)
                    await session.commit()
            raise self.retry(exc=e)

    async def _fetch_export_data(job_type: str, params: Dict):
        from datetime import date, datetime

        from sqlalchemy import and_, select
        from src.core.database import AsyncSessionLocal

        if job_type == "transactions":
            from src.models.finance import FinancialTransaction

            headers = ["日期", "类型", "分类", "子分类", "金额(元)", "描述", "支付方式", "门店ID"]
            async with AsyncSessionLocal() as session:
                conditions = []
                if params.get("store_id"):
                    conditions.append(FinancialTransaction.store_id == params["store_id"])
                if params.get("transaction_type"):
                    conditions.append(FinancialTransaction.transaction_type == params["transaction_type"])
                if params.get("start_date"):
                    conditions.append(FinancialTransaction.transaction_date >= date.fromisoformat(params["start_date"]))
                if params.get("end_date"):
                    conditions.append(FinancialTransaction.transaction_date <= date.fromisoformat(params["end_date"]))
                stmt = select(FinancialTransaction)
                if conditions:
                    stmt = stmt.where(and_(*conditions))
                stmt = stmt.order_by(FinancialTransaction.transaction_date.desc())
                result = await session.execute(stmt)
                rows = [
                    [
                        t.transaction_date.isoformat() if t.transaction_date else "",
                        t.transaction_type or "",
                        t.category or "",
                        t.subcategory or "",
                        round((t.amount or 0) / 100, 2),
                        t.description or "",
                        t.payment_method or "",
                        t.store_id or "",
                    ]
                    for t in result.scalars().all()
                ]
            return rows, headers

        elif job_type == "audit_logs":
            from src.models.audit_log import AuditLog

            headers = ["时间", "用户ID", "用户名", "角色", "操作", "资源类型", "资源ID", "描述", "IP", "状态", "门店ID"]
            async with AsyncSessionLocal() as session:
                conditions = []
                if params.get("user_id"):
                    conditions.append(AuditLog.user_id == params["user_id"])
                if params.get("action"):
                    conditions.append(AuditLog.action == params["action"])
                if params.get("store_id"):
                    conditions.append(AuditLog.store_id == params["store_id"])
                if params.get("start_date"):
                    conditions.append(AuditLog.created_at >= datetime.fromisoformat(params["start_date"]))
                if params.get("end_date"):
                    conditions.append(AuditLog.created_at <= datetime.fromisoformat(params["end_date"]))
                stmt = select(AuditLog)
                if conditions:
                    stmt = stmt.where(and_(*conditions))
                stmt = stmt.order_by(AuditLog.created_at.desc())
                result = await session.execute(stmt)
                rows = [
                    [
                        log.created_at.isoformat() if log.created_at else "",
                        str(log.user_id) if log.user_id else "",
                        log.username or "",
                        log.user_role or "",
                        log.action or "",
                        log.resource_type or "",
                        str(log.resource_id) if log.resource_id else "",
                        log.description or "",
                        log.ip_address or "",
                        log.status or "",
                        str(log.store_id) if log.store_id else "",
                    ]
                    for log in result.scalars().all()
                ]
            return rows, headers

        elif job_type == "orders":
            from src.models.order import Order

            headers = ["订单号", "状态", "总金额(元)", "桌号", "门店ID", "下单时间"]
            async with AsyncSessionLocal() as session:
                conditions = []
                if params.get("store_id"):
                    conditions.append(Order.store_id == params["store_id"])
                if params.get("status"):
                    conditions.append(Order.status == params["status"])
                if params.get("start_date"):
                    conditions.append(Order.created_at >= datetime.fromisoformat(params["start_date"]))
                if params.get("end_date"):
                    conditions.append(Order.created_at <= datetime.fromisoformat(params["end_date"]))
                stmt = select(Order)
                if conditions:
                    stmt = stmt.where(and_(*conditions))
                stmt = stmt.order_by(Order.created_at.desc())
                result = await session.execute(stmt)
                rows = [
                    [
                        o.order_number or str(o.id),
                        o.status.value if hasattr(o.status, "value") else str(o.status or ""),
                        round((o.total_amount or 0) / 100, 2),
                        o.table_number or "",
                        o.store_id or "",
                        o.created_at.isoformat() if o.created_at else "",
                    ]
                    for o in result.scalars().all()
                ]
            return rows, headers

        else:
            raise ValueError(f"不支持的导出类型: {job_type}，可选: transactions/audit_logs/orders")

    def _write_csv(file_path: str, headers: list, rows: list) -> int:
        with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(rows)
        return os.path.getsize(file_path)

    def _write_xlsx(file_path: str, headers: list, rows: list) -> int:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("请安装 openpyxl: pip install openpyxl")
        wb = openpyxl.Workbook()
        ws = wb.active
        hf = Font(bold=True, color="FFFFFF")
        hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hf
            c.fill = hfill
            ws.column_dimensions[get_column_letter(ci)].width = 16
        for ri, row in enumerate(rows, 2):
            for ci, v in enumerate(row, 1):
                ws.cell(row=ri, column=ci, value=v)
        wb.save(file_path)
        return os.path.getsize(file_path)

    return asyncio.run(_run())


# ---------------------------------------------------------------------------
# 增量备份任务
# ---------------------------------------------------------------------------


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="run_backup",
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY", "60")),
)
def run_backup(self, job_id: str) -> Dict[str, Any]:
    """
    执行全量/增量备份任务
    - 全量：导出所有指定表的数据为 JSON，打包成 tar.gz
    - 增量：仅导出 since_timestamp 之后有变更的行（依赖 updated_at 字段）
    """
    import hashlib
    import json
    import tarfile
    import tempfile
    from datetime import datetime, timezone

    async def _run():
        from sqlalchemy import text
        from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
        from sqlalchemy.orm import sessionmaker

        db_url = os.environ["DATABASE_URL"]
        backup_dir = os.getenv("BACKUP_TMP_DIR", "/tmp/backups")
        os.makedirs(backup_dir, exist_ok=True)

        engine = create_async_engine(db_url, echo=False)
        async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

        async with async_session() as session:
            # 读取 BackupJob
            from src.models.backup_job import BackupJob, BackupStatus

            result = await session.execute(
                text("SELECT * FROM backup_jobs WHERE id = :id"),
                {"id": job_id},
            )
            row = result.mappings().first()
            if not row:
                raise ValueError(f"BackupJob {job_id} 不存在")

            backup_type = row["backup_type"]
            since_ts = row["since_timestamp"]
            tables_filter = row["tables"] or []

            # 标记 RUNNING
            await session.execute(
                text("UPDATE backup_jobs SET status='running', celery_task_id=:tid, updated_at=NOW() WHERE id=:id"),
                {"tid": self.request.id, "id": job_id},
            )
            await session.commit()

        # 获取所有用户表
        async with async_session() as session:
            res = await session.execute(text("SELECT tablename FROM pg_tables WHERE schemaname='public' ORDER BY tablename"))
            all_tables = [r[0] for r in res.fetchall()]

        target_tables = [t for t in all_tables if not tables_filter or t in tables_filter]
        # 排除备份相关表，避免递归
        target_tables = [t for t in target_tables if t not in ("backup_jobs", "export_jobs")]

        total = len(target_tables)
        row_counts: Dict[str, int] = {}
        tmp_dir = tempfile.mkdtemp(dir=backup_dir)

        try:
            for idx, table in enumerate(target_tables):
                async with async_session() as session:
                    # 表名来自 pg_tables（系统目录）+ tables_filter（用户可控），
                    # 必须通过白名单校验后才能拼入 SQL（表名无法参数化）
                    if not _SAFE_TABLE_RE.match(table):
                        logger.warning("backup.skip_unsafe_table", table=table)
                        row_counts[table] = 0
                        continue
                    if backup_type == "incremental" and since_ts:
                        # 增量：只取 updated_at > since_timestamp 的行
                        try:
                            res = await session.execute(
                                text(f"SELECT * FROM {table} WHERE updated_at > :ts"),
                                {"ts": since_ts},
                            )
                        except Exception:
                            # 表没有 updated_at 字段时跳过
                            row_counts[table] = 0
                            continue
                    else:
                        res = await session.execute(text(f"SELECT * FROM {table}"))

                    cols = list(res.keys())
                    rows_data = [dict(zip(cols, r)) for r in res.fetchall()]

                    # 序列化（UUID/datetime 转字符串）
                    def _serialize(v):
                        if hasattr(v, "isoformat"):
                            return v.isoformat()
                        if hasattr(v, "__str__") and not isinstance(v, (int, float, bool, str, type(None))):
                            return str(v)
                        return v

                    rows_data = [{k: _serialize(v) for k, v in r.items()} for r in rows_data]
                    row_counts[table] = len(rows_data)

                    table_file = os.path.join(tmp_dir, f"{table}.json")
                    with open(table_file, "w", encoding="utf-8") as f:
                        json.dump({"table": table, "rows": rows_data}, f, ensure_ascii=False, indent=2)

                # 更新进度
                progress = int((idx + 1) / total * 90)
                async with async_session() as session:
                    await session.execute(
                        text("UPDATE backup_jobs SET progress=:p, updated_at=NOW() WHERE id=:id"),
                        {"p": progress, "id": job_id},
                    )
                    await session.commit()

            # 打包 tar.gz
            ts_str = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
            archive_name = f"backup_{backup_type}_{ts_str}_{job_id[:8]}.tar.gz"
            archive_path = os.path.join(backup_dir, archive_name)
            with tarfile.open(archive_path, "w:gz") as tar:
                tar.add(tmp_dir, arcname="backup")

            # 计算 SHA256
            sha256 = hashlib.sha256()
            with open(archive_path, "rb") as f:
                for chunk in iter(lambda: f.read(65536), b""):
                    sha256.update(chunk)
            checksum = sha256.hexdigest()
            file_size = os.path.getsize(archive_path)
            completed_at = datetime.now(timezone.utc).isoformat()

            async with async_session() as session:
                await session.execute(
                    text(
                        "UPDATE backup_jobs SET status='completed', progress=100, "
                        "file_path=:fp, file_size_bytes=:fs, checksum=:cs, "
                        "row_counts=:rc, completed_at=:ca, updated_at=NOW() WHERE id=:id"
                    ),
                    {
                        "fp": archive_path,
                        "fs": file_size,
                        "cs": checksum,
                        "rc": json.dumps(row_counts),
                        "ca": completed_at,
                        "id": job_id,
                    },
                )
                await session.commit()

            logger.info("备份任务完成", job_id=job_id, archive=archive_path, checksum=checksum)
            return {"job_id": job_id, "file_path": archive_path, "checksum": checksum}

        except Exception as e:
            logger.error("备份任务失败", job_id=job_id, error=str(e))
            async with async_session() as session:
                await session.execute(
                    text("UPDATE backup_jobs SET status='failed', error_message=:err, updated_at=NOW() WHERE id=:id"),
                    {"err": str(e)[:1000], "id": job_id},
                )
                await session.commit()
            raise self.retry(exc=e)

        finally:
            import shutil

            shutil.rmtree(tmp_dir, ignore_errors=True)

    return asyncio.run(_run())


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY_LONG", "300")),
)
def generate_daily_hub(
    self,
    store_id: str = None,
) -> Dict[str, Any]:
    """
    生成 T+1 经营统筹备战板

    Args:
        store_id: 门店ID (None 表示为所有活跃门店生成)

    Returns:
        生成结果
    """

    async def _run():
        from datetime import date, timedelta

        from sqlalchemy import select

        from ..core.database import get_db_session
        from ..models.store import Store
        from ..services.daily_hub_service import daily_hub_service

        target_date = date.today() + timedelta(days=1)

        async with get_db_session() as session:
            if store_id:
                result = await session.execute(select(Store).where(Store.id == store_id, Store.is_active == True))
            else:
                result = await session.execute(select(Store).where(Store.is_active == True))
            stores = result.scalars().all()

        generated = 0
        for store in stores:
            try:
                await daily_hub_service.generate_battle_board(store_id=str(store.id), target_date=target_date)
                generated += 1
                logger.info("备战板生成成功", store_id=str(store.id), target_date=str(target_date))
            except Exception as e:
                logger.error("备战板生成失败", store_id=str(store.id), error=str(e))

        return {"success": True, "generated": generated, "target_date": str(target_date)}

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


# ═══════════════════════════════════════════════════════════════════════════════
# Phase 3 — 损耗推理 / 规则评估 / 本体日同步
# ═══════════════════════════════════════════════════════════════════════════════


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.process_waste_event",
    max_retries=3,
    default_retry_delay=30,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=300,
    retry_jitter=True,
)
def process_waste_event(
    self,
    event_id: str,
) -> Dict[str, Any]:
    """
    损耗事件五步推理（异步，由 WasteEventService._enqueue_analysis 投递）

    步骤：
      1. 从 PostgreSQL 加载 WasteEvent
      2. 调用 WasteReasoningEngine.infer_root_cause(event_id)
      3. 将根因 / 置信度 / 证据链回写 PostgreSQL
      4. 同步推理结论到 Neo4j

    Args:
        event_id: WasteEvent.event_id（格式 WE-XXXXXXXX）

    Returns:
        {"success": True, "event_id": ..., "root_cause": ..., "confidence": ...}
    """

    async def _run():
        from ..core.database import get_db_session
        from ..services.waste_event_service import WasteEventService

        async with get_db_session() as session:
            svc = WasteEventService(session)

            # 标记推理中
            from sqlalchemy import update as _update

            from ..models.waste_event import WasteEvent, WasteEventStatus

            await session.execute(
                _update(WasteEvent).where(WasteEvent.event_id == event_id).values(status=WasteEventStatus.ANALYZING)
            )
            await session.commit()

        # 调用推理引擎（同步驱动，独立 session）
        try:
            from ..ontology.reasoning import WasteReasoningEngine

            engine = WasteReasoningEngine()
            result = engine.infer_root_cause(event_id)
        except Exception as e:
            logger.warning("推理引擎调用失败", event_id=event_id, error=str(e))
            return {"success": False, "event_id": event_id, "error": str(e)}

        if not result.get("success"):
            return {"success": False, "event_id": event_id, "error": result.get("error")}

        # 写回分析结论
        async with get_db_session() as session:
            svc = WasteEventService(session)
            await svc.write_back_analysis(
                event_id=event_id,
                root_cause=result.get("root_cause", "unknown"),
                confidence=result.get("confidence", 0.0),
                evidence=result.get("evidence_chain", {}),
                scores=result.get("scores", {}),
            )
            # 同步到 Neo4j
            ev = await svc.get_event(event_id)
            if ev:
                await svc._sync_analysis_to_neo4j(ev)
            await session.commit()

        logger.info(
            "损耗推理任务完成",
            event_id=event_id,
            root_cause=result.get("root_cause"),
            confidence=result.get("confidence"),
        )
        return {
            "success": True,
            "event_id": event_id,
            "root_cause": result.get("root_cause"),
            "confidence": result.get("confidence"),
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


# ═══════════════════════════════════════════════════════════════════════════════
# L3 — 跨店知识聚合夜间物化（凌晨 2:30 触发）
# ═══════════════════════════════════════════════════════════════════════════════


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.nightly_cross_store_sync",
    max_retries=2,
    default_retry_delay=300,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=1800,
    retry_jitter=False,
)
def nightly_cross_store_sync(
    self,
    store_ids: list = None,
) -> Dict[str, Any]:
    """
    L3 跨店知识聚合夜间物化任务（建议凌晨 2:30 触发）

    执行步骤：
      1. 计算两两门店相似度矩阵，写入 store_similarity_cache
      2. 重建同伴组 store_peer_groups（tier + region 分组）
      3. 物化昨日 cross_store_metrics（6 项指标 × 全门店）
      4. 同步 Neo4j 图（Store 节点 + SIMILAR_TO / BENCHMARK_OF / SHARES_RECIPE 边）

    Args:
        store_ids: 指定门店列表（None = 全部活跃门店）

    Returns:
        {
          "similarity_pairs":  N,
          "peer_groups":       N,
          "metrics_upserted":  N,
          "graph_synced":      bool,
          "errors":            [...]
        }
    """

    async def _run():
        from ..core.database import get_db_session
        from ..services.cross_store_knowledge_service import CrossStoreKnowledgeService

        errors = []

        async with get_db_session() as session:
            svc = CrossStoreKnowledgeService(session)

            # Step 1: 相似度矩阵
            try:
                sim_result = await svc.compute_pairwise_similarity(store_ids=store_ids)
                similarity_pairs = sim_result.get("pairs_computed", 0)
                logger.info("跨店相似度矩阵已计算", pairs=similarity_pairs)
            except Exception as e:
                errors.append({"step": "similarity", "error": str(e)})
                similarity_pairs = 0
                logger.error("相似度矩阵计算失败", error=str(e))

            # Step 2: 同伴组重建
            try:
                pg_result = await svc.build_peer_groups(store_ids=store_ids)
                peer_groups = pg_result.get("groups_built", 0)
                logger.info("同伴组重建完成", groups=peer_groups)
            except Exception as e:
                errors.append({"step": "peer_groups", "error": str(e)})
                peer_groups = 0
                logger.error("同伴组重建失败", error=str(e))

            # Step 3: 日维度指标物化
            try:
                mat_result = await svc.materialize_metrics(store_ids=store_ids)
                metrics_upserted = mat_result.get("upserted", 0)
                logger.info("跨店指标物化完成", upserted=metrics_upserted)
            except Exception as e:
                errors.append({"step": "materialize", "error": str(e)})
                metrics_upserted = 0
                logger.error("指标物化失败", error=str(e))

            await session.commit()

            # Step 4: Neo4j 图同步（独立异常处理，不影响前序结果）
            graph_synced = False
            try:
                graph_result = await svc.sync_store_graph(store_ids=store_ids)
                graph_synced = not graph_result.get("skipped", False)
                logger.info("Neo4j 跨店图同步完成", result=graph_result)
            except Exception as e:
                errors.append({"step": "graph_sync", "error": str(e)})
                logger.error("Neo4j 跨店图同步失败", error=str(e))

        logger.info(
            "L3 跨店知识聚合夜间任务完成",
            similarity_pairs=similarity_pairs,
            peer_groups=peer_groups,
            metrics_upserted=metrics_upserted,
            graph_synced=graph_synced,
            errors=len(errors),
        )
        return {
            "success": len(errors) == 0,
            "similarity_pairs": similarity_pairs,
            "peer_groups": peer_groups,
            "metrics_upserted": metrics_upserted,
            "graph_synced": graph_synced,
            "errors": errors,
        }

    try:
        return asyncio.run(_run())
    except Exception as nightly_e:
        raise self.retry(exc=nightly_e)


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.evaluate_store_rules",
    max_retries=2,
    default_retry_delay=60,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=600,
    retry_jitter=True,
)
def evaluate_store_rules(
    self,
    store_id: str,
    kpi_context: Dict[str, Any],
    industry_type: str = "general",
) -> Dict[str, Any]:
    """
    对门店 KPI 上下文运行推理规则库，匹配触发规则并自动推送企微告警

    工作流程：
      1. 从规则库加载 ACTIVE 规则
      2. 对 kpi_context 执行规则匹配，获取 Top-10 匹配规则
      3. 对命中规则写入 RuleExecution 日志
      4. 置信度 >= 0.70 时，自动创建企微 P1/P2 Action
      5. 行业基准对比（若 industry_type 非 "general"）

    Args:
        store_id: 门店ID
        kpi_context: 当前 KPI 指标字典，如
            {"waste_rate": 0.18, "labor_cost_ratio": 0.36, ...}
        industry_type: 行业类型（seafood / hotpot / fastfood / general）

    Returns:
        {"matched": [...], "actions_created": N, "store_id": ...}
    """

    async def _run():
        from ..core.database import get_db_session
        from ..services.knowledge_rule_service import KnowledgeRuleService

        matched_rules = []
        actions_created = 0

        async with get_db_session() as session:
            rule_svc = KnowledgeRuleService(session)

            # 规则匹配（全品类）
            matched = await rule_svc.match_rules(kpi_context)

            for hit in matched:
                matched_rules.append(hit)

                # 写入执行日志
                rule_obj = await rule_svc.get_by_code(hit["rule_code"])
                if rule_obj:
                    await rule_svc.log_execution(
                        rule=rule_obj,
                        store_id=store_id,
                        event_id=None,
                        condition_values=kpi_context,
                        conclusion_output=hit.get("conclusion", {}),
                        confidence_score=hit["confidence"],
                    )

                # 置信度 ≥ 0.70 → 推送企微告警
                if hit["confidence"] >= 0.70:
                    try:
                        from ..services.wechat_action_fsm import ActionCategory, ActionPriority, get_wechat_fsm

                        fsm = get_wechat_fsm()
                        priority = ActionPriority.P1 if hit["confidence"] >= 0.80 else ActionPriority.P2
                        conclusion = hit.get("conclusion", {})
                        action_text = conclusion.get("action") or conclusion.get("conclusion", "请检查相关指标")
                        await fsm.create_action(
                            store_id=store_id,
                            category=ActionCategory.KPI_ALERT,
                            priority=priority,
                            title=f"规则告警：{hit['rule_code']}",
                            content=(f"**{hit['name']}**\n" f"置信度：{hit['confidence']:.0%}\n" f"建议：{action_text}"),
                            receiver_user_id="store_manager",
                            source_event_id=f"RULE-{hit['rule_code']}-{store_id}",
                            evidence={"kpi_context": kpi_context, "rule_code": hit["rule_code"]},
                        )
                        actions_created += 1
                    except Exception as e:
                        logger.warning(
                            "规则触发企微告警失败",
                            store_id=store_id,
                            rule_code=hit["rule_code"],
                            error=str(e),
                        )

            # 行业基准对比（非 general 时）
            benchmark_summary = []
            if industry_type != "general":
                benchmark_results = await rule_svc.compare_to_benchmark(industry_type, kpi_context)
                for br in benchmark_results:
                    if br["percentile_band"] == "bottom_25":
                        benchmark_summary.append(br)

            await session.commit()

        logger.info(
            "门店规则评估完成",
            store_id=store_id,
            matched=len(matched_rules),
            actions_created=actions_created,
        )
        return {
            "success": True,
            "store_id": store_id,
            "matched": matched_rules,
            "actions_created": actions_created,
            "bottom_25_benchmarks": benchmark_summary,
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.daily_ontology_sync",
    max_retries=2,
    default_retry_delay=300,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=1800,
    retry_jitter=False,
)
def daily_ontology_sync(
    self,
    store_id: str = None,
) -> Dict[str, Any]:
    """
    每日全量 PostgreSQL → Neo4j 本体同步（建议凌晨 3:00 触发）

    范围：
      - 活跃 BOMTemplate 及明细行 → BOM 节点 + HAS_INGREDIENT 边
      - WasteEvent（最近 30 天）→ WasteEvent 节点 + WASTE_OF 边

    Args:
        store_id: 指定门店（None = 全部活跃门店）

    Returns:
        {"synced_stores": N, "nodes_upserted": N, "errors": [...]}
    """

    async def _run():
        from ..core.database import get_db_session
        from ..services.ontology_sync_service import sync_ontology_from_pg

        errors = []

        async with get_db_session() as session:
            try:
                # 使用统一同步入口（含 stores/dishes/ingredients/staff/orders/suppliers/boms/waste_events）
                tenant_id = store_id or "default"
                result = await sync_ontology_from_pg(session, tenant_id, store_id)
                nodes_upserted = sum(result.values())
                synced_stores = result.get("stores", 0)
            except Exception as e:
                errors.append({"store_id": store_id or "all", "error": str(e)})
                logger.error("本体同步失败", error=str(e))
                nodes_upserted = 0
                synced_stores = 0

        logger.info(
            "日常本体同步完成",
            synced_stores=synced_stores,
            nodes_upserted=nodes_upserted,
            errors=len(errors),
        )
        return {
            "success": True,
            "synced_stores": synced_stores,
            "nodes_upserted": nodes_upserted,
            "errors": errors,
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


# ═══════════════════════════════════════════════════════════════════════════════
# L4 — 全平台推理扫描夜间任务（凌晨 3:30 触发）
# ═══════════════════════════════════════════════════════════════════════════════


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.nightly_reasoning_scan",
    max_retries=2,
    default_retry_delay=300,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=1800,
    retry_jitter=False,
)
def nightly_reasoning_scan(
    self,
    store_ids: list = None,
) -> Dict[str, Any]:
    """
    L4 全平台推理扫描夜间任务（建议凌晨 3:30 触发，在 L3 nightly_cross_store_sync 之后）

    执行步骤：
      1. 从 PostgreSQL 拉取所有活跃门店最近 24h KPI 快照
      2. 调用 DiagnosisService.run_full_diagnosis() 全维度推理
      3. 结论写入 reasoning_reports（upsert，幂等）
      4. 对 P1/P2 报告同步写入 Neo4j ReasoningReport 节点
      5. 置信度 ≥ 0.70 的 P1/P2 告警自动推送企微

    Args:
        store_ids: 指定门店列表（None = 全部活跃门店）

    Returns:
        {
          "stores_scanned":  N,
          "p1_alerts":       N,
          "p2_alerts":       N,
          "wechat_sent":     N,
          "errors":          [...]
        }
    """

    async def _run():
        from sqlalchemy import and_, func, select

        from ..core.database import get_db_session
        from ..models.reasoning import ReasoningReport
        from ..models.store import Store
        from ..services.diagnosis_service import DiagnosisService
        from ..services.reasoning_engine import ALL_DIMENSIONS

        errors = []
        stores_scanned = 0
        p1_count = 0
        p2_count = 0
        wechat_sent = 0

        async with get_db_session() as session:
            # 拉取活跃门店列表
            if store_ids:
                stmt = select(Store).where(and_(Store.id.in_(store_ids), Store.is_active.is_(True)))
            else:
                stmt = select(Store).where(Store.is_active.is_(True))
            stores = (await session.execute(stmt)).scalars().all()

            svc = DiagnosisService(session)

            for store in stores:
                sid = str(store.id)
                try:
                    # 构造门店 KPI 快照（从近 30 天 cross_store_metrics 物化数据）
                    kpi_context = await _build_kpi_context(session, sid)
                    if not kpi_context:
                        logger.debug("门店无 KPI 数据，跳过推理", store_id=sid)
                        continue

                    # 全维度诊断
                    report = await svc.run_full_diagnosis(
                        store_id=sid,
                        kpi_context=kpi_context,
                    )
                    stores_scanned += 1

                    # 统计 P1/P2
                    for dim, c in report.dimensions.items():
                        if c.severity == "P1":
                            p1_count += 1
                        elif c.severity == "P2":
                            p2_count += 1

                    # P1/P2 → 同步 Neo4j + 企微推送
                    for dim, c in report.dimensions.items():
                        if c.severity in ("P1", "P2") and c.confidence >= 0.70:
                            # Neo4j 同步
                            try:
                                import uuid as _uuid
                                from datetime import date

                                # 查找刚写入的 reasoning_report id
                                from ..models.reasoning import ReasoningReport as RR
                                from ..ontology.data_sync import OntologyDataSync

                                rr_stmt = (
                                    select(RR)
                                    .where(
                                        and_(
                                            RR.store_id == sid,
                                            RR.report_date == date.today(),
                                            RR.dimension == dim,
                                        )
                                    )
                                    .limit(1)
                                )
                                rr = (await session.execute(rr_stmt)).scalar_one_or_none()
                                if rr:
                                    with OntologyDataSync() as sync:
                                        sync.upsert_reasoning_report(
                                            report_id=str(rr.id),
                                            store_id=sid,
                                            report_date=rr.report_date.isoformat(),
                                            dimension=dim,
                                            severity=c.severity,
                                            root_cause=c.root_cause,
                                            confidence=c.confidence,
                                            triggered_rules=c.triggered_rules,
                                        )
                            except Exception as neo4j_err:
                                logger.warning(
                                    "Neo4j ReasoningReport 同步失败",
                                    store_id=sid,
                                    error=str(neo4j_err),
                                )

                            # 企微推送
                            try:
                                from ..services.wechat_action_fsm import ActionCategory, ActionPriority, get_wechat_fsm

                                fsm = get_wechat_fsm()
                                priority = ActionPriority.P1 if c.severity == "P1" else ActionPriority.P2
                                action_text = c.recommended_actions[0] if c.recommended_actions else "请查看推理报告并采取行动"
                                await fsm.create_action(
                                    store_id=sid,
                                    category=ActionCategory.KPI_ALERT,
                                    priority=priority,
                                    title=f"L4推理告警：{dim} 维度 {c.severity}",
                                    content=(
                                        f"**{dim}** 维度异常\n"
                                        f"根因: {c.root_cause or '待分析'}\n"
                                        f"置信度: {c.confidence:.0%}\n"
                                        f"建议: {action_text}"
                                    ),
                                    receiver_user_id="store_manager",
                                    source_event_id=f"L4-{sid}-{dim}",
                                    evidence={"dimension": dim, "severity": c.severity},
                                )
                                wechat_sent += 1
                            except Exception as wx_err:
                                logger.warning(
                                    "L4 企微告警推送失败",
                                    store_id=sid,
                                    dim=dim,
                                    error=str(wx_err),
                                )

                except Exception as e:
                    errors.append({"store_id": sid, "error": str(e)})
                    logger.error("门店推理扫描失败", store_id=sid, error=str(e))

            await session.commit()

        logger.info(
            "L4 夜间推理扫描完成",
            stores_scanned=stores_scanned,
            p1_alerts=p1_count,
            p2_alerts=p2_count,
            wechat_sent=wechat_sent,
            errors=len(errors),
        )
        return {
            "success": len(errors) == 0,
            "stores_scanned": stores_scanned,
            "p1_alerts": p1_count,
            "p2_alerts": p2_count,
            "wechat_sent": wechat_sent,
            "errors": errors,
        }

    async def _build_kpi_context(session, store_id: str) -> Dict[str, Any]:
        """从 cross_store_metrics 物化表拉取近期 KPI 值"""
        from datetime import date, timedelta

        from sqlalchemy import and_, select

        from ..models.cross_store import CrossStoreMetric

        yesterday = date.today() - timedelta(days=1)
        stmt = select(
            CrossStoreMetric.metric_name,
            CrossStoreMetric.value,
        ).where(
            and_(
                CrossStoreMetric.store_id == store_id,
                CrossStoreMetric.metric_date == yesterday,
            )
        )
        rows = (await session.execute(stmt)).all()
        return {r[0]: r[1] for r in rows}

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


# ── L5 夜间行动派发任务 ───────────────────────────────────────────────────────


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.nightly_action_dispatch",
    max_retries=2,
    default_retry_delay=300,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=1800,
    retry_jitter=False,
)
def nightly_action_dispatch(
    self,
    store_ids: list = None,
    days_back: int = 1,
) -> Dict[str, Any]:
    """
    L5 夜间行动批量派发任务（建议调度时间: 04:30，L4 nightly_reasoning_scan 完成后执行）

    执行步骤：
      1. 查询近 days_back 天内所有未派发行动的 P1/P2 推理报告
      2. 按 severity 优先级（P1 优先）逐一触发 ActionDispatchService.dispatch_from_report()
      3. 汇总派发统计并返回

    Args:
        store_ids: 指定门店列表（None = 全平台）
        days_back: 回溯天数（默认 1 = 仅处理昨日和今日报告）

    Returns:
        {success, stores_covered, plans_created, dispatched, partial, skipped, errors}
    """
    import asyncio
    from typing import Any, Dict

    async def _run():
        from datetime import date, timedelta

        from sqlalchemy import and_, select
        from src.core.database import async_session_factory
        from src.models.action_plan import ActionPlan
        from src.models.reasoning import ReasoningReport
        from src.services.action_dispatch_service import ActionDispatchService

        total_stats = {
            "plans_created": 0,
            "dispatched": 0,
            "partial": 0,
            "skipped": 0,
            "errors": 0,
        }
        stores_covered: set = set()

        async with async_session_factory() as session:
            # 确定目标门店
            if store_ids:
                target_stores = store_ids
            else:
                from src.models.store import Store

                rows = (await session.execute(select(Store.id).where(Store.is_active == True))).all()  # noqa: E712
                target_stores = [r[0] for r in rows]

            svc = ActionDispatchService(session)
            for sid in target_stores:
                try:
                    stats = await svc.dispatch_pending_alerts(store_id=sid, days_back=days_back)
                    if stats["plans_created"] > 0 or stats["dispatched"] > 0:
                        stores_covered.add(sid)
                    for k in ("plans_created", "dispatched", "partial", "skipped", "errors"):
                        total_stats[k] += stats.get(k, 0)
                except Exception as e:
                    total_stats["errors"] += 1
                    logger.error("L5 门店行动派发失败", store_id=sid, error=str(e))

            await session.commit()

        logger.info(
            "L5 夜间行动派发完成",
            stores_covered=len(stores_covered),
            **total_stats,
        )
        return {
            "success": total_stats["errors"] == 0,
            "stores_covered": len(stores_covered),
            **total_stats,
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


# ── 多阶段工作流 Celery 任务 ──────────────────────────────────────────────────


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.start_evening_planning_all_stores",
    max_retries=2,
    default_retry_delay=120,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=600,
    retry_jitter=True,
)
def start_evening_planning_all_stores(
    self,
    store_ids: list = None,
) -> Dict[str, Any]:
    """
    每日 17:00 触发：为全平台所有活跃门店启动 Day N+1 规划工作流，
    并立即触发 initial_plan 阶段的快速规划（Fast Mode，<30s）。

    调度建议：beat_schedule 中设置 crontab(hour=17, minute=0)

    Args:
        store_ids: 指定门店列表（None = 全平台活跃门店）

    Returns:
        {success, stores_started, fast_plan_ok, fast_plan_failed, errors}
    """

    async def _run():
        from datetime import date, timedelta

        from src.core.database import async_session_factory
        from src.services.fast_planning_service import FastPlanningService
        from src.services.workflow_engine import WorkflowEngine

        plan_date = date.today() + timedelta(days=1)
        started = 0
        fast_plan_ok = 0
        fast_plan_fail = 0
        errors = 0

        async with async_session_factory() as session:
            # 确定目标门店
            if store_ids:
                target_stores = store_ids
            else:
                from sqlalchemy import select
                from src.models.store import Store

                rows = (await session.execute(select(Store.id).where(Store.is_active == True))).all()  # noqa: E712
                target_stores = [str(r[0]) for r in rows]

            engine = WorkflowEngine(session)
            fast_svc = FastPlanningService(session)

            for sid in target_stores:
                try:
                    # 1. 启动工作流（幂等）
                    wf = await engine.start_daily_workflow(
                        store_id=sid,
                        plan_date=plan_date,
                    )
                    started += 1

                    # 2. 获取 initial_plan 阶段，触发快速规划
                    try:
                        init_phase = await engine.get_phase(wf.id, "initial_plan")
                        if init_phase:
                            content = await fast_svc.generate_initial_plan(sid, plan_date)
                            await engine.submit_decision(
                                phase_id=init_phase.id,
                                content=content,
                                submitted_by="system",
                                mode="fast",
                                data_completeness=content.get("data_completeness", 0.7),
                                confidence=content.get("confidence", 0.75),
                                change_reason="系统 17:00 快速规划自动生成",
                            )
                            fast_plan_ok += 1
                    except Exception as fp_err:
                        fast_plan_fail += 1
                        logger.warning(
                            "快速规划触发失败（非致命）",
                            store_id=sid,
                            error=str(fp_err),
                        )

                except Exception as e:
                    errors += 1
                    logger.error("启动工作流失败", store_id=sid, error=str(e))

            await session.commit()

        logger.info(
            "晚间规划工作流批量启动完成",
            plan_date=str(plan_date),
            stores_started=started,
            fast_plan_ok=fast_plan_ok,
            fast_plan_failed=fast_plan_fail,
            errors=errors,
        )
        return {
            "success": errors == 0,
            "plan_date": str(plan_date),
            "stores_started": started,
            "fast_plan_ok": fast_plan_ok,
            "fast_plan_failed": fast_plan_fail,
            "errors": errors,
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.check_workflow_deadlines",
    max_retries=1,
    default_retry_delay=30,
)
def check_workflow_deadlines(self) -> Dict[str, Any]:
    """
    每 5 分钟扫描全平台所有 running/reviewing 工作流阶段：
      - 距 deadline ≤ 10 分钟 → 发送企微预警
      - 已过 deadline          → 自动锁定阶段并推进到下一阶段

    调度建议：beat_schedule 中设置 crontab(minute="*/5")

    Returns:
        {success, auto_locked, locked_phases}
    """

    async def _run():
        from src.core.database import async_session_factory
        from src.services.timing_service import TimingService

        async with async_session_factory() as session:
            timing = TimingService(session)
            locked = await timing.check_and_auto_lock_all()
            await session.commit()

        return {
            "success": True,
            "auto_locked": len(locked),
            "locked_phases": locked,
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.release_expired_room_locks",
    max_retries=2,
    default_retry_delay=60,
    autoretry_for=(Exception,),
    retry_backoff=True,
)
def release_expired_room_locks(self) -> Dict[str, Any]:
    """
    每天凌晨 01:00 扫描全平台超时锁台预约并自动释放。

    锁台（room_lock）超过 ROOM_LOCK_TIMEOUT_DAYS（默认 7 天）未签约，
    自动回退到意向（intent）阶段，释放场地资源。

    调度建议：beat_schedule 中设置 crontab(hour=1, minute=0)

    Returns:
        {success, released_count, released_ids}
    """

    async def _run():
        from src.core.database import async_session_factory
        from src.services.banquet_lifecycle_service import BanquetLifecycleService

        async with async_session_factory() as session:
            svc = BanquetLifecycleService(session)
            released = await svc.release_expired_locks()
            await session.commit()

        return {
            "success": True,
            "released_count": len(released),
            "released_ids": released,
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.monthly_save_fct_tax",
    max_retries=2,
    default_retry_delay=300,
)
def monthly_save_fct_tax(self, year: int = 0, month: int = 0) -> Dict[str, Any]:
    """
    月度税务记录自动保存（FCT 业财税一体化）

    每月1日凌晨 01:00 自动运行：
      - 遍历所有门店
      - 调用 FCTService.estimate_monthly_tax() 估算上月税务
      - 将结果持久化到 fct_tax_records 表

    Args:
        year:  被保存的年份（默认 0 表示自动推断上个月所在年）
        month: 被保存的月份（默认 0 表示自动推断上个月）
    """
    from datetime import date, timedelta

    async def _run():
        from sqlalchemy import select
        from src.core.database import async_session_factory
        from src.models.store import Store
        from src.services.fct_service import FCTService

        # 自动推断上个月
        today = date.today()
        if year == 0 or month == 0:
            first_of_this_month = today.replace(day=1)
            last_month_end = first_of_this_month - timedelta(days=1)
            _year = last_month_end.year
            _month = last_month_end.month
        else:
            _year, _month = year, month

        saved, failed = 0, 0

        async with async_session_factory() as session:
            stores_result = await session.execute(select(Store.id))
            store_ids = [row[0] for row in stores_result.all()]

        for store_id in store_ids:
            try:
                async with async_session_factory() as session:
                    svc = FCTService(session)
                    await svc.estimate_monthly_tax(
                        store_id=store_id,
                        year=_year,
                        month=_month,
                        save=True,
                    )
                    await session.commit()
                    saved += 1
            except Exception as exc:
                logger.warning(
                    "月度税务保存失败",
                    store_id=store_id,
                    year=_year,
                    month=_month,
                    error=str(exc),
                )
                failed += 1

        logger.info(
            "月度税务批量保存完成",
            year=_year,
            month=_month,
            saved=saved,
            failed=failed,
        )
        return {"year": _year, "month": _month, "saved": saved, "failed": failed}

    try:
        return asyncio.run(_run())
    except Exception as e:
        raise self.retry(exc=e)


# ============================================================
# ARCH-003: 门店记忆层更新任务
# ============================================================


@celery_app.task(bind=True, max_retries=3, default_retry_delay=300)
def update_store_memory(self, store_id: str = None, brand_id: str = None):
    """每日凌晨2AM更新门店记忆层（峰时模式/菜品健康度/员工基线）"""
    import asyncio

    async def _run():
        from src.core.database import get_db_session
        from src.services.store_memory_service import StoreMemoryService

        async with get_db_session() as db:
            service = StoreMemoryService(db_session=db)
            if store_id:
                await service.refresh_store_memory(store_id=store_id, brand_id=brand_id)
            else:
                # 全量刷新：从 DB 获取所有门店
                from sqlalchemy import select
                from src.models.store import Store

                result = await db.execute(select(Store.id, Store.brand_id))
                stores = result.all()
                for row in stores:
                    try:
                        await service.refresh_store_memory(
                            store_id=str(row.id),
                            brand_id=str(row.brand_id) if row.brand_id else None,
                        )
                    except Exception as exc:
                        logger.warning(
                            "update_store_memory.store_failed",
                            store_id=str(row.id),
                            error=str(exc),
                        )

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


@celery_app.task(bind=True, max_retries=3, default_retry_delay=30)
def realtime_anomaly_check(self, store_id: str, event: dict):
    """StaffAction 写入后触发的实时异常检测"""
    import asyncio

    async def _run():
        from src.core.database import get_db_session
        from src.services.store_memory_service import StoreMemoryService

        async with get_db_session() as db:
            service = StoreMemoryService(db_session=db)
            await service.detect_anomaly(store_id=store_id, event=event)

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# FEAT-002: 预测性备料推送任务
# ============================================================


@celery_app.task(bind=True, max_retries=3, default_retry_delay=300)
def push_daily_forecast(self, store_id: str = None):
    """每日9AM推送预测性备料建议，confidence=low 时附\"数据积累中\"提示"""
    import asyncio
    from datetime import date, timedelta

    async def _run():
        from src.core.database import get_db_session
        from src.services.demand_forecaster import DemandForecaster
        from src.services.wechat_service import wechat_service

        target_date = date.today() + timedelta(days=1)

        async with get_db_session() as db:
            forecaster = DemandForecaster(db_session=db)

            stores_to_forecast = []
            if store_id:
                stores_to_forecast = [store_id]
            else:
                from sqlalchemy import select
                from src.models.store import Store

                result = await db.execute(select(Store.id))
                stores_to_forecast = [str(row.id) for row in result.all()]

            for sid in stores_to_forecast:
                try:
                    result = await forecaster.predict(store_id=sid, target_date=target_date)
                    message_data = {
                        "store_id": sid,
                        "target_date": str(result.target_date),
                        "estimated_revenue": result.estimated_revenue,
                        "confidence": result.confidence,
                        "basis": result.basis,
                        "note": result.note or "",
                        "item_count": len(result.items),
                    }
                    await wechat_service.send_templated_message(
                        template="daily_forecast",
                        data=message_data,
                        to_user_id=f"store_{sid}",
                    )
                except Exception as exc:
                    logger.warning(
                        "push_daily_forecast.store_failed",
                        store_id=sid,
                        error=str(exc),
                    )

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# L8: 07:00 今日人力建议推送任务
# ============================================================


@celery_app.task(
    bind=True,
    max_retries=3,
    default_retry_delay=300,
    name="tasks.push_daily_workforce_advice",
)
def push_daily_workforce_advice(self, store_id: str = None):
    """每日 07:00 推送人力建议（默认遍历所有活跃门店）。"""
    import asyncio

    async def _run():
        from sqlalchemy import select
        from src.core.database import get_db_session
        from src.models.store import Store, StoreStatus
        from src.services.workforce_push_service import WorkforcePushService

        async with get_db_session() as db:
            if store_id:
                stores = [(store_id, "")]
            else:
                active_rows = (
                    await db.execute(
                        select(Store.id, Store.name).where(
                            Store.is_active.is_(True),
                            Store.status == StoreStatus.ACTIVE.value,
                        )
                    )
                ).all()
                if active_rows:
                    stores = [(str(r.id), r.name or "") for r in active_rows]
                else:
                    rows = (await db.execute(select(Store.id, Store.name))).all()
                    stores = [(str(r.id), r.name or "") for r in rows]

            for sid, sname in stores:
                try:
                    await WorkforcePushService.push_daily_staffing_advice(
                        store_id=sid,
                        db=db,
                        store_name=sname,
                        recipient_user_id=_get_store_recipient(sid),
                    )
                except Exception as exc:
                    logger.warning(
                        "push_daily_workforce_advice.store_failed",
                        store_id=sid,
                        error=str(exc),
                    )

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# L8: 自动排班（预算硬约束 + 异常提醒）
# ============================================================


@celery_app.task(
    bind=True,
    max_retries=3,
    default_retry_delay=300,
    name="tasks.auto_generate_workforce_schedule",
)
def auto_generate_workforce_schedule(self, store_id: str = None):
    """每日自动生成当日排班，并在异常时提醒门店负责人。"""
    import asyncio
    from datetime import date as _date

    async def _run():
        from sqlalchemy import select
        from src.core.database import get_db_session
        from src.models.store import Store, StoreStatus
        from src.services.workforce_auto_schedule_service import WorkforceAutoScheduleService

        async with get_db_session() as db:
            if store_id:
                stores = [str(store_id)]
            else:
                active_rows = (
                    await db.execute(
                        select(Store.id).where(
                            Store.is_active.is_(True),
                            Store.status == StoreStatus.ACTIVE.value,
                        )
                    )
                ).all()
                stores = [str(r.id) for r in active_rows] if active_rows else []
                if not stores:
                    rows = (await db.execute(select(Store.id))).all()
                    stores = [str(r.id) for r in rows]

            target_date = _date.today()
            for sid in stores:
                try:
                    await WorkforceAutoScheduleService.generate_schedule_with_constraints(
                        store_id=sid,
                        schedule_date=target_date,
                        db=db,
                        auto_publish=True,
                        notify_on_anomaly=True,
                        recipient_user_id=_get_store_recipient(sid),
                    )
                except Exception as exc:
                    logger.warning(
                        "auto_generate_workforce_schedule.store_failed",
                        store_id=sid,
                        error=str(exc),
                    )

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# INFRA-002: 企微消息重试任务
# ============================================================


@celery_app.task(bind=True, max_retries=1)
def retry_failed_wechat_messages(self):
    """每5分钟从失败队列取出企微消息重试（最多3次）"""
    import asyncio

    async def _run():
        from src.services.wechat_service import wechat_service

        await wechat_service.retry_failed_messages(max_retries=3, batch_size=10)

    try:
        asyncio.run(_run())
    except Exception as exc:
        logger.warning("retry_failed_wechat_messages.error", error=str(exc))


# ============================================================
# v2.0 决策效果反馈（48h 后检查）
# ============================================================


@celery_app.task(bind=True, max_retries=2, default_retry_delay=300)
def check_decision_impact(self, decision_id: str):
    """
    审批后 48h 效果回检。

    对比决策前后的关键指标，向审批人发送效果反馈消息，
    并将决策 outcome 设为 PENDING（待人工核实）。
    """
    import asyncio

    async def _run():
        from sqlalchemy import select
        from src.core.database import get_db_session
        from src.models.decision_log import DecisionLog, DecisionOutcome, DecisionStatus
        from src.services.decision_push_service import _APPROVAL_BASE_URL
        from src.services.wechat_service import wechat_service

        async with get_db_session() as db:
            result = await db.execute(select(DecisionLog).where(DecisionLog.id == decision_id))
            dl = result.scalar_one_or_none()
            if not dl:
                logger.warning("check_decision_impact.not_found", decision_id=decision_id)
                return

            # 仅处理已批准且尚无结果的决策
            if dl.decision_status not in (DecisionStatus.APPROVED, DecisionStatus.EXECUTED):
                logger.info("check_decision_impact.skip_status", decision_id=decision_id, status=str(dl.decision_status))
                return
            if dl.outcome is None:
                dl.outcome = DecisionOutcome.PENDING
                await db.commit()

            # 向审批人发送 48h 效果反馈提醒
            try:
                title = f"【48h效果反馈】{dl.decision_type}"
                description = (
                    f"您于48小时前批准的决策\n"
                    f"来源：{dl.agent_type} / {dl.agent_method}\n"
                    f"门店：{dl.store_id}\n"
                    f"请核实执行效果并在系统中记录结果"
                )
                action_url = f"{_APPROVAL_BASE_URL}/{decision_id}/outcome"
                recipient = dl.manager_id or f"store_{dl.store_id}"
                await wechat_service.send_decision_card(
                    title=title,
                    description=description,
                    action_url=action_url,
                    btntxt="记录结果",
                    to_user_id=recipient,
                )
                logger.info("check_decision_impact.feedback_sent", decision_id=decision_id, recipient=recipient)
            except Exception as exc:
                logger.warning("check_decision_impact.feedback_failed", decision_id=decision_id, error=str(exc))

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# v2.0 决策型企微推送（4时间点）
# ============================================================


def _get_store_recipient(store_id: str) -> str:
    """获取门店负责人的企微 user_id（优先从环境变量读，兜底返回 store_{id}）。"""
    return os.getenv(f"WECHAT_RECIPIENT_{store_id.upper()}", f"store_{store_id}")


@celery_app.task(bind=True, max_retries=3, default_retry_delay=120)
def push_morning_decisions(self, store_id: str = None):
    """08:00晨推：为所有（或指定）门店推送今日 Top3 决策卡片。"""
    import asyncio

    async def _run():
        from sqlalchemy import select
        from src.core.database import get_db_session
        from src.models.store import Store
        from src.services.decision_push_service import DecisionPushService

        async with get_db_session() as db:
            if store_id:
                stores = [(store_id, "")]
            else:
                rows = (await db.execute(select(Store.id, Store.name))).all()
                stores = [(str(r.id), r.name or "") for r in rows]

            for sid, sname in stores:
                try:
                    await DecisionPushService.push_morning_decisions(
                        store_id=sid,
                        brand_id="",
                        recipient_user_id=_get_store_recipient(sid),
                        db=db,
                        store_name=sname,
                    )
                except Exception as exc:
                    logger.warning("push_morning_decisions.store_failed", store_id=sid, error=str(exc))

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


@celery_app.task(bind=True, max_retries=3, default_retry_delay=120)
def push_noon_anomaly(self, store_id: str = None):
    """12:00午推：上午损耗/成本率异常汇总（仅在存在异常时推送）。"""
    import asyncio

    async def _run():
        from sqlalchemy import select
        from src.core.database import get_db_session
        from src.models.store import Store
        from src.services.decision_push_service import DecisionPushService

        async with get_db_session() as db:
            if store_id:
                stores = [(store_id, "")]
            else:
                rows = (await db.execute(select(Store.id, Store.name))).all()
                stores = [(str(r.id), r.name or "") for r in rows]

            for sid, sname in stores:
                try:
                    await DecisionPushService.push_noon_anomaly(
                        store_id=sid,
                        brand_id="",
                        recipient_user_id=_get_store_recipient(sid),
                        db=db,
                        store_name=sname,
                    )
                except Exception as exc:
                    logger.warning("push_noon_anomaly.store_failed", store_id=sid, error=str(exc))

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


@celery_app.task(bind=True, max_retries=3, default_retry_delay=120)
def push_prebattle_decisions(self, store_id: str = None):
    """17:30战前推：库存/排班备战核查（仅在有库存或紧急决策时推送）。"""
    import asyncio

    async def _run():
        from sqlalchemy import select
        from src.core.database import get_db_session
        from src.models.store import Store
        from src.services.decision_push_service import DecisionPushService

        async with get_db_session() as db:
            if store_id:
                stores = [(store_id, "")]
            else:
                rows = (await db.execute(select(Store.id, Store.name))).all()
                stores = [(str(r.id), r.name or "") for r in rows]

            for sid, sname in stores:
                try:
                    await DecisionPushService.push_prebattle_decisions(
                        store_id=sid,
                        brand_id="",
                        recipient_user_id=_get_store_recipient(sid),
                        db=db,
                        store_name=sname,
                    )
                except Exception as exc:
                    logger.warning("push_prebattle_decisions.store_failed", store_id=sid, error=str(exc))

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


@celery_app.task(bind=True, max_retries=3, default_retry_delay=120)
def push_evening_recap(self, store_id: str = None):
    """20:30晚推：当日回顾+待批决策提醒。"""
    import asyncio

    async def _run():
        from sqlalchemy import select
        from src.core.database import get_db_session
        from src.models.store import Store
        from src.services.decision_push_service import DecisionPushService

        async with get_db_session() as db:
            if store_id:
                stores = [(store_id, "")]
            else:
                rows = (await db.execute(select(Store.id, Store.name))).all()
                stores = [(str(r.id), r.name or "") for r in rows]

            for sid, sname in stores:
                try:
                    await DecisionPushService.push_evening_recap(
                        store_id=sid,
                        brand_id="",
                        recipient_user_id=_get_store_recipient(sid),
                        db=db,
                        store_name=sname,
                    )
                except Exception as exc:
                    logger.warning("push_evening_recap.store_failed", store_id=sid, error=str(exc))

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# v2.0 KPI 食材成本率告警（09:30 日检）
# ============================================================


@celery_app.task(bind=True, max_retries=2, default_retry_delay=300)
def check_food_cost_kpi_alert(self):
    """
    每日 09:30 检查所有门店食材成本率是否超过 AlertThresholdsPage 中配置的阈值。
    超标时发送企业微信告警（warning=橙 / critical=红）。
    """
    import asyncio

    async def _run():
        from src.core.database import get_db_session
        from src.services.kpi_alert_service import KPIAlertService

        async with get_db_session() as db:
            result = await KPIAlertService.run_and_notify(db=db)
            logger.info(
                "food_cost_kpi_alert_done",
                total=result["total"],
                alerts=result["alert_count"],
                sent=result["sent_count"],
                failed=result["failed_count"],
            )

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# v2.0 食材成本率趋势预测告警（09:45 日检）
# ============================================================


@celery_app.task(bind=True, max_retries=2, default_retry_delay=300)
def check_food_cost_trend_alert(self):
    """
    每日 09:45 检查所有门店成本率趋势，对趋势恶化（预计 7 天内超标）的门店
    提前发送企业微信趋势预警，防患于未然。
    """
    import asyncio

    async def _run():
        from src.core.database import get_db_session
        from src.services.kpi_alert_service import KPIAlertService

        async with get_db_session() as db:
            result = await KPIAlertService.run_trend_alerts(db=db)
            logger.info(
                "food_cost_trend_alert_done",
                total=result["total"],
                trend_alerts=result["trend_alert_count"],
                sent=result["sent_count"],
                failed=result["failed_count"],
            )

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)
    """
    每日 06:00 扫描私域会员生命周期变更。

    逻辑：
      1. 查询 recency_days > 45 且状态非终态的会员 → 尝试 CHURN_WARNING 触发
      2. 查询 recency_days > 90 且状态为 at_risk 的会员  → 尝试 INACTIVITY_LONG 触发
      3. 成功转移到 at_risk 或 dormant 的会员，自动触发 dormant_wakeup 旅程

    每次每店最多处理 50 人，避免长时间锁表。
    """
    import asyncio

    async def _run():
        from sqlalchemy import text as _text
        from src.core.database import get_db_session
        from src.models.member_lifecycle import StateTransitionTrigger
        from src.services.journey_orchestrator import JourneyOrchestrator
        from src.services.lifecycle_state_machine import LifecycleStateMachine
        from src.services.member_context_store import MemberContextStore, get_context_store

        sm = LifecycleStateMachine()
        orch = JourneyOrchestrator()
        ctx_store = await get_context_store() or MemberContextStore(None)  # None = no-op
        stats = {"scanned": 0, "transitioned": 0, "journeys_triggered": 0}

        async with get_db_session() as db:
            # 获取近30天有交易的活跃门店
            stores_sql = _text("""
                SELECT DISTINCT store_id FROM orders
                WHERE created_at >= NOW() - (30 * INTERVAL '1 day')
            """)
            store_rows = (await db.execute(stores_sql)).fetchall()
            store_ids = [r[0] for r in store_rows]

            for store_id in store_ids:
                # ── 批次1：churn_warning 候选（recency > 45，非终态）──────────
                churn_sql = _text("""
                    SELECT customer_id, wechat_openid
                    FROM private_domain_members
                    WHERE store_id = :store_id
                      AND recency_days > 45
                      AND COALESCE(lifecycle_state, 'repeat')
                          NOT IN ('at_risk', 'dormant', 'lost',
                                  'lead', 'registered', 'first_order_pending')
                    LIMIT 50
                """)
                churn_rows = (await db.execute(churn_sql, {"store_id": store_id})).fetchall()

                for row in churn_rows:
                    stats["scanned"] += 1
                    result = await sm.apply_trigger(
                        row[0],
                        store_id,
                        StateTransitionTrigger.CHURN_WARNING,
                        db,
                        changed_by="scan_lifecycle",
                    )
                    if result.get("transitioned"):
                        stats["transitioned"] += 1
                        await ctx_store.invalidate(store_id, row[0])
                        jr = await orch.trigger(
                            row[0],
                            store_id,
                            "dormant_wakeup",
                            db,
                            wechat_user_id=row[1],
                        )
                        if "error" not in jr:
                            stats["journeys_triggered"] += 1

                # ── 批次2：inactivity_long 候选（recency > 90，at_risk）────────
                dormant_sql = _text("""
                    SELECT customer_id, wechat_openid
                    FROM private_domain_members
                    WHERE store_id = :store_id
                      AND recency_days > 90
                      AND lifecycle_state = 'at_risk'
                    LIMIT 50
                """)
                dormant_rows = (await db.execute(dormant_sql, {"store_id": store_id})).fetchall()

                for row in dormant_rows:
                    stats["scanned"] += 1
                    result = await sm.apply_trigger(
                        row[0],
                        store_id,
                        StateTransitionTrigger.INACTIVITY_LONG,
                        db,
                        changed_by="scan_lifecycle",
                    )
                    if result.get("transitioned"):
                        stats["transitioned"] += 1
                        await ctx_store.invalidate(store_id, row[0])
                        jr = await orch.trigger(
                            row[0],
                            store_id,
                            "dormant_wakeup",
                            db,
                            wechat_user_id=row[1],
                        )
                        if "error" not in jr:
                            stats["journeys_triggered"] += 1

        logger.info("lifecycle.scan_done", **stats)
        return stats

    try:
        return asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# 私域增长：旅程步骤延迟执行
# ============================================================


@celery_app.task(bind=True, max_retries=3, default_retry_delay=60)
def execute_journey_step(
    self,
    journey_db_id: str,
    step_index: int,
    wechat_user_id: str = None,
):
    """
    执行旅程的单个步骤（由 JourneyOrchestrator.trigger() 调度）。

    Args:
        journey_db_id:  private_domain_journeys.id（UUID 字符串）
        step_index:     步骤序号（0-based）
        wechat_user_id: 接收人企微 user_id（可为 None，跳过发送）
    """
    import asyncio

    async def _run():
        from src.core.database import get_db_session
        from src.services.journey_narrator import JourneyNarrator
        from src.services.journey_orchestrator import JourneyOrchestrator

        # 懒初始化企微服务（WECHAT_CORP_ID 未配置时降级为 None）
        wechat_svc = None
        try:
            from src.services.wechat_service import wechat_service as _ws

            if _ws.corp_id and _ws.corp_secret:
                wechat_svc = _ws
        except Exception:
            pass  # 企微未配置，静默跳过

        narrator = JourneyNarrator()  # 内部懒初始化 LLM，无 API KEY 自动降级

        # 懒初始化频控引擎（REDIS_URL 未配置时降级为 None）
        freq_cap = None
        try:
            import redis.asyncio as aioredis
            from src.services.frequency_cap_engine import FrequencyCapEngine

            _redis_url = os.getenv("REDIS_URL", "redis://localhost:6379")
            _redis_client = await aioredis.from_url(_redis_url, encoding="utf-8", decode_responses=True)
            freq_cap = FrequencyCapEngine(_redis_client)
        except Exception:
            pass  # Redis 未配置，频控静默禁用

        async with get_db_session() as db:
            orchestrator = JourneyOrchestrator()
            result = await orchestrator.execute_step(
                journey_db_id,
                step_index,
                db,
                wechat_user_id=wechat_user_id,
                wechat_service=wechat_svc,
                freq_cap_engine=freq_cap,
                narrator=narrator,
            )
            logger.info(
                "journey.celery_step_done",
                journey_db_id=journey_db_id,
                step_index=step_index,
                wechat_wired=wechat_svc is not None,
                freq_cap_wired=freq_cap is not None,
                result=result,
            )
            return result

    try:
        asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# 私域增长：旅程 catch-up dispatcher（每 5 分钟）
# ============================================================


@celery_app.task(bind=True, max_retries=2, default_retry_delay=60)
def dispatch_stale_journeys(self) -> Dict[str, Any]:
    """
    扫描所有 RUNNING 旅程中 next_action_at 已过期的步骤，重新调度执行。

    解决场景：Celery worker 重启 / countdown 任务丢失 / Redis broker 清空后，
    旅程卡在 running 状态无人处理。

    逻辑：
      1. 查询 status='running' AND next_action_at <= NOW() 的旅程（上限 100）
      2. 对每条旅程，调度 execute_journey_step(current_step) 立即执行
      3. 将 next_action_at 设为 NULL 避免重复调度（execute_step 完成后会设置下一步的时间）

    调度周期：celery beat 每 5 分钟触发一次
    """
    import asyncio

    async def _run() -> Dict[str, Any]:
        from src.core.database import get_db_session

        dispatched = 0
        errors = []

        async with get_db_session() as session:
            from sqlalchemy import text as sa_text

            # 查询过期的 running 旅程
            result = await session.execute(
                sa_text("""
                    SELECT id, journey_type, current_step, total_steps
                    FROM private_domain_journeys
                    WHERE status = 'running'
                      AND next_action_at IS NOT NULL
                      AND next_action_at <= NOW()
                    ORDER BY next_action_at ASC
                    LIMIT 100
                """)
            )
            stale_journeys = result.fetchall()

            if not stale_journeys:
                return {"dispatched": 0, "total_stale": 0}

            for journey in stale_journeys:
                j_id = str(journey.id)
                step_idx = journey.current_step or 0

                if step_idx >= (journey.total_steps or 0):
                    # 步骤已超出范围，标记完成
                    await session.execute(
                        sa_text("""
                            UPDATE private_domain_journeys
                            SET status = 'completed', next_action_at = NULL,
                                completed_at = NOW(), updated_at = NOW()
                            WHERE id = :id
                        """),
                        {"id": j_id},
                    )
                    continue

                try:
                    # 清除 next_action_at 防止下次轮询重复调度
                    await session.execute(
                        sa_text("""
                            UPDATE private_domain_journeys
                            SET next_action_at = NULL, updated_at = NOW()
                            WHERE id = :id AND next_action_at IS NOT NULL
                        """),
                        {"id": j_id},
                    )

                    # 调度立即执行（countdown=0）
                    execute_journey_step.apply_async(
                        args=[j_id, step_idx, None],
                        countdown=0,
                    )
                    dispatched += 1
                except Exception as e:
                    errors.append({"journey_id": j_id, "error": str(e)[:100]})

            await session.commit()

        logger.info(
            "dispatch_stale_journeys.done",
            total_stale=len(stale_journeys),
            dispatched=dispatched,
            errors=len(errors),
        )
        return {
            "total_stale": len(stale_journeys),
            "dispatched": dispatched,
            "errors": errors[:5],
        }

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("dispatch_stale_journeys.failed", error=str(exc))
        raise self.retry(exc=exc)


# ============================================================
# 私域增长：RFM 数据刷新（每日凌晨 03:00）
# ============================================================


@celery_app.task(bind=True, max_retries=2, default_retry_delay=120)
def refresh_private_domain_rfm(self):
    """
    每日刷新私域会员 RFM 三指标：
      - recency_days  : 距最近一次订单的天数
      - frequency     : 历史订单总笔数
      - monetary      : 历史消费总金额（分）

    使用单条 UPDATE...FROM 语句批量更新，避免逐行查询。
    仅更新有历史订单的会员；新增会员若尚无订单则 recency_days 保持 NULL。
    """
    import asyncio

    async def _run():
        from sqlalchemy import text as _text
        from src.core.database import get_db_session
        from src.services.member_context_store import MemberContextStore, get_context_store

        sql = _text("""
            UPDATE private_domain_members AS m
            SET
                recency_days = EXTRACT(
                    DAY FROM (NOW() - agg.last_order_at)
                )::int,
                frequency    = agg.order_count,
                monetary     = agg.total_amount
            FROM (
                SELECT
                    customer_id,
                    store_id,
                    MAX(created_at)                AS last_order_at,
                    COUNT(*)                       AS order_count,
                    COALESCE(SUM(total_amount), 0) AS total_amount
                FROM orders
                WHERE customer_id IS NOT NULL
                GROUP BY customer_id, store_id
            ) AS agg
            WHERE m.customer_id = agg.customer_id
              AND m.store_id    = agg.store_id
        """)

        async with get_db_session() as db:
            result = await db.execute(sql)
            await db.commit()
            updated = result.rowcount

        # DB 更新后批量清除 Redis 缓存（按门店），下次读时重新填充
        invalidated = 0
        ctx_store = await get_context_store() or MemberContextStore(None)
        try:
            stores_sql = _text("""
                SELECT DISTINCT store_id FROM private_domain_members
            """)
            async with get_db_session() as db2:
                store_result = await db2.execute(stores_sql)
                store_rows = await _maybe_await(store_result.fetchall())
            for (sid,) in store_rows:
                invalidated += await ctx_store.invalidate_store(sid)
        except Exception as exc:
            logger.warning("rfm.ctx_invalidate_failed", error=str(exc))

        logger.info("rfm.refresh_done", updated_rows=updated, cache_invalidated=invalidated)
        return {"updated": updated, "cache_invalidated": invalidated}

    try:
        return asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ============================================================
# 私域增长：新会员旅程自动触发（每小时）
# ============================================================


@celery_app.task(bind=True, max_retries=2, default_retry_delay=60)
def trigger_new_member_journeys(self):
    """
    每小时扫描过去 70 分钟内新加入的私域会员（留 10 分钟缓冲防漏），
    对尚未有 member_activation 旅程的成员自动触发激活旅程。

    每批最多 100 人，避免大量并发写入。
    """
    import asyncio

    async def _run():
        from sqlalchemy import text as _text
        from src.core.database import get_db_session
        from src.services.journey_orchestrator import JourneyOrchestrator

        orch = JourneyOrchestrator()
        stats = {"scanned": 0, "triggered": 0, "skipped": 0}

        sql = _text("""
            SELECT m.customer_id, m.store_id, m.wechat_openid
            FROM private_domain_members m
            WHERE m.created_at >= NOW() - (70 * INTERVAL '1 minute')
              AND NOT EXISTS (
                  SELECT 1
                  FROM private_domain_journeys j
                  WHERE j.customer_id  = m.customer_id
                    AND j.store_id     = m.store_id
                    AND j.journey_type = 'member_activation'
              )
            LIMIT 100
        """)

        async with get_db_session() as db:
            result = await db.execute(sql)
            rows = await _maybe_await(result.fetchall())
            for row in rows:
                customer_id, store_id, wechat_openid = row[0], row[1], row[2]
                stats["scanned"] += 1
                result = await orch.trigger(
                    customer_id,
                    store_id,
                    "member_activation",
                    db,
                    wechat_user_id=wechat_openid,
                )
                if "error" in result:
                    stats["skipped"] += 1
                else:
                    stats["triggered"] += 1

        logger.info("new_member_journey.done", **stats)
        return stats

    try:
        return asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ── Agent-13: 需求预测主动触达 ──────────────────────────────────────────────────


@celery_app.task(bind=True, max_retries=2, default_retry_delay=120)
def trigger_demand_predictions(self):
    """
    每日 09:15 扫描所有门店，找出 72h 内即将到店的高频会员，
    触发 proactive_remind 旅程（发企微提醒，提升选择率）。

    设计：
    - DemandPredictor.scan_upcoming_visitors(horizon_days=3)
    - 每门店最多处理 100 人（SQL LIMIT 保证）
    - 已有进行中 proactive_remind 旅程的会员自动跳过（SQL WHERE NOT EXISTS 保证）
    - 任意门店失败不影响其他门店
    """
    import asyncio

    async def _run():
        from sqlalchemy import text as _text
        from src.core.database import get_db_session
        from src.services.demand_predictor import DemandPredictor
        from src.services.journey_orchestrator import JourneyOrchestrator

        predictor = DemandPredictor()
        orch = JourneyOrchestrator()
        stats: dict = {"stores": 0, "candidates": 0, "triggered": 0, "skipped": 0}

        # 取所有有会员的门店
        store_sql = _text(
            "SELECT DISTINCT store_id FROM private_domain_members " "WHERE lifecycle_state NOT IN ('lost', 'lead') LIMIT 200"
        )

        async with get_db_session() as db:
            store_result = await db.execute(store_sql)
            stores = await _maybe_await(store_result.fetchall())
            stats["stores"] = len(stores)

            for (store_id,) in stores:
                try:
                    candidates = await predictor.scan_upcoming_visitors(store_id, db, horizon_days=3)
                    for c in candidates:
                        stats["candidates"] += 1
                        result = await orch.trigger(
                            c.customer_id,
                            c.store_id,
                            "proactive_remind",
                            db,
                            wechat_user_id=c.wechat_openid,
                        )
                        if "error" in result:
                            stats["skipped"] += 1
                        else:
                            stats["triggered"] += 1
                except Exception as exc:
                    logger.warning(
                        "demand_predictions.store_failed",
                        store_id=store_id,
                        error=str(exc),
                    )

        logger.info("demand_predictions.done", **stats)
        return stats

    try:
        return asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ── EventScheduler: 生日/入会周年触达 ─────────────────────────────────────────


@celery_app.task(bind=True, max_retries=2, default_retry_delay=120)
def trigger_birthday_reminders(self):
    """
    每日 10:00 扫描所有门店，找出 3 天内生日或入会周年的会员，
    分别触发 birthday_greeting / anniversary_greeting 旅程。

    设计：
    - BirthdayReminderService.scan_upcoming_events(horizon_days=3)
    - 30 天内已有进行中旅程的会员自动跳过（SQL WHERE NOT EXISTS 保证）
    - 任意门店失败不影响其他门店
    """
    import asyncio

    async def _run():
        from sqlalchemy import text as _text
        from src.core.database import get_db_session
        from src.services.birthday_reminder_service import BirthdayReminderService
        from src.services.journey_orchestrator import JourneyOrchestrator

        svc = BirthdayReminderService()
        orch = JourneyOrchestrator()
        stats: dict = {
            "stores": 0,
            "birthday_triggered": 0,
            "anniversary_triggered": 0,
            "skipped": 0,
        }

        store_sql = _text("SELECT DISTINCT store_id FROM private_domain_members " "WHERE is_active = true LIMIT 200")

        _EVENT_TO_JOURNEY = {
            "birthday": "birthday_greeting",
            "anniversary": "anniversary_greeting",
        }

        async with get_db_session() as db:
            stores = (await db.execute(store_sql)).fetchall()
            stats["stores"] = len(stores)

            for (store_id,) in stores:
                try:
                    events = await svc.scan_upcoming_events(store_id, db, horizon_days=3)
                    for ev in events:
                        journey_id = _EVENT_TO_JOURNEY[ev.event_type]
                        result = await orch.trigger(
                            ev.customer_id,
                            ev.store_id,
                            journey_id,
                            db,
                            wechat_user_id=ev.wechat_openid,
                        )
                        if "error" in result:
                            stats["skipped"] += 1
                        else:
                            stats[f"{ev.event_type}_triggered"] += 1
                except Exception as exc:
                    logger.warning(
                        "birthday_reminders.store_failed",
                        store_id=store_id,
                        error=str(exc),
                    )

        logger.info("birthday_reminders.done", **stats)
        return stats

    try:
        return asyncio.run(_run())
    except Exception as exc:
        raise self.retry(exc=exc)


# ── 天财商龙 POS 日单拉取（凌晨 02:00）──────────────────────────────────────


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY_LONG", "300")),
)
def pull_tiancai_daily_orders(self) -> Dict[str, Any]:
    """
    拉取天财商龙昨日全量已支付订单并写入 orders 表（每日凌晨 02:00 执行，
    在 03:00 POS 对账任务之前完成数据入库）。

    环境变量：
      TIANCAI_BASE_URL    — API 基础地址（必填；缺失则整体跳过）
      TIANCAI_BRAND_ID    — 品牌 ID（用于 to_order 映射）
      TIANCAI_APPID       — Terminal ID（鉴权用，全局共享）
      TIANCAI_ACCESSID    — Terminal authorization ID（鉴权用，全局共享）
      TIANCAI_CENTER_ID   — 集团 centerId（全局或门店级）
      TIANCAI_SHOP_ID     — 默认门店 shopId
      TIANCAI_SHOP_ID_{store_id} — 门店级 shopId（优先于全局）

    Returns:
        {success, date, stores_processed, orders_upserted, errors}
    """

    async def _run():
        from datetime import date, timedelta

        from sqlalchemy import select
        from sqlalchemy import text as _text

        from ..core.database import get_db_session
        from ..models.store import Store

        base_url = os.getenv("TIANCAI_BASE_URL", "")
        if not base_url:
            logger.warning("tiancai_pull.skipped", reason="TIANCAI_BASE_URL not configured")
            return {
                "success": True,
                "skipped": True,
                "stores_processed": 0,
                "orders_upserted": 0,
                "errors": [],
            }

        brand_id = os.getenv("TIANCAI_BRAND_ID", "")
        yesterday = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")

        from packages.api_adapters.tiancai_shanglong.src.adapter import TiancaiShanglongAdapter  # noqa: E402

        stores_processed = 0
        orders_upserted = 0
        errors = []

        async with get_db_session() as session:
            result = await session.execute(select(Store).where(Store.is_active == True))
            stores = result.scalars().all()

            # 全局凭据（各门店可用 TIANCAI_APP_ID_{store_id} 覆盖）
            global_app_id = os.getenv("TIANCAI_APP_ID", "")
            global_app_secret = os.getenv("TIANCAI_APP_SECRET", "")

            for store in stores:
                sid = str(store.id)
                # 门店级凭据优先于全局凭据
                app_id = os.getenv(f"TIANCAI_APP_ID_{sid}") or global_app_id
                app_secret = os.getenv(f"TIANCAI_APP_SECRET_{sid}") or global_app_secret

                # 无凭据 → 静默跳过（不计入 errors）
                if not app_id or not app_secret:
                    logger.debug("tiancai_pull.store_skipped", store_id=sid, reason="no credentials")
                    continue

                try:
                    adapter = TiancaiShanglongAdapter({
                        "base_url": base_url,
                        "app_id": app_id,
                        "app_secret": app_secret,
                        "store_id": sid,
                        "brand_id": brand_id,
                    })
                    order_list = await adapter.pull_daily_orders(yesterday, brand_id)

                    for schema in order_list:
                        total_cents = int(schema.total * 100)
                        discount_cents = int(schema.discount * 100)
                        await session.execute(
                            _text("""
                                INSERT INTO orders
                                    (id, store_id, table_number, status,
                                     total_amount, discount_amount, final_amount,
                                     order_time, waiter_id, sales_channel, notes,
                                     order_metadata, created_at, updated_at)
                                VALUES
                                    (:id, :store_id, :table_number, :status,
                                     :total_amount, :discount_amount, :final_amount,
                                     :order_time, :waiter_id, 'tiancai', :notes,
                                     '{}', NOW(), NOW())
                                ON CONFLICT (id) DO UPDATE SET
                                    status          = EXCLUDED.status,
                                    total_amount    = EXCLUDED.total_amount,
                                    discount_amount = EXCLUDED.discount_amount,
                                    final_amount    = EXCLUDED.final_amount,
                                    updated_at      = NOW()
                            """),
                            {
                                "id": schema.order_id,
                                "store_id": sid,
                                "table_number": schema.table_number,
                                "status": schema.order_status.value,
                                "total_amount": total_cents,
                                "discount_amount": discount_cents,
                                "final_amount": total_cents,
                                "order_time": schema.created_at,
                                "waiter_id": schema.waiter_id,
                                "notes": schema.notes,
                            },
                        )
                        orders_upserted += 1

                    await session.commit()
                    stores_processed += 1
                    logger.info(
                        "tiancai_pull.store_done",
                        store_id=sid,
                        date=yesterday,
                        orders=len(order_list),
                    )

                except Exception as e:
                    await session.rollback()
                    errors.append({"store_id": sid, "error": str(e)})
                    logger.error("tiancai_pull.store_failed", store_id=sid, error=str(e))

        logger.info(
            "tiancai_pull_daily_orders.done",
            date=yesterday,
            stores_processed=stores_processed,
            orders_upserted=orders_upserted,
            errors=len(errors),
        )
        return {
            "success": True,
            "date": yesterday,
            "stores_processed": stores_processed,
            "orders_upserted": orders_upserted,
            "errors": errors,
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        logger.error("tiancai_pull_daily_orders.failed", error=str(e))
        raise self.retry(exc=e)


# ── 品智日营业数据同步（每日 01:30 执行，早于天财 02:00 及对账 03:00）────────────


@celery_app.task(
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY_LONG", "300")),
)
def pull_pinzhi_daily_data(self) -> Dict[str, Any]:
    """
    拉取品智 POS 昨日营业数据（订单 + 营业汇总 + 出品明细）并写入 orders 表。

    环境变量：
      PINZHI_BASE_URL   — API 基础地址（必填；缺失则跳过）
      PINZHI_TOKEN      — API Token（必填）
      PINZHI_BRAND_ID   — 品牌 ID（映射标准 schema 用）

    Returns:
        {success, date, stores_processed, orders_upserted, summaries_saved, errors}
    """

    async def _run():
        from datetime import date, timedelta

        from sqlalchemy import select
        from sqlalchemy import text as _text

        from ..core.database import get_db_session
        from ..models.integration import ExternalSystem, IntegrationType
        from ..models.store import Store

        global_base_url = os.getenv("PINZHI_BASE_URL", "")
        global_token = os.getenv("PINZHI_TOKEN", "")
        global_brand_id = os.getenv("PINZHI_BRAND_ID", "")
        yesterday = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")

        from packages.api_adapters.pinzhi.src.adapter import PinzhiAdapter

        stores_processed = 0
        orders_upserted = 0
        summaries_saved = 0
        errors = []

        try:
            async with get_db_session() as session:
                # 获取所有活跃门店
                result = await session.execute(select(Store).where(Store.is_active.is_(True)))
                stores = result.scalars().all()

                # 预加载品智 ExternalSystem 凭证（store_id → ExternalSystem）
                ext_result = await session.execute(
                    select(ExternalSystem).where(
                        ExternalSystem.provider == "pinzhi",
                        ExternalSystem.type == IntegrationType.POS,
                    )
                )
                ext_by_store: dict = {str(e.store_id): e for e in ext_result.scalars().all() if e.store_id}

                # 若全局凭证和 ExternalSystem 均无配置，跳过整个任务
                if not global_base_url and not global_token and not ext_by_store:
                    logger.warning("pinzhi_pull.skipped", reason="无全局环境变量且接入配置管理中无品智凭证")
                    return {
                        "success": True,
                        "skipped": True,
                        "stores_processed": 0,
                        "orders_upserted": 0,
                        "summaries_saved": 0,
                        "errors": [],
                    }

                for store in stores:
                    sid = str(store.id)
                    cfg = store.config if isinstance(store.config, dict) else {}
                    ext = ext_by_store.get(sid)
                    ext_cfg: dict = ext.config if ext and isinstance(ext.config, dict) else {}

                    # 凭证优先级：store.config > ExternalSystem > 全局环境变量
                    base_url = (
                        cfg.get("pinzhi_base_url")
                        or ext_cfg.get("pinzhi_base_url")
                        or (ext.api_endpoint if ext else None)
                        or global_base_url
                    )
                    token = (
                        cfg.get("pinzhi_token")
                        or ext_cfg.get("pinzhi_store_token")
                        or (ext.api_secret if ext else None)
                        or (ext.api_key if ext else None)
                        or global_token
                    )
                    brand_id = (
                        cfg.get("pinzhi_brand_id")
                        or ext_cfg.get("brand_id")
                        or getattr(store, "brand_id", None)
                        or global_brand_id
                    )
                    ognid = str(
                        cfg.get("pinzhi_ognid")
                        or ext_cfg.get("pinzhi_oms_id")
                        or ext_cfg.get("pinzhi_store_id")
                        or store.code
                        or sid
                    )

                    if not base_url or not token:
                        logger.debug("pinzhi_pull.store_skipped", store_id=sid, reason="无凭证")
                        continue

                    adapter = PinzhiAdapter(
                        {
                            "base_url": base_url,
                            "token": token,
                            "timeout": int(os.getenv("PINZHI_TIMEOUT", "30")),
                            "retry_times": int(os.getenv("PINZHI_RETRY_TIMES", "3")),
                        }
                    )

                    try:
                        # 1) 拉取订单明细
                        page = 1
                        store_orders = 0
                        while True:
                            raw_orders = await adapter.query_orders(
                                ognid=ognid,
                                begin_date=yesterday,
                                end_date=yesterday,
                                page_index=page,
                                page_size=100,
                            )
                            if not raw_orders:
                                break

                            for raw in raw_orders:
                                order_schema = adapter.to_order(raw, sid, brand_id)
                                total_cents = int(order_schema.total * 100)
                                discount_cents = int(order_schema.discount * 100)
                                # CDP: 从原始数据提取会员手机号（供 consumer_id 回填）
                                vip_phone = raw.get("vipMobile") or raw.get("mobile") or ""
                                vip_name = raw.get("vipName") or ""
                                await session.execute(
                                    _text("""
                                        INSERT INTO orders
                                            (id, store_id, table_number, status,
                                             total_amount, discount_amount, final_amount,
                                             order_time, waiter_id, sales_channel, notes,
                                             customer_phone, customer_name,
                                             order_metadata, created_at, updated_at)
                                        VALUES
                                            (:id, :store_id, :table_number, :status,
                                             :total_amount, :discount_amount, :final_amount,
                                             :order_time, :waiter_id, 'pinzhi', :notes,
                                             :customer_phone, :customer_name,
                                             '{}', NOW(), NOW())
                                        ON CONFLICT (id) DO UPDATE SET
                                            status          = EXCLUDED.status,
                                            total_amount    = EXCLUDED.total_amount,
                                            discount_amount = EXCLUDED.discount_amount,
                                            final_amount    = EXCLUDED.final_amount,
                                            customer_phone  = COALESCE(NULLIF(EXCLUDED.customer_phone, ''), orders.customer_phone),
                                            customer_name   = COALESCE(NULLIF(EXCLUDED.customer_name, ''), orders.customer_name),
                                            updated_at      = NOW()
                                    """),
                                    {
                                        "id": order_schema.order_id,
                                        "store_id": sid,
                                        "table_number": order_schema.table_number,
                                        "status": order_schema.order_status.value,
                                        "total_amount": total_cents,
                                        "discount_amount": discount_cents,
                                        "final_amount": total_cents - discount_cents,
                                        "order_time": order_schema.created_at,
                                        "waiter_id": order_schema.waiter_id,
                                        "notes": order_schema.notes,
                                        "customer_phone": vip_phone,
                                        "customer_name": vip_name,
                                    },
                                )
                                store_orders += 1

                            if len(raw_orders) < 100:
                                break
                            page += 1

                        orders_upserted += store_orders

                        # 2) 拉取营业汇总（日报用）
                        summary = await adapter.query_order_summary(ognid, yesterday)
                        if summary:
                            await session.execute(
                                _text("""
                                    INSERT INTO daily_summaries
                                        (store_id, business_date, revenue_cents,
                                         order_count, customer_count, avg_ticket_cents,
                                         source, raw_data, created_at)
                                    VALUES
                                        (:store_id, :biz_date, :revenue,
                                         :order_count, :customer_count, :avg_ticket,
                                         'pinzhi', :raw_data, NOW())
                                    ON CONFLICT (store_id, business_date, source) DO UPDATE SET
                                        revenue_cents   = EXCLUDED.revenue_cents,
                                        order_count     = EXCLUDED.order_count,
                                        customer_count  = EXCLUDED.customer_count,
                                        avg_ticket_cents = EXCLUDED.avg_ticket_cents,
                                        raw_data        = EXCLUDED.raw_data,
                                        created_at      = NOW()
                                """),
                                {
                                    "store_id": sid,
                                    "biz_date": yesterday,
                                    "revenue": int(float(summary.get("realPrice", 0))),
                                    "order_count": int(summary.get("orderCount", 0)),
                                    "customer_count": int(summary.get("customerCount", 0)),
                                    "avg_ticket": int(float(summary.get("avgPrice", 0))),
                                    "raw_data": __import__("json").dumps(summary, ensure_ascii=False, default=str),
                                },
                            )
                            summaries_saved += 1

                        await session.commit()
                        stores_processed += 1
                        logger.info(
                            "pinzhi_pull.store_done",
                            store_id=sid,
                            date=yesterday,
                            orders=store_orders,
                        )

                    except Exception as e:
                        await session.rollback()
                        errors.append({"store_id": sid, "error": str(e)})
                        logger.error("pinzhi_pull.store_failed", store_id=sid, error=str(e))
                    finally:
                        await adapter.close()

        except Exception as exc:
            logger.error("pinzhi_pull.session_error", error=str(exc))
            raise

        logger.info(
            "pinzhi_pull_daily_data.done",
            date=yesterday,
            stores_processed=stores_processed,
            orders_upserted=orders_upserted,
            summaries_saved=summaries_saved,
            errors=len(errors),
        )
        return {
            "success": True,
            "date": yesterday,
            "stores_processed": stores_processed,
            "orders_upserted": orders_upserted,
            "summaries_saved": summaries_saved,
            "errors": errors,
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        logger.error("pinzhi_pull_daily_data.failed", error=str(e))
        raise self.retry(exc=e)


# ── Onboarding Pipeline Task ───────────────────────────────────────────────────


@celery_app.task(bind=True, max_retries=1, default_retry_delay=60)
def run_onboarding_pipeline(self, store_id: str, task_id: str = "") -> Dict[str, Any]:
    """
    5-stage Onboarding Pipeline (async wrapped for Celery):
      1. data_cleaning → 2. kpi_calculation → 3. baseline_compare
      4. vector_embedding → 5. knowledge_summary
    """

    async def _run():
        from ..core.database import get_db_session
        from ..services.onboarding_pipeline_service import OnboardingPipelineService

        async with get_db_session() as session:
            return await OnboardingPipelineService.run(store_id=store_id, db=session)

    try:
        return asyncio.run(_run())
    except Exception as e:
        logger.error("onboarding_pipeline.failed", store_id=store_id, error=str(e))
        raise self.retry(exc=e)


# ── Onboarding Historical Backfill Task ───────────────────────────────────────


@celery_app.task(bind=True, max_retries=2, default_retry_delay=120)
def pull_historical_backfill(
    self,
    store_id: str,
    adapter: str,
    credentials: Dict[str, Any],
) -> Dict[str, Any]:
    """
    历史数据回灌任务（Onboarding Phase 2）。
    由 POST /api/v1/onboarding/connect/{adapter} 触发。

    目前支持的适配器：
      tiancai — 天财商龙（直接复用 TiancaiShanglongAdapter）
    其他适配器返回 skipped=True，待后续集成。

    执行结果写回 OnboardingTask.step="connect" 的进度字段。

    Returns:
        {success, store_id, adapter, records_imported, skipped, errors}
    """

    async def _run() -> Dict[str, Any]:
        from sqlalchemy import select
        from sqlalchemy import text as _text

        from ..core.database import get_db_session
        from ..models.onboarding import OnboardingTask

        async def _update_progress(
            session,
            total: int,
            imported: int,
            failed: int,
            status: str,
            extra: Dict[str, Any],
        ) -> None:
            result = await session.execute(
                select(OnboardingTask).where(
                    OnboardingTask.store_id == store_id,
                    OnboardingTask.step == "connect",
                )
            )
            task = result.scalar_one_or_none()
            if task:
                task.total_records = total
                task.imported_records = imported
                task.failed_records = failed
                task.status = status
                task.extra = extra
                await session.commit()

        # ── 天财商龙适配器 ─────────────────────────────────────────────────────
        if adapter == "tiancai":
            from datetime import date, timedelta

            from packages.api_adapters.tiancai_shanglong.src.adapter import TiancaiShanglongAdapter  # noqa: E402

            base_url = credentials.get("base_url") or os.getenv("TIANCAI_BASE_URL", "")
            if not base_url:
                logger.warning(
                    "backfill.tiancai.skipped",
                    store_id=store_id,
                    reason="TIANCAI_BASE_URL not configured",
                )
                async with get_db_session() as session:
                    await _update_progress(session, 0, 0, 0, "skipped", {"reason": "TIANCAI_BASE_URL 未配置"})
                return {
                    "success": True,
                    "store_id": store_id,
                    "adapter": adapter,
                    "records_imported": 0,
                    "skipped": True,
                    "errors": [],
                }

            app_id = credentials.get("app_id") or os.getenv(f"TIANCAI_APP_ID_{store_id}") or os.getenv("TIANCAI_APP_ID", "")
            app_secret = (
                credentials.get("app_secret")
                or os.getenv(f"TIANCAI_APP_SECRET_{store_id}")
                or os.getenv("TIANCAI_APP_SECRET", "")
            )
            brand_id = credentials.get("brand_id") or os.getenv("TIANCAI_BRAND_ID", "")

            adapter_instance = TiancaiShanglongAdapter(
                base_url=base_url,
                app_id=app_id,
                app_secret=app_secret,
                brand_id=brand_id,
            )

            # 默认回灌最近 30 天
            backfill_days = int(os.getenv("ONBOARDING_BACKFILL_DAYS", "30"))
            today = date.today()
            records_imported = 0
            errors: list = []

            async with get_db_session() as session:
                await _update_progress(session, backfill_days, 0, 0, "in_progress", {"adapter": adapter, "phase": "pulling"})

            for i in range(backfill_days):
                target_date = (today - timedelta(days=i + 1)).strftime("%Y-%m-%d")
                try:
                    orders = await adapter_instance.pull_daily_orders(
                        store_id=store_id,
                        target_date=target_date,
                    )
                    if orders:
                        async with get_db_session() as session:
                            for order in orders:
                                await session.execute(
                                    _text("""
                                        INSERT INTO orders
                                            (id, store_id, table_number, status,
                                             total_amount, discount_amount, final_amount,
                                             order_time, source_system, order_metadata,
                                             created_at, updated_at)
                                        VALUES
                                            (:id, :store_id, :table_number, :status,
                                             :total_amount, :discount_amount, :final_amount,
                                             :order_time, 'tiancai', '{}', NOW(), NOW())
                                        ON CONFLICT (id) DO NOTHING
                                    """),
                                    {
                                        "id": order.order_id,
                                        "store_id": store_id,
                                        "table_number": order.table_no or "",
                                        "status": order.status or "completed",
                                        "total_amount": int(float(order.total_amount or 0) * 100),
                                        "discount_amount": int(float(order.discount_amount or 0) * 100),
                                        "final_amount": int(float(order.paid_amount or 0) * 100),
                                        "order_time": order.order_time,
                                    },
                                )
                                records_imported += 1
                            await session.commit()
                except Exception as e:
                    logger.warning(
                        "backfill.tiancai.day_failed",
                        store_id=store_id,
                        date=target_date,
                        error=str(e),
                    )
                    errors.append({"date": target_date, "error": str(e)})

            async with get_db_session() as session:
                await _update_progress(
                    session,
                    total=backfill_days,
                    imported=records_imported,
                    failed=len(errors),
                    status="completed",
                    extra={
                        "adapter": adapter,
                        "days_pulled": backfill_days,
                        "records_imported": records_imported,
                        "errors": errors[:10],  # 最多记录前10条错误
                    },
                )

            logger.info(
                "backfill.tiancai.done",
                store_id=store_id,
                records_imported=records_imported,
                errors=len(errors),
            )
            return {
                "success": True,
                "store_id": store_id,
                "adapter": adapter,
                "records_imported": records_imported,
                "skipped": False,
                "errors": errors[:10],
            }

        # ── 客如云适配器 ───────────────────────────────────────────────────────
        elif adapter == "keruyun":
            from datetime import date, timedelta

            from packages.api_adapters.keruyun.src.adapter import KeruyunAdapter

            client_id = (
                credentials.get("client_id")
                or os.getenv(f"KERUYUN_CLIENT_ID_{store_id}")
                or os.getenv("KERUYUN_CLIENT_ID", "")
            )
            client_secret = (
                credentials.get("client_secret")
                or os.getenv(f"KERUYUN_CLIENT_SECRET_{store_id}")
                or os.getenv("KERUYUN_CLIENT_SECRET", "")
            )
            if not client_id or not client_secret:
                logger.warning("backfill.keruyun.skipped", store_id=store_id, reason="KERUYUN credentials not configured")
                async with get_db_session() as session:
                    await _update_progress(session, 0, 0, 0, "skipped", {"reason": "KERUYUN_CLIENT_ID/SECRET 未配置"})
                return {
                    "success": True,
                    "store_id": store_id,
                    "adapter": adapter,
                    "records_imported": 0,
                    "skipped": True,
                    "errors": [],
                }

            adapter_instance = KeruyunAdapter(
                config={
                    "base_url": credentials.get("base_url") or os.getenv("KERUYUN_BASE_URL", "https://api.keruyun.com"),
                    "client_id": client_id,
                    "client_secret": client_secret,
                    "store_id": store_id,
                }
            )

            backfill_days = int(os.getenv("ONBOARDING_BACKFILL_DAYS", "30"))
            today = date.today()
            records_imported = 0
            errors: list = []

            async with get_db_session() as session:
                await _update_progress(session, backfill_days, 0, 0, "in_progress", {"adapter": adapter, "phase": "pulling"})

            for i in range(backfill_days):
                target_date = today - timedelta(days=i + 1)
                start_time = target_date.strftime("%Y-%m-%d 00:00:00")
                end_time = target_date.strftime("%Y-%m-%d 23:59:59")
                try:
                    result = await adapter_instance.query_order(start_time=start_time, end_time=end_time)
                    orders = result if isinstance(result, list) else result.get("orders", result.get("list", []))
                    if orders and isinstance(orders, list):
                        async with get_db_session() as session:
                            for order in orders:
                                oid = order.get("order_id") or order.get("orderId") or order.get("id")
                                if not oid:
                                    continue
                                await session.execute(
                                    _text("""
                                        INSERT INTO orders
                                            (id, store_id, table_number, status,
                                             total_amount, discount_amount, final_amount,
                                             order_time, source_system, order_metadata,
                                             created_at, updated_at)
                                        VALUES
                                            (:id, :store_id, :table_number, :status,
                                             :total_amount, :discount_amount, :final_amount,
                                             :order_time, 'keruyun', '{}', NOW(), NOW())
                                        ON CONFLICT (id) DO NOTHING
                                    """),
                                    {
                                        "id": str(oid),
                                        "store_id": store_id,
                                        "table_number": order.get("tableId") or order.get("table_id") or "",
                                        "status": str(order.get("status") or "completed"),
                                        "total_amount": int(
                                            float(order.get("totalAmount") or order.get("total_amount") or 0) * 100
                                        ),
                                        "discount_amount": int(
                                            float(order.get("discountAmount") or order.get("discount_amount") or 0) * 100
                                        ),
                                        "final_amount": int(
                                            float(
                                                order.get("paidAmount")
                                                or order.get("paid_amount")
                                                or order.get("finalAmount")
                                                or 0
                                            )
                                            * 100
                                        ),
                                        "order_time": order.get("orderTime") or order.get("order_time") or start_time,
                                    },
                                )
                                records_imported += 1
                            await session.commit()
                except Exception as e:
                    logger.warning(
                        "backfill.keruyun.day_failed", store_id=store_id, date=target_date.isoformat(), error=str(e)
                    )
                    errors.append({"date": target_date.isoformat(), "error": str(e)})

            async with get_db_session() as session:
                await _update_progress(
                    session,
                    total=backfill_days,
                    imported=records_imported,
                    failed=len(errors),
                    status="completed",
                    extra={
                        "adapter": adapter,
                        "days_pulled": backfill_days,
                        "records_imported": records_imported,
                        "errors": errors[:10],
                    },
                )
            logger.info("backfill.keruyun.done", store_id=store_id, records_imported=records_imported, errors=len(errors))
            return {
                "success": True,
                "store_id": store_id,
                "adapter": adapter,
                "records_imported": records_imported,
                "skipped": False,
                "errors": errors[:10],
            }

        # ── 品智适配器 ─────────────────────────────────────────────────────────
        elif adapter == "pinzhi":
            from datetime import date, timedelta

            from packages.api_adapters.pinzhi.src.adapter import PinzhiAdapter

            base_url = credentials.get("base_url") or os.getenv("PINZHI_BASE_URL", "")
            token = credentials.get("token") or os.getenv(f"PINZHI_TOKEN_{store_id}") or os.getenv("PINZHI_TOKEN", "")
            if not base_url or not token:
                logger.warning("backfill.pinzhi.skipped", store_id=store_id, reason="PINZHI_BASE_URL/TOKEN not configured")
                async with get_db_session() as session:
                    await _update_progress(session, 0, 0, 0, "skipped", {"reason": "PINZHI_BASE_URL/TOKEN 未配置"})
                return {
                    "success": True,
                    "store_id": store_id,
                    "adapter": adapter,
                    "records_imported": 0,
                    "skipped": True,
                    "errors": [],
                }

            ognid = credentials.get("ognid") or os.getenv(f"PINZHI_OGNID_{store_id}") or os.getenv("PINZHI_OGNID", store_id)
            adapter_instance = PinzhiAdapter(config={"base_url": base_url, "token": token})

            backfill_days = int(os.getenv("ONBOARDING_BACKFILL_DAYS", "30"))
            today = date.today()
            records_imported = 0
            errors: list = []
            page_size = 50

            async with get_db_session() as session:
                await _update_progress(session, backfill_days, 0, 0, "in_progress", {"adapter": adapter, "phase": "pulling"})

            for i in range(backfill_days):
                target_date = today - timedelta(days=i + 1)
                date_str = target_date.strftime("%Y-%m-%d")
                try:
                    page = 1
                    while True:
                        orders = await adapter_instance.query_orders(
                            ognid=ognid,
                            begin_date=date_str,
                            end_date=date_str,
                            page_index=page,
                            page_size=page_size,
                        )
                        if not orders:
                            break
                        async with get_db_session() as session:
                            for order in orders:
                                oid = order.get("orderId") or order.get("order_id") or order.get("id")
                                if not oid:
                                    continue
                                await session.execute(
                                    _text("""
                                        INSERT INTO orders
                                            (id, store_id, table_number, status,
                                             total_amount, discount_amount, final_amount,
                                             order_time, source_system, order_metadata,
                                             created_at, updated_at)
                                        VALUES
                                            (:id, :store_id, :table_number, :status,
                                             :total_amount, :discount_amount, :final_amount,
                                             :order_time, 'pinzhi', '{}', NOW(), NOW())
                                        ON CONFLICT (id) DO NOTHING
                                    """),
                                    {
                                        "id": str(oid),
                                        "store_id": store_id,
                                        "table_number": order.get("tableId") or order.get("table_id") or "",
                                        "status": str(order.get("orderStatus") or order.get("status") or "completed"),
                                        "total_amount": int(
                                            float(order.get("totalAmount") or order.get("total_amount") or 0) * 100
                                        ),
                                        "discount_amount": int(
                                            float(order.get("discountAmount") or order.get("discount_amount") or 0) * 100
                                        ),
                                        "final_amount": int(
                                            float(
                                                order.get("actualAmount")
                                                or order.get("paidAmount")
                                                or order.get("final_amount")
                                                or 0
                                            )
                                            * 100
                                        ),
                                        "order_time": order.get("orderTime") or order.get("order_time") or date_str,
                                    },
                                )
                                records_imported += 1
                            await session.commit()
                        if len(orders) < page_size:
                            break
                        page += 1
                except Exception as e:
                    logger.warning("backfill.pinzhi.day_failed", store_id=store_id, date=date_str, error=str(e))
                    errors.append({"date": date_str, "error": str(e)})

            async with get_db_session() as session:
                await _update_progress(
                    session,
                    total=backfill_days,
                    imported=records_imported,
                    failed=len(errors),
                    status="completed",
                    extra={
                        "adapter": adapter,
                        "days_pulled": backfill_days,
                        "records_imported": records_imported,
                        "errors": errors[:10],
                    },
                )
            logger.info("backfill.pinzhi.done", store_id=store_id, records_imported=records_imported, errors=len(errors))
            return {
                "success": True,
                "store_id": store_id,
                "adapter": adapter,
                "records_imported": records_imported,
                "skipped": False,
                "errors": errors[:10],
            }

        # ── 美团SAAS / 奥琦玮 / 一订 — 不支持历史批量回灌 ──────────────────────
        elif adapter in ("meituan", "meituan-saas"):
            # 美团SAAS只提供单单查询接口（query_order by order_id/day_seq），
            # 无法按日期批量拉取历史订单，标记 skipped。
            logger.warning(
                "backfill.meituan.skipped", store_id=store_id, reason="meituan-saas adapter has no batch historical pull API"
            )
            async with get_db_session() as session:
                await _update_progress(
                    session, 0, 0, 0, "skipped", {"adapter": adapter, "reason": "美团SAAS仅支持单单查询，不支持历史批量回灌"}
                )
            return {
                "success": True,
                "store_id": store_id,
                "adapter": adapter,
                "records_imported": 0,
                "skipped": True,
                "errors": [],
            }

        elif adapter == "aoqiwei":
            # 奥琦玮为上传推送模型（POS→奥琦玮），无历史订单拉取接口，标记 skipped。
            logger.warning(
                "backfill.aoqiwei.skipped",
                store_id=store_id,
                reason="aoqiwei is an upload-push model with no historical pull API",
            )
            async with get_db_session() as session:
                await _update_progress(
                    session, 0, 0, 0, "skipped", {"adapter": adapter, "reason": "奥琦玮为上传推送模型，无历史数据拉取接口"}
                )
            return {
                "success": True,
                "store_id": store_id,
                "adapter": adapter,
                "records_imported": 0,
                "skipped": True,
                "errors": [],
            }

        elif adapter == "yiding":
            # 一订为预订管理系统，非POS订单系统，orders表回灌不适用，标记 skipped。
            logger.warning(
                "backfill.yiding.skipped", store_id=store_id, reason="yiding is a reservation system, not a POS order system"
            )
            async with get_db_session() as session:
                await _update_progress(
                    session, 0, 0, 0, "skipped", {"adapter": adapter, "reason": "一订为预订系统，不适用于orders表历史回灌"}
                )
            return {
                "success": True,
                "store_id": store_id,
                "adapter": adapter,
                "records_imported": 0,
                "skipped": True,
                "errors": [],
            }

        # ── 未知适配器 ─────────────────────────────────────────────────────────
        logger.warning(
            "backfill.adapter_not_implemented",
            store_id=store_id,
            adapter=adapter,
        )
        async with get_db_session() as session:
            await _update_progress(
                session,
                0,
                0,
                0,
                "skipped",
                {"adapter": adapter, "reason": f"{adapter} 历史回灌尚未集成"},
            )
        return {
            "success": True,
            "store_id": store_id,
            "adapter": adapter,
            "records_imported": 0,
            "skipped": True,
            "errors": [],
        }

    try:
        return asyncio.run(_run())
    except Exception as e:
        logger.error("backfill.failed", store_id=store_id, adapter=adapter, error=str(e))
        raise self.retry(exc=e)


# ============================================================
# P3: OpsAgent 定时巡检 + 企微 P0 告警推送
# ============================================================


def _build_ops_alert_markdown(store_id: str, store_name: str, dashboard: Dict[str, Any]) -> str:
    """将 get_store_dashboard 结果格式化为企微 Markdown 告警消息。"""
    score = dashboard.get("overall_score", 0)
    status = dashboard.get("overall_status", "unknown")
    active = dashboard.get("active_alerts", 0)
    layers = dashboard.get("layers", {})

    l1 = layers.get("l1_device", {})
    l2 = layers.get("l2_network", {})
    l3 = layers.get("l3_system", {})
    fs = dashboard.get("food_safety", {})

    label = store_name or store_id
    status_emoji = "🔴" if status == "critical" else "🟡"

    lines = [
        f"{status_emoji} **【运维P0告警】{label}**",
        f"> 健康分：**{score}/100** | 活跃告警：{active} 条",
        "",
        f"**L1 设备层**：{l1.get('score', '-')} 分，告警 {l1.get('alert_count', 0)} 条",
        f"**L2 网络层**：{l2.get('score', '-')} 分，" f"可用率 {l2.get('availability_pct', 100):.1f}%",
        f"**L3 系统层**：{l3.get('score', '-')} 分，"
        f"宕机系统 {l3.get('down_systems', 0)} 个" + (f"（其中 P0 = {l3['p0_down']} 个）" if l3.get("p0_down") else ""),
    ]

    down_list = l3.get("down_list") or []
    if down_list:
        lines.append(f"**已停服**：{', '.join(down_list[:5])}")

    violations = fs.get("violations", 0)
    if violations:
        rate = fs.get("compliance_rate_pct", 0)
        lines.append(f"**食安违规**：{violations} 条（合规率 {rate:.1f}%）")

    lines += ["", f"*巡检时间：{dashboard.get('generated_at', '')}*"]
    return "\n".join(lines)


@celery_app.task(bind=True, max_retries=2, default_retry_delay=60)
def ops_patrol(self, store_id: str = None):
    """
    OpsAgent 定时巡检（默认每15分钟）。

    对所有（或指定）门店执行 get_store_dashboard；
    当 overall_status == 'critical' 或存在 P0 系统宕机时，
    立即通过企微 Markdown 推送 P0 告警给门店负责人。
    warning 状态仅记录日志，不推送（避免告警风暴）。
    """

    async def _run():
        from sqlalchemy import select
        from src.core.database import get_db_session
        from src.models.store import Store
        from src.services.ops_monitor_service import OpsMonitorService
        from src.services.wechat_service import wechat_service

        svc = OpsMonitorService()

        async with get_db_session() as db:
            if store_id:
                stores = [(store_id, "")]
            else:
                rows = (await db.execute(select(Store.id, Store.name))).all()
                stores = [(str(r.id), r.name or "") for r in rows]

            pushed = 0
            for sid, sname in stores:
                try:
                    dashboard = await svc.get_store_dashboard(db, store_id=sid, window_minutes=15)
                    status = dashboard.get("overall_status", "healthy")
                    l3 = dashboard.get("layers", {}).get("l3_system", {})
                    p0_down = l3.get("p0_down", 0)

                    # 仅 critical 或 P0 系统宕机时推送
                    if status != "critical" and p0_down == 0:
                        logger.debug("ops_patrol.ok", store_id=sid, status=status)
                        continue

                    markdown = _build_ops_alert_markdown(sid, sname, dashboard)
                    recipient = _get_store_recipient(sid)
                    await wechat_service.send_markdown_message(
                        content=markdown,
                        touser=recipient,
                    )
                    pushed += 1
                    logger.info(
                        "ops_patrol.alert_pushed",
                        store_id=sid,
                        status=status,
                        p0_down=p0_down,
                        recipient=recipient,
                    )
                except Exception as exc:
                    logger.warning("ops_patrol.store_failed", store_id=sid, error=str(exc))

        logger.info("ops_patrol.done", total_stores=len(stores), pushed=pushed)
        return {"stores_checked": len(stores), "alerts_pushed": pushed}

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("ops_patrol.failed", error=str(exc))
        raise self.retry(exc=exc)


# ============================================================
# P5: Agent 间通信协议 — 异步分发任务
# ============================================================


@celery_app.task(bind=True, max_retries=2, default_retry_delay=30)
def dispatch_agent_message(
    self,
    from_agent: str,
    to_agent: str,
    action: str,
    payload: Dict[str, Any],
    store_id: str = "",
    priority: int = 5,
    trace_id: str = "",
    msg_id: str = "",
) -> Dict[str, Any]:
    """
    即发即忘模式下的 Agent 消息异步分发（AgentBus.fire_and_forget 的 Celery 后端）。

    收到消息后实例化目标 Agent 并执行 action，
    结果写入日志（不推送，调用方不等待回复）。
    """

    async def _run():
        from ..core.agent_bus import AgentBus, AgentMessage

        msg = AgentMessage(
            from_agent=from_agent,
            to_agent=to_agent,
            action=action,
            payload=payload,
            store_id=store_id,
            priority=priority,
            msg_id=msg_id or str(__import__("uuid").uuid4()),
        )
        if trace_id:
            msg.trace_id = trace_id

        bus = AgentBus.get()
        reply = await bus.send(msg)

        logger.info(
            "dispatch_agent_message.done",
            from_agent=from_agent,
            to_agent=to_agent,
            action=action,
            success=reply.success,
            trace_id=msg.trace_id,
        )
        return {
            "success": reply.success,
            "from_agent": from_agent,
            "to_agent": to_agent,
            "action": action,
            "error": reply.error,
            "trace_id": msg.trace_id,
        }

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("dispatch_agent_message.failed", to_agent=to_agent, action=action, error=str(exc))
        raise self.retry(exc=exc)


# ── Marketing: 每日批量挽回流失客户 ──────────────────────────────────────────────


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=2,
    default_retry_delay=120,
)
def marketing_auto_outreach(self, store_id: str = None) -> Dict[str, Any]:
    """
    每日批量企微挽回（10:30 自动执行）。
    遍历所有活跃门店（或指定门店），对 at_risk/lost 客户发送个性化挽回消息。
    通过 FrequencyCapEngine 控制发送频次，避免骚扰。
    """

    async def _run():
        from datetime import datetime

        from ..core.database import get_db_session
        from ..models.store import Store
        from ..services.marketing_agent_service import MarketingAgentService

        results = []
        async with get_db_session() as session:
            if store_id:
                store_ids = [store_id]
            else:
                # 取最近30天有订单的门店
                from datetime import timedelta

                from sqlalchemy import distinct, func, select

                from ..models.order import Order

                cutoff = datetime.utcnow() - timedelta(days=30)
                rows = (await session.execute(select(distinct(Order.store_id)).where(Order.created_at >= cutoff))).all()
                store_ids = [r[0] for r in rows if r[0]]

        svc = MarketingAgentService(db=None)
        for sid in store_ids:
            try:
                result = await svc.trigger_batch_churn_recovery(sid, dry_run=False)
                results.append(result)
                logger.info("marketing_auto_outreach.store_done", store_id=sid, sent=result["sent"])
            except Exception as e:
                logger.warning("marketing_auto_outreach.store_failed", store_id=sid, error=str(e))
                results.append({"store_id": sid, "error": str(e)})

        return {"stores_processed": len(results), "results": results}

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("marketing_auto_outreach.failed", error=str(exc))
        raise self.retry(exc=exc)


@celery_app.task(
    base=CallbackTask,
    bind=True,
    name="tasks.check_edge_hub_heartbeats",
    max_retries=2,
    default_retry_delay=60,
    autoretry_for=(Exception,),
    retry_backoff=True,
)
def check_edge_hub_heartbeats(self) -> Dict[str, Any]:
    """
    每 3 分钟扫描全平台边缘主机心跳。

    离线逻辑（last_heartbeat 超过阈值）：
      - 将主机 status 设为 offline
      - 若无同类 open EdgeAlert(hub_offline)，则创建 P1 告警

    恢复逻辑（last_heartbeat 在阈值内，但 status 为 offline）：
      - 将主机 status 设为 online
      - 自动 RESOLVE 该主机全部 open hub_offline 告警

    返回：{offline_marked, recovered, alerts_created, alerts_resolved}

    调度建议：beat_schedule 中设置 crontab(minute="*/3")
    """
    import uuid as _uuid
    from datetime import timedelta

    OFFLINE_THRESHOLD_MINUTES = int(os.getenv("EDGE_HUB_OFFLINE_THRESHOLD_MIN", "5"))

    async def _run():
        from sqlalchemy import and_, select, update
        from src.core.database import AsyncSessionLocal
        from src.models.edge_hub import AlertLevel, AlertStatus, EdgeAlert, EdgeHub, HubStatus
        from src.models.user import User, UserRole
        from src.services.wechat_alert_service import wechat_alert_service

        offline_marked = 0
        recovered = 0
        alerts_created = 0
        alerts_resolved = 0
        now = __import__("datetime").datetime.utcnow()
        stale_cutoff = now - timedelta(minutes=OFFLINE_THRESHOLD_MINUTES)

        async with AsyncSessionLocal() as db:
            hubs = (await db.execute(select(EdgeHub))).scalars().all()

            for hub in hubs:
                heartbeat_stale = hub.last_heartbeat is None or hub.last_heartbeat < stale_cutoff

                if heartbeat_stale and hub.status != HubStatus.OFFLINE:
                    # ── 心跳超时 → 标记离线 + 创建 P1 告警 ─────────────────
                    hub.status = HubStatus.OFFLINE
                    offline_marked += 1

                    existing = (
                        await db.execute(
                            select(EdgeAlert).where(
                                and_(
                                    EdgeAlert.hub_id == hub.id,
                                    EdgeAlert.alert_type == "hub_offline",
                                    EdgeAlert.status == AlertStatus.OPEN,
                                )
                            )
                        )
                    ).scalar_one_or_none()

                    if not existing:
                        db.add(
                            EdgeAlert(
                                id=str(_uuid.uuid4()),
                                hub_id=hub.id,
                                store_id=hub.store_id,
                                level=AlertLevel.P1,
                                alert_type="hub_offline",
                                message=f"边缘主机 {hub.hub_code} 心跳超时（>{OFFLINE_THRESHOLD_MINUTES}min）",
                                status=AlertStatus.OPEN,
                            )
                        )
                        alerts_created += 1
                        logger.warning(
                            "edge_hub.heartbeat_lost",
                            hub_id=hub.id,
                            hub_code=hub.hub_code,
                            store_id=hub.store_id,
                        )

                        # ── 推送企微通知给对应门店店长 ───────────────────────
                        try:
                            managers = (
                                (
                                    await db.execute(
                                        select(User).where(
                                            User.store_id == hub.store_id,
                                            User.is_active == True,
                                            User.role.in_([UserRole.STORE_MANAGER, UserRole.ADMIN]),
                                            User.wechat_user_id.isnot(None),
                                        )
                                    )
                                )
                                .scalars()
                                .all()
                            )
                            recipient_ids = [m.wechat_user_id for m in managers]
                            if recipient_ids:
                                await wechat_alert_service.send_hardware_alert(
                                    hub_id=hub.id,
                                    hub_code=hub.hub_code,
                                    store_id=hub.store_id,
                                    alert_type="hub_offline",
                                    recipient_ids=recipient_ids,
                                    extra={"last_heartbeat": hub.last_heartbeat},
                                )
                        except Exception as _wechat_exc:
                            logger.error(
                                "edge_hub.wechat_push_failed",
                                hub_id=hub.id,
                                error=str(_wechat_exc),
                            )

                elif not heartbeat_stale and hub.status == HubStatus.OFFLINE:
                    # ── 心跳恢复 → 标记在线 + 自动 resolve 告警 ──────────────
                    hub.status = HubStatus.ONLINE
                    recovered += 1

                    open_alerts = (
                        (
                            await db.execute(
                                select(EdgeAlert).where(
                                    and_(
                                        EdgeAlert.hub_id == hub.id,
                                        EdgeAlert.alert_type == "hub_offline",
                                        EdgeAlert.status == AlertStatus.OPEN,
                                    )
                                )
                            )
                        )
                        .scalars()
                        .all()
                    )

                    for alert in open_alerts:
                        alert.status = AlertStatus.RESOLVED
                        alert.resolved_at = now
                        alert.resolved_by = "system"
                        alerts_resolved += 1

                    if recovered:
                        logger.info(
                            "edge_hub.heartbeat_recovered",
                            hub_id=hub.id,
                            hub_code=hub.hub_code,
                            store_id=hub.store_id,
                        )

            await db.commit()

        return {
            "success": True,
            "offline_marked": offline_marked,
            "recovered": recovered,
            "alerts_created": alerts_created,
            "alerts_resolved": alerts_resolved,
        }

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("check_edge_hub_heartbeats.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── SignalBus 周期扫描任务 ─────────────────────────────────────────────────────


@celery_app.task(
    bind=True,
    max_retries=2,
    default_retry_delay=60,
)
def run_signal_bus_scan(self) -> Dict[str, Any]:
    """
    SignalBus 周期扫描：每2小时遍历全部门店，执行3条核心路由规则
      ① 差评 → 私域修复旅程
      ② 临期/低库存 → 废料预警推送
      ③ 大桌预订(≥6人) → 裂变场景识别
    """

    async def _run():
        from sqlalchemy import text

        from ..core.database import AsyncSessionLocal
        from ..services.signal_bus import run_periodic_scan

        results = []
        async with AsyncSessionLocal() as db:
            try:
                rows = (await db.execute(text("SELECT id FROM stores WHERE is_active = true"))).fetchall()
                store_ids = [r[0] for r in rows]
            except Exception:
                store_ids = ["store_001"]  # 降级到默认门店

            for sid in store_ids:
                try:
                    r = await run_periodic_scan(sid, db)
                    results.append(r)
                except Exception as exc:
                    logger.warning("signal_bus.task.store_failed", store_id=sid, error=str(exc))

        total = sum(r.get("total_routed", 0) for r in results)
        logger.info("signal_bus.task.done", stores=len(results), total_routed=total)
        return {"stores": len(results), "total_routed": total, "details": results}

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("signal_bus.task.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── 店长版每日简报任务 ────────────────────────────────────────────────────────


@celery_app.task(
    bind=True,
    max_retries=2,
    default_retry_delay=60,
)
def push_sm_daily_briefing(self) -> Dict[str, Any]:
    """
    店长版每日简报：08:00 推送
      - 私域健康分（今日综合得分）
      - Top3 决策（¥预期影响 + 置信度）
      - 昨日经营快照（营收/成本率/损耗）
      - 流失预警（高风险会员数）
      - 今日1条核心行动建议
    """

    async def _run():
        from sqlalchemy import text

        from ..core.database import AsyncSessionLocal
        from ..services.store_manager_briefing_service import push_briefing

        results = []
        async with AsyncSessionLocal() as db:
            try:
                rows = (await db.execute(text("SELECT id FROM stores WHERE is_active = true"))).fetchall()
                store_ids = [r[0] for r in rows]
            except Exception:
                store_ids = ["store_001"]

            for sid in store_ids:
                try:
                    r = await push_briefing(sid, db)
                    results.append(
                        {
                            "store_id": sid,
                            "pushed": r["pushed"],
                            "score": r["briefing"]["health"]["total_score"],
                        }
                    )
                except Exception as exc:
                    logger.warning("push_sm_briefing.store_failed", store_id=sid, error=str(exc))

        total_pushed = sum(1 for r in results if r.get("pushed"))
        logger.info("push_sm_briefing.done", stores=len(results), pushed=total_pushed)
        return {"stores": len(results), "pushed": total_pushed, "details": results}

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("push_sm_briefing.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── 老板多店版简报任务 ────────────────────────────────────────────────────────


@celery_app.task(
    bind=True,
    max_retries=2,
    default_retry_delay=60,
)
def push_hq_daily_briefing(self) -> Dict[str, Any]:
    """
    老板多店版每日简报：08:10 推送（晚于店长版 5 分钟）
      - 各门店健康分排名（红绿灯）
      - 预警门店（< 50 分）
      - 全局 Top3 决策
      - 昨日全店合并营收/均成本率
    """

    async def _run():
        from ..core.database import AsyncSessionLocal
        from ..services.hq_briefing_service import push_hq_briefing

        async with AsyncSessionLocal() as db:
            result = await push_hq_briefing(db, dry_run=False)

        logger.info(
            "push_hq_briefing.done",
            stores=result["briefing"].get("store_count", 0),
            alerts=len(result["briefing"].get("alerts", [])),
            pushed=result["pushed"],
        )
        return {
            "store_count": result["briefing"].get("store_count", 0),
            "pushed": result["pushed"],
        }

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("push_hq_briefing.failed", error=str(exc))
        raise self.retry(exc=exc)


# ============================================================
# Sprint 1 CDP: POS拉取后回填 consumer_id
# ============================================================
@celery_app.task(bind=True, max_retries=2, default_retry_delay=120)
def cdp_sync_consumer_ids(self) -> Dict[str, Any]:
    """
    POS数据拉取后，为所有未解析的订单/预订回填 consumer_id。
    每日 02:30 执行（紧跟 01:30 品智 + 02:00 天财拉取之后）。

    Sprint 1 KPI: consumer_id 填充率 ≥ 80%
    """

    async def _run():
        from src.core.database import get_db_session
        from src.services.cdp_sync_service import cdp_sync_service

        async with get_db_session() as session:
            result = await cdp_sync_service.sync_all_stores(session, batch_size=500)
            await session.commit()

        logger.info("cdp_sync_consumer_ids.done", **result)
        return result

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("cdp_sync_consumer_ids.failed", error=str(exc))
        raise self.retry(exc=exc)


# ============================================================
# Sprint 2 CDP: consumer_id 驱动的 RFM 重算
# ============================================================
@celery_app.task(bind=True, max_retries=2, default_retry_delay=120)
def cdp_rfm_recalculate(self) -> Dict[str, Any]:
    """
    基于 consumer_id 重算全量 RFM。
    每日 03:00 执行（02:30 consumer_id 回填完成后）。

    步骤：
    1. 先回填 PrivateDomainMember → ConsumerIdentity 链接
    2. 再基于 consumer_id 重算 RFM
    3. 计算偏差率，验证 < 5%

    Sprint 2 KPI: RFM 偏差 < 5%
    """

    async def _run():
        from src.core.database import get_db_session
        from src.services.cdp_rfm_service import cdp_rfm_service

        async with get_db_session() as session:
            # Step 1: 回填 member → consumer link
            link_result = await cdp_rfm_service.backfill_members(session)
            # Step 2: 重算 RFM
            rfm_result = await cdp_rfm_service.recalculate_all(session)
            # Step 3: 偏差校验
            deviation = await cdp_rfm_service.compute_deviation(session)
            await session.commit()

        result = {
            "link": link_result,
            "rfm": rfm_result,
            "deviation": deviation,
        }
        logger.info("cdp_rfm_recalculate.done", **result)
        return result

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("cdp_rfm_recalculate.failed", error=str(exc))
        raise self.retry(exc=exc)


# ============================================================
# Sprint 3 MemberAgent: 沉睡会员自动唤醒扫描
# ============================================================
@celery_app.task(bind=True, max_retries=2, default_retry_delay=120)
def member_agent_dormant_sweep(self) -> Dict[str, Any]:
    """
    自动扫描所有门店的沉睡会员并触发唤醒旅程。
    每日 06:30 执行（RFM 重算 03:00 完成后，开店前触发）。

    流程：
    1. 查询所有有沉睡会员的门店
    2. 每店批量触发 dormant_wakeup 旅程（每店最多 10 条/天）
    3. 汇总结果，追踪 Sprint 3 KPI: ≥50条/周

    每店每天限 10 条，避免集中骚扰；7天累计 ≥50 达标。
    """

    async def _run():
        from sqlalchemy import func, select
        from src.core.database import get_db_session
        from src.models.private_domain import PrivateDomainMember
        from src.services.member_agent_service import member_agent_service

        async with get_db_session() as session:
            # Step 1: 找到有沉睡会员的门店
            stmt = (
                select(PrivateDomainMember.store_id)
                .where(
                    PrivateDomainMember.is_active.is_(True),
                    PrivateDomainMember.consumer_id.isnot(None),
                    PrivateDomainMember.recency_days >= 30,
                )
                .group_by(PrivateDomainMember.store_id)
            )
            result = await session.execute(stmt)
            store_ids = [row[0] for row in result.all()]

            # Step 2: 每店触发唤醒（每店每天最多 10 条）
            total = {"stores": 0, "scanned": 0, "eligible": 0, "triggered": 0}
            for sid in store_ids:
                try:
                    r = await member_agent_service.batch_trigger_wakeup(
                        session,
                        sid,
                        min_recency_days=30,
                        max_count=10,
                        dry_run=False,
                    )
                    total["stores"] += 1
                    total["scanned"] += r.get("scanned", 0)
                    total["eligible"] += r.get("eligible", 0)
                    total["triggered"] += r.get("triggered", 0)
                except Exception as e:
                    logger.warning(
                        "member_agent_dormant_sweep.store_failed",
                        store_id=sid,
                        error=str(e),
                    )

            await session.commit()

        logger.info("member_agent_dormant_sweep.done", **total)
        return total

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("member_agent_dormant_sweep.failed", error=str(exc))
        raise self.retry(exc=exc)


# ============================================================
# Sprint 4: 增收月报自动生成（每月1日）
# ============================================================
@celery_app.task(bind=True, max_retries=2, default_retry_delay=300)
def revenue_growth_monthly_report(self) -> Dict[str, Any]:
    """
    每月1日 08:00 自动生成上月增收月报。

    遍历所有门店，生成各 Agent ¥贡献汇总。
    """

    async def _run():
        from sqlalchemy import func, select
        from src.core.database import get_db_session
        from src.models.order import Order
        from src.services.revenue_growth_service import revenue_growth_service

        async with get_db_session() as session:
            # 查询所有有订单的门店
            stmt = select(Order.store_id).group_by(Order.store_id)
            result = await session.execute(stmt)
            store_ids = [row[0] for row in result.all()]

            reports = []
            for sid in store_ids:
                try:
                    report = await revenue_growth_service.generate_monthly_report(
                        session,
                        sid,
                        month_offset=-1,
                    )
                    reports.append(report)
                    logger.info(
                        "revenue_growth_report.store_done",
                        store_id=sid,
                        delta_yuan=report["revenue"]["delta_yuan"],
                        agent_total=report["agent_contribution"]["agent_total_yuan"],
                    )
                except Exception as e:
                    logger.warning(
                        "revenue_growth_report.store_failed",
                        store_id=sid,
                        error=str(e),
                    )

        total = {
            "stores": len(reports),
            "total_delta_yuan": sum(r["revenue"]["delta_yuan"] for r in reports),
            "total_agent_contribution_yuan": sum(r["agent_contribution"]["agent_total_yuan"] for r in reports),
        }
        logger.info("revenue_growth_monthly_report.done", **total)
        return total

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("revenue_growth_monthly_report.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── CDP 全量回填管道 ────────────────────────────────────────────


@celery_app.task(bind=True, max_retries=1, default_retry_delay=300)
def cdp_full_backfill(self, store_id=None, batch_size=500) -> Dict[str, Any]:
    """
    CDP 全量回填管道（异步版）

    步骤：订单回填 → 会员回填 → RFM重算 → 偏差校验
    """

    async def _run():
        from src.core.database import get_db_session
        from src.services.cdp_monitor_service import cdp_monitor_service

        async with get_db_session() as db:
            result = await cdp_monitor_service.run_full_backfill(
                db,
                store_id=store_id,
                batch_size=batch_size,
            )
            logger.info(
                "cdp_full_backfill.done",
                store_id=store_id,
                kpi_all_met=result["kpi_summary"]["all_met"],
            )
            return result

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("cdp_full_backfill.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── P2: 决策效果闭环评估 ──────────────────────────────────────────────────


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=2,
    default_retry_delay=300,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_jitter=True,
)
def evaluate_decision_effects(self) -> Dict[str, Any]:
    """
    每日 03:30 扫描已执行但未评估的 DecisionLog，
    根据 decision_type 匹配评估策略，计算偏差，更新 outcome/trust_score。
    """

    async def _run():
        from src.core.database import get_db_session
        from src.services.effect_evaluator import EffectEvaluator

        async with get_db_session() as db:
            evaluator = EffectEvaluator(db)
            result = await evaluator.run_evaluation_sweep()
            logger.info("evaluate_decision_effects.done", **result)

            # 评估完成后推送企微通知（仅在有评估结果时）
            evaluated = result.get("evaluated", 0)
            if evaluated > 0:
                try:
                    from src.services.wechat_service import wechat_service

                    if wechat_service:
                        success = result.get("success", 0)
                        failure = result.get("failure", 0)
                        msg = (
                            f"【AI效果评估完成】\n"
                            f"评估决策 {evaluated} 条\n"
                            f"成功 {success} | 失败 {failure}\n"
                            f"查看详情：/sm/decisions"
                        )
                        await wechat_service.send_text(msg)
                except Exception as push_exc:
                    logger.warning("evaluate_decision_effects.push_failed", error=str(push_exc))

            return result

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("evaluate_decision_effects.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── 奥琦玮供应链：每日库存 + 采购单拉取 ──────────────────────────────────────


@celery_app.task(
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY_LONG", "300")),
)
def pull_aoqiwei_daily_supply(self) -> Dict[str, Any]:
    """
    拉取奥琦玮供应链昨日采购入库单 + 当前库存快照，写入 DB。

    注意：奥琦玮 POS 订单为 push 模型（POS 向我方推送），无 pull API；
    此任务仅处理供应链（inventory/purchase）数据。

    环境变量：
      AOQIWEI_BASE_URL    — API 基础地址（默认 https://openapi.acescm.cn）
      AOQIWEI_APP_KEY     — AppKey（必填；缺失则跳过）
      AOQIWEI_APP_SECRET  — AppSecret（必填；缺失则跳过）
      AOQIWEI_SHOP_CODE   — 默认门店 shopCode
      AOQIWEI_SHOP_CODE_{store_id} — 门店级 shopCode（优先）

    Returns:
        {success, date, stores_processed, purchase_orders_saved, stock_snapshots, errors}
    """

    async def _run():
        from datetime import date, timedelta

        from sqlalchemy import select

        from ..core.database import get_db_session
        from ..models.store import Store

        app_key = os.getenv("AOQIWEI_APP_KEY", "")
        app_secret = os.getenv("AOQIWEI_APP_SECRET", "")

        if not app_key or not app_secret:
            logger.warning("aoqiwei_supply_pull.skipped", reason="AOQIWEI_APP_KEY or AOQIWEI_APP_SECRET not configured")
            return {
                "success": True,
                "skipped": True,
                "stores_processed": 0,
                "purchase_orders_saved": 0,
                "stock_snapshots": 0,
                "errors": [],
            }

        yesterday = (date.today() - timedelta(days=1)).strftime("%Y-%m-%d")
        base_url = os.getenv("AOQIWEI_BASE_URL", "https://openapi.acescm.cn")

        from packages.api_adapters.aoqiwei.src.adapter import AoqiweiAdapter

        adapter = AoqiweiAdapter(
            {
                "base_url": base_url,
                "app_key": app_key,
                "app_secret": app_secret,
                "timeout": int(os.getenv("AOQIWEI_TIMEOUT", "30")),
                "retry_times": int(os.getenv("AOQIWEI_RETRY_TIMES", "3")),
            }
        )

        stores_processed = 0
        purchase_orders_saved = 0
        stock_snapshots = 0
        errors = []

        try:
            async with get_db_session() as session:
                result = await session.execute(select(Store).where(Store.is_active.is_(True)))
                stores = result.scalars().all()

                for store in stores:
                    sid = str(store.id)
                    shop_code = (
                        os.getenv(f"AOQIWEI_SHOP_CODE_{sid}")
                        or os.getenv("AOQIWEI_SHOP_CODE", "")
                        or getattr(store, "code", None)
                        or sid
                    )

                    try:
                        # 1) 拉取昨日采购入库单
                        page = 1
                        store_po_count = 0
                        while True:
                            po_resp = await adapter.query_purchase_orders(
                                start_date=yesterday,
                                end_date=yesterday,
                                page=page,
                                page_size=100,
                            )
                            po_list = po_resp.get("list", []) if isinstance(po_resp, dict) else []
                            if not po_list:
                                break

                            for po in po_list:
                                order_no = po.get("purchaseOrderNo") or po.get("orderNo") or ""
                                depot_code = po.get("depotCode", "")
                                total_amount = float(po.get("totalAmount") or po.get("amount") or 0)
                                order_date = po.get("orderDate") or yesterday
                                raw_json = __import__("json").dumps(po, ensure_ascii=False)
                                # 写入 inventory_transactions 或专用采购表（降级：只记录摘要到 daily_summaries source='aoqiwei_supply'）
                                from sqlalchemy import text as _text

                                await session.execute(
                                    _text("""
                                        INSERT INTO daily_summaries
                                            (store_id, business_date, revenue_cents,
                                             order_count, source, raw_data, created_at)
                                        VALUES
                                            (:store_id, :business_date, :revenue_cents,
                                             :order_count, 'aoqiwei_supply', :raw_data, NOW())
                                        ON CONFLICT (store_id, business_date, source) DO UPDATE SET
                                            revenue_cents = EXCLUDED.revenue_cents + daily_summaries.revenue_cents,
                                            order_count   = EXCLUDED.order_count   + daily_summaries.order_count,
                                            raw_data      = EXCLUDED.raw_data,
                                            created_at    = NOW()
                                    """),
                                    {
                                        "store_id": sid,
                                        "business_date": order_date,
                                        "revenue_cents": int(total_amount * 100),
                                        "order_count": 1,
                                        "raw_data": raw_json,
                                    },
                                )
                                store_po_count += 1

                            total_count = po_resp.get("total", 0) if isinstance(po_resp, dict) else 0
                            if page * 100 >= total_count or len(po_list) < 100:
                                break
                            page += 1

                        purchase_orders_saved += store_po_count

                        # 2) 库存快照（当前库存，不依赖日期）
                        stock_list = await adapter.query_stock(shop_code=shop_code)
                        if stock_list:
                            stock_snapshots += len(stock_list)
                            # 库存数据写入 inventory_items（仅记录有 good_code 的行）
                            from sqlalchemy import text as _text2

                            for item in stock_list:
                                good_code = item.get("goodCode") or item.get("good_code") or ""
                                good_name = item.get("goodName") or item.get("good_name") or ""
                                qty = float(item.get("qty") or item.get("remainQty") or 0)
                                unit = item.get("unit") or ""
                                if not good_code:
                                    continue
                                await session.execute(
                                    _text2("""
                                        INSERT INTO inventory_items
                                            (store_id, sku_code, name, unit, quantity,
                                             category, low_stock_threshold, created_at, updated_at)
                                        VALUES
                                            (:store_id, :sku_code, :name, :unit, :qty,
                                             'aoqiwei', 0, NOW(), NOW())
                                        ON CONFLICT (store_id, sku_code) DO UPDATE SET
                                            quantity   = EXCLUDED.quantity,
                                            updated_at = NOW()
                                    """),
                                    {
                                        "store_id": sid,
                                        "sku_code": good_code,
                                        "name": good_name,
                                        "unit": unit,
                                        "qty": qty,
                                    },
                                )

                        stores_processed += 1
                        logger.info(
                            "aoqiwei_supply_pull.store_done",
                            store_id=sid,
                            purchase_orders=store_po_count,
                            stock_items=len(stock_list) if stock_list else 0,
                        )

                    except Exception as store_exc:
                        err_msg = f"store {sid}: {store_exc}"
                        errors.append(err_msg)
                        logger.error("aoqiwei_supply_pull.store_error", store_id=sid, error=str(store_exc))

                await session.commit()

        except Exception as exc:
            logger.error("aoqiwei_supply_pull.fatal", error=str(exc))
            errors.append(str(exc))

        return {
            "success": len(errors) == 0,
            "date": yesterday,
            "stores_processed": stores_processed,
            "purchase_orders_saved": purchase_orders_saved,
            "stock_snapshots": stock_snapshots,
            "errors": errors,
        }

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("pull_aoqiwei_daily_supply.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── 奥琦玮微生活会员：基于 POS 订单手机号增强会员数据 ──────────────────────────


@celery_app.task(
    bind=True,
    max_retries=int(os.getenv("CELERY_MAX_RETRIES", "3")),
    default_retry_delay=int(os.getenv("CELERY_RETRY_DELAY_LONG", "300")),
)
def enrich_members_from_aoqiwei_crm(self) -> Dict[str, Any]:
    """
    从奥琦玮微生活 CRM 拉取会员积分/余额/等级，写入 MemberSync 表。

    执行流程：
      1. 从 ExternalSystem 表查找 provider='aoqiwei_crm' 的已启用配置
      2. 从 orders.customer_phone 提取近 30 天有消费记录的会员手机号
      3. 逐个调用 CRM /member/info 补全会员数据
      4. 写入 member_syncs 表（支持 upsert）

    注意：奥琦玮 CRM 无批量导出 API，仅支持按手机号/卡号逐条查询。
    MemberSync 表记录单次同步结果，重复手机号只更新最新数据。

    环境变量（优先于 ExternalSystem 表配置）：
      AOQIWEI_CRM_APPID    — CRM AppID
      AOQIWEI_CRM_APPKEY   — CRM AppKey（签名密钥）
      AOQIWEI_CRM_BASE_URL — 默认 https://welcrm.com
    """

    async def _run():
        import json as _json
        from datetime import date, timedelta

        from sqlalchemy import select
        from sqlalchemy import text as _text

        from ..core.database import get_db_session
        from ..models.integration import ExternalSystem, IntegrationStatus

        appid = os.getenv("AOQIWEI_CRM_APPID", "")
        appkey = os.getenv("AOQIWEI_CRM_APPKEY", "")
        base_url = os.getenv("AOQIWEI_CRM_BASE_URL", "https://welcrm.com")

        members_enriched = 0
        members_not_found = 0
        errors: list = []

        async with get_db_session() as session:
            # 优先用 DB 中配置的凭证（如有），fallback 到环境变量
            sys_result = await session.execute(_text("""
                    SELECT api_key, api_secret, api_endpoint
                    FROM external_systems
                    WHERE provider = 'aoqiwei_crm'
                      AND status != 'inactive'
                    ORDER BY created_at ASC
                    LIMIT 1
                """))
            sys_row = sys_result.fetchone()
            if sys_row:
                appid = sys_row[0] or appid
                appkey = sys_row[1] or appkey
                base_url = sys_row[2] or base_url

            if not appid or not appkey:
                logger.warning(
                    "aoqiwei_crm_enrich.skipped",
                    reason="AOQIWEI_CRM_APPID 或 AOQIWEI_CRM_APPKEY 未配置（环境变量或DB均未找到）",
                )
                return {
                    "success": True,
                    "skipped": True,
                    "members_enriched": 0,
                    "members_not_found": 0,
                    "errors": [],
                }

            # 取近 30 天有手机号的会员（去重）
            thirty_days_ago = (date.today() - timedelta(days=30)).strftime("%Y-%m-%d")
            phones_result = await session.execute(
                _text("""
                    SELECT DISTINCT customer_phone
                    FROM orders
                    WHERE customer_phone IS NOT NULL
                      AND customer_phone <> ''
                      AND order_time >= :since
                    LIMIT 2000
                """),
                {"since": thirty_days_ago},
            )
            phones = [row[0] for row in phones_result.fetchall() if row[0]]

            if not phones:
                logger.info("aoqiwei_crm_enrich.no_phones", reason="近30天无含手机号的订单，跳过本次同步")
                return {
                    "success": True,
                    "skipped": True,
                    "members_enriched": 0,
                    "members_not_found": 0,
                    "errors": [],
                }

            logger.info("aoqiwei_crm_enrich.start", total_phones=len(phones))

            from packages.api_adapters.aoqiwei.src.crm_adapter import AoqiweiCrmAdapter

            adapter = AoqiweiCrmAdapter(
                {
                    "base_url": base_url,
                    "appid": appid,
                    "appkey": appkey,
                    "timeout": int(os.getenv("AOQIWEI_CRM_TIMEOUT", "15")),
                    "retry_times": 2,
                }
            )

            for phone in phones:
                try:
                    info = await adapter.get_member_info(mobile=phone)
                    if not info:
                        members_not_found += 1
                        continue

                    # 提取字段（适配奥琦玮 CRM 实际返回格式）
                    cno = info.get("cno") or info.get("card_no") or ""
                    name = info.get("name") or info.get("real_name") or ""
                    level = info.get("level_name") or info.get("level") or ""
                    points = int(info.get("credits") or info.get("points") or 0)
                    balance_fen = int(info.get("balance") or 0)
                    balance_yuan = round(balance_fen / 100, 2)
                    raw_json = _json.dumps(info, ensure_ascii=False)

                    await session.execute(
                        _text("""
                            INSERT INTO member_syncs
                                (id, system_id, member_id, external_member_id,
                                 phone, name, level, points, balance,
                                 sync_status, synced_at, raw_data,
                                 created_at, updated_at)
                            VALUES
                                (gen_random_uuid(),
                                 (SELECT id FROM external_systems
                                  WHERE provider='aoqiwei_crm' LIMIT 1),
                                 :phone, :cno,
                                 :phone, :name, :level, :points, :balance,
                                 'success', NOW(), :raw_data::jsonb,
                                 NOW(), NOW())
                            ON CONFLICT (phone, system_id) DO UPDATE SET
                                external_member_id = EXCLUDED.external_member_id,
                                name               = EXCLUDED.name,
                                level              = EXCLUDED.level,
                                points             = EXCLUDED.points,
                                balance            = EXCLUDED.balance,
                                sync_status        = 'success',
                                synced_at          = NOW(),
                                raw_data           = EXCLUDED.raw_data,
                                updated_at         = NOW()
                        """),
                        {
                            "phone": phone,
                            "cno": cno,
                            "name": name,
                            "level": level,
                            "points": points,
                            "balance": balance_yuan,
                            "raw_data": raw_json,
                        },
                    )
                    members_enriched += 1

                    # 限速：每 10 条请求休眠 200ms，避免触发 CRM 限流
                    if members_enriched % 10 == 0:
                        await asyncio.sleep(0.2)

                except Exception as phone_exc:
                    err_str = str(phone_exc)
                    # "会员不存在" 不算错误
                    not_found_keywords = ["会员不存在", "用户不存在", "member not found", "status=0", "status=2", "无效用户"]
                    if any(kw.lower() in err_str.lower() for kw in not_found_keywords):
                        members_not_found += 1
                    else:
                        errors.append(f"{phone}: {err_str}")
                        logger.warning("aoqiwei_crm_enrich.phone_error", phone=phone, error=err_str)

            await adapter.aclose()
            await session.commit()

        logger.info(
            "aoqiwei_crm_enrich.done",
            members_enriched=members_enriched,
            members_not_found=members_not_found,
            errors=len(errors),
        )
        return {
            "success": len(errors) == 0,
            "members_enriched": members_enriched,
            "members_not_found": members_not_found,
            "errors": errors[:20],  # 最多返回20条错误
        }

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("enrich_members_from_aoqiwei_crm.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── IM 通讯录定时同步 ─────────────────────────────────────────────────────


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=2,
    default_retry_delay=300,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_jitter=True,
)
def scheduled_im_roster_sync(self) -> Dict[str, Any]:
    """
    定时同步所有已启用 sync_enabled 的品牌 IM 通讯录。
    默认每日 02:00 执行（通过 beat_schedule 配置）。
    """

    async def _run():
        from sqlalchemy import and_, select
        from src.core.database import get_db_session
        from src.models.brand_im_config import BrandIMConfig
        from src.services.im_sync_service import IMSyncService

        results = []
        async with get_db_session() as db:
            configs = await db.execute(
                select(BrandIMConfig).where(
                    and_(
                        BrandIMConfig.sync_enabled.is_(True),
                        BrandIMConfig.is_active.is_(True),
                    )
                )
            )
            brand_configs = configs.scalars().all()

            for config in brand_configs:
                try:
                    sync_svc = IMSyncService(db)
                    result = await sync_svc.sync_roster(
                        brand_id=config.brand_id,
                        trigger="scheduled",
                    )
                    results.append(
                        {
                            "brand_id": config.brand_id,
                            "status": "success",
                            "added": result.get("added", 0),
                            "updated": result.get("updated", 0),
                            "disabled": result.get("disabled", 0),
                        }
                    )
                    logger.info(
                        "scheduled_im_sync.brand_done",
                        brand_id=config.brand_id,
                        added=result.get("added", 0),
                    )
                except Exception as exc:
                    results.append(
                        {
                            "brand_id": config.brand_id,
                            "status": "failed",
                            "error": str(exc),
                        }
                    )
                    logger.error(
                        "scheduled_im_sync.brand_failed",
                        brand_id=config.brand_id,
                        error=str(exc),
                    )

        summary = {
            "total_brands": len(results),
            "success": sum(1 for r in results if r["status"] == "success"),
            "failed": sum(1 for r in results if r["status"] == "failed"),
        }
        logger.info("scheduled_im_sync.done", **summary)
        return summary

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("scheduled_im_sync.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── 钉钉消息重试 ──────────────────────────────────────────────────────────


@celery_app.task(bind=True, max_retries=1)
def retry_failed_dingtalk_messages(self):
    """每5分钟从失败队列取出钉钉消息重试（最多3次）"""
    import asyncio

    async def _run():
        from src.services.dingtalk_service import dingtalk_service

        await dingtalk_service.retry_failed_messages(max_retries=3, batch_size=10)

    try:
        asyncio.run(_run())
    except Exception as exc:
        logger.warning("retry_failed_dingtalk_messages.error", error=str(exc))


# ── IM 考勤数据定时同步 ───────────────────────────────────────────────────


@celery_app.task(
    base=CallbackTask,
    bind=True,
    max_retries=2,
    default_retry_delay=300,
)
def scheduled_im_attendance_sync(self) -> Dict[str, Any]:
    """
    每日 06:00 同步昨日 IM 打卡数据到屯象OS考勤表。
    """
    from datetime import date, timedelta

    async def _run():
        from sqlalchemy import and_, select
        from src.core.database import get_db_session
        from src.models.brand_im_config import BrandIMConfig
        from src.services.im_attendance_sync import IMAttendanceSyncService

        yesterday = date.today() - timedelta(days=1)
        results = []

        async with get_db_session() as db:
            configs = await db.execute(
                select(BrandIMConfig).where(
                    and_(
                        BrandIMConfig.sync_enabled.is_(True),
                        BrandIMConfig.is_active.is_(True),
                    )
                )
            )
            for config in configs.scalars().all():
                try:
                    service = IMAttendanceSyncService(db)
                    result = await service.sync_attendance(
                        config.brand_id,
                        yesterday,
                        yesterday,
                    )
                    results.append({"brand_id": config.brand_id, **result})
                except Exception as exc:
                    results.append({"brand_id": config.brand_id, "error": str(exc)})
                    logger.warning("scheduled_attendance_sync.brand_failed", brand_id=config.brand_id, error=str(exc))

        summary = {
            "total_brands": len(results),
            "total_synced": sum(r.get("synced", 0) for r in results),
        }
        logger.info("scheduled_attendance_sync.done", **summary)
        return summary

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.error("scheduled_attendance_sync.failed", error=str(exc))
        raise self.retry(exc=exc)


# ── Phase 4: 入职引导提醒 ────────────────────────────────────────────────


@celery_app.task(bind=True, max_retries=1)
def remind_incomplete_onboarding(self) -> Dict[str, Any]:
    """
    每日 09:00 提醒入职超过7天但任务未完成的新员工。
    """

    async def _run():
        from src.core.database import get_db_session
        from src.services.im_onboarding_robot import IMOnboardingRobot

        async with get_db_session() as db:
            robot = IMOnboardingRobot(db)
            return await robot.remind_incomplete_onboarding(days_threshold=7)

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.warning("remind_onboarding.failed", error=str(exc))
        return {"error": str(exc)}


# ── Phase 4: 里程碑通知扫描 ──────────────────────────────────────────────


@celery_app.task(bind=True, max_retries=1)
def sweep_milestone_notifications(self) -> Dict[str, Any]:
    """
    每日 10:00 扫描未推送的里程碑，批量发送 IM 庆祝通知。
    """

    async def _run():
        from src.core.database import get_db_session
        from src.services.im_milestone_notifier import IMMilestoneNotifier

        async with get_db_session() as db:
            notifier = IMMilestoneNotifier(db)
            return await notifier.sweep_unnotified_milestones()

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.warning("sweep_milestones.failed", error=str(exc))
        return {"error": str(exc)}


@celery_app.task(name="check_compliance_alerts")
def check_compliance_alerts():
    """每日合规扫描：健康证/合同/身份证到期告警"""
    import asyncio

    from sqlalchemy import select
    from src.core.database import AsyncSessionLocal
    from src.models.store import Store
    from src.services.compliance_alert_service import ComplianceAlertService

    async def _run():
        async with AsyncSessionLocal() as db:
            # 扫描所有活跃门店
            result = await db.execute(select(Store.id).where(Store.is_active.is_(True)))
            store_ids = [r[0] for r in result.all()]

            total_alerts = 0
            for store_id in store_ids:
                try:
                    svc = ComplianceAlertService(store_id)
                    result = await svc.send_compliance_alerts(db)
                    if result.get("sent"):
                        total_alerts += result.get("sent_count", 0)
                except Exception as e:
                    logger.warning("compliance_check_failed", store_id=store_id, error=str(e))

            logger.info("compliance_check_completed", stores=len(store_ids), alerts_sent=total_alerts)

    asyncio.run(_run())


# ── W2-1: 审批超期检查 ────────────────────────────────────────────────


@celery_app.task(name="src.core.celery_tasks.run_decision_effect_reviews", bind=True, max_retries=1)
def run_decision_effect_reviews(self) -> Dict[str, Any]:
    """每日04:00 扫描已执行决策的30/60/90天效果回顾 — Palantir闭环核心"""

    async def _run():
        from src.core.database import get_db_session
        from src.services.decision_flywheel_service import DecisionFlywheelService

        service = DecisionFlywheelService()
        async with get_db_session() as db:
            result = await service.run_effect_reviews(db)
            logger.info(
                "decision_flywheel.effect_review.done",
                reviewed_30d=result.get("reviewed_30d", 0),
                reviewed_60d=result.get("reviewed_60d", 0),
                reviewed_90d=result.get("reviewed_90d", 0),
            )
            return result

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.warning("decision_flywheel.effect_review.failed", error=str(exc))
        return {"error": str(exc)}


@celery_app.task(name="check_approval_timeouts", bind=True, max_retries=1)
def check_approval_timeouts(self) -> Dict[str, Any]:
    """每小时检查超期审批 — 自动升级到下一级/推送催办通知"""

    async def _run():
        from src.core.database import get_db_session
        from src.services.approval_engine import approval_engine

        async with get_db_session() as db:
            result = await approval_engine.check_timeouts(db)
            logger.info(
                "approval_timeout_check.done",
                total=result.get("total_timed_out", 0),
                escalated=result.get("escalated", 0),
                reminded=result.get("reminded", 0),
            )
            return result

    try:
        return asyncio.run(_run())
    except Exception as exc:
        logger.warning("approval_timeout_check.failed", error=str(exc))
        return {"error": str(exc)}
