"""P2 Payroll Engine tests — SocialInsurance / Tax / PayrollService / HRExport

All tests use mocked AsyncSession — no real database required.
"""
import uuid
from datetime import date, datetime, timezone
from io import BytesIO
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from src.services.hr.social_insurance_service import SocialInsuranceService
from src.services.hr.tax_service import TaxService
from src.services.hr.payroll_service import PayrollService
from src.services.hr.hr_export_service import HRExportService

pytestmark = pytest.mark.asyncio


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_scalar_result(value):
    """Mock result whose scalar_one_or_none() returns value."""
    mock_result = MagicMock()
    mock_result.scalar_one_or_none.return_value = value
    return mock_result


def _make_scalars_result(values):
    """Mock result whose scalars().all() returns values."""
    mock_result = MagicMock()
    mock_scalars = MagicMock()
    mock_scalars.all.return_value = values
    mock_result.scalars.return_value = mock_scalars
    return mock_result


def _make_scalar_only_result(value):
    """Mock result whose scalar() returns value (for aggregate queries)."""
    mock_result = MagicMock()
    mock_result.scalar.return_value = value
    return mock_result


def _make_batch(batch_id=None, org_node_id="store-001", year=2026, month=3,
                status="draft"):
    batch = MagicMock()
    batch.id = batch_id or uuid.uuid4()
    batch.org_node_id = org_node_id
    batch.period_year = year
    batch.period_month = month
    batch.status = status
    batch.total_gross_fen = 0
    batch.total_net_fen = 0
    return batch


def _make_assignment(asn_id=None, org_node_id="store-001"):
    asn = MagicMock()
    asn.id = asn_id or uuid.uuid4()
    asn.org_node_id = org_node_id
    asn.status = "active"
    return asn


def _make_contract(assignment_id, pay_scheme=None):
    contract = MagicMock()
    contract.assignment_id = assignment_id
    contract.pay_scheme = pay_scheme or {"type": "fixed_monthly", "base_salary_fen": 600000}
    contract.valid_from = date(2025, 1, 1)
    contract.valid_to = None
    return contract


def _make_attendance(status="normal", work_minutes=480, overtime_minutes=0):
    att = MagicMock()
    att.status = status
    att.work_minutes = work_minutes
    att.overtime_minutes = overtime_minutes
    att.date = date(2026, 3, 1)
    return att


# ===========================================================================
# SocialInsuranceService tests (4)
# ===========================================================================

class TestSocialInsuranceService:

    def test_si_normal_salary(self):
        """6000元 → 各项按费率计算"""
        svc = SocialInsuranceService()
        result = svc.calculate_employee_portion(6000.0)
        assert result["base_yuan"] == 6000.0
        assert result["pension_yuan"] == 480.0      # 6000 * 0.08
        assert result["medical_yuan"] == 120.0       # 6000 * 0.02
        assert result["unemployment_yuan"] == 30.0   # 6000 * 0.005
        assert result["housing_fund_yuan"] == 300.0  # 6000 * 0.05
        assert result["total_yuan"] == 930.0

    def test_si_floor_capped(self):
        """2000元 → base被下限3800兜底"""
        svc = SocialInsuranceService()
        result = svc.calculate_employee_portion(2000.0)
        assert result["base_yuan"] == 3800.0
        assert result["pension_yuan"] == 304.0       # 3800 * 0.08
        assert result["total_yuan"] == round(
            3800 * 0.08 + 3800 * 0.02 + 3800 * 0.005 + 3800 * 0.05, 2
        )

    def test_si_ceiling_capped(self):
        """25000元 → base被上限19000封顶"""
        svc = SocialInsuranceService()
        result = svc.calculate_employee_portion(25000.0)
        assert result["base_yuan"] == 19000.0
        assert result["pension_yuan"] == 1520.0  # 19000 * 0.08

    def test_si_custom_housing_fund_rate(self):
        """自定义12%公积金比例"""
        svc = SocialInsuranceService()
        result = svc.calculate_employee_portion(10000.0, housing_fund_rate=0.12)
        assert result["housing_fund_yuan"] == 1200.0  # 10000 * 0.12
        # 养老800 + 医疗200 + 失业50 + 公积金1200 = 2250
        assert result["total_yuan"] == 2250.0


# ===========================================================================
# TaxService tests (5)
# ===========================================================================

class TestTaxService:

    def test_tax_below_threshold(self):
        """应纳税所得额<=0 → 税额=0"""
        svc = TaxService()
        assert svc.calculate_monthly_tax(0, 0) == 0.0
        assert svc.calculate_monthly_tax(0, -100) == 0.0

    def test_tax_bracket_1(self):
        """月应纳税所得额3000元 → 3%档"""
        svc = TaxService()
        tax = svc.calculate_monthly_tax(0, 3000)
        assert tax == 90.0  # 3000 * 0.03

    def test_tax_bracket_2(self):
        """月应纳税所得额10000元（累计10000 <= 36000）→ still in 3% bracket for 1st month"""
        svc = TaxService()
        tax = svc.calculate_monthly_tax(0, 10000)
        # 累计10000 <= 36000 → 10000 * 0.03 - 0 = 300
        assert tax == 300.0

    def test_tax_cumulative_jan_to_mar(self):
        """1月~3月逐月累进：每月taxable=10000"""
        svc = TaxService()
        # 1月：ytd=0, current=10000 → 累计10000 → 300
        tax_jan = svc.calculate_monthly_tax(0, 10000)
        assert tax_jan == 300.0

        # 2月：ytd=10000, current=10000 → 累计20000 → 20000*0.03=600, 减去已缴300=300
        tax_feb = svc.calculate_monthly_tax(10000, 10000)
        assert tax_feb == 300.0

        # 3月：ytd=20000, current=10000 → 累计30000 → 30000*0.03=900, 减去已缴600=300
        tax_mar = svc.calculate_monthly_tax(20000, 10000)
        assert tax_mar == 300.0

    def test_tax_high_income(self):
        """月应纳税所得额50000元（高收入，跨档）"""
        svc = TaxService()
        # 1月：ytd=0, current=50000 → 累计50000
        # 50000 在 36000~144000 档 → 50000*0.10-2520=2480
        tax = svc.calculate_monthly_tax(0, 50000)
        assert tax == 2480.0


# ===========================================================================
# PayrollService tests (5)
# ===========================================================================

class TestPayrollService:

    @pytest.fixture
    def svc(self):
        return PayrollService()

    @pytest.fixture
    def mock_session(self):
        session = AsyncMock()
        session.flush = AsyncMock()
        session.add = MagicMock()
        return session

    def _setup_calculate_mocks(self, session, batch, assignments, contract,
                                att_rows, ytd_fen=0):
        """Setup session.execute side effects for calculate().

        Call order:
        1. select PayrollBatch → batch
        2. select EmploymentAssignment → assignments
        Per assignment:
          3. select EmploymentContract → contract
          4. select DailyAttendance → att_rows
          5. select ytd aggregate → ytd_fen
        """
        side_effects = [
            _make_scalar_result(batch),        # 1. batch lookup
            _make_scalars_result(assignments),  # 2. assignments
        ]
        for _ in assignments:
            side_effects.append(_make_scalar_result(contract))    # 3. contract
            side_effects.append(_make_scalars_result(att_rows))   # 4. attendance
            side_effects.append(_make_scalar_only_result(ytd_fen))  # 5. ytd
        session.execute = AsyncMock(side_effect=side_effects)

    async def test_calculate_fixed_monthly(self, svc, mock_session):
        """固定月薪合同 → base_salary_fen 从 pay_scheme 读取"""
        batch = _make_batch()
        asn = _make_assignment()
        contract = _make_contract(asn.id, {"type": "fixed_monthly", "base_salary_fen": 800000})
        att_rows = [_make_attendance() for _ in range(22)]  # 22天正常出勤

        self._setup_calculate_mocks(mock_session, batch, [asn], contract, att_rows)

        items = await svc.calculate(batch.id, mock_session)
        assert len(items) == 1
        assert items[0].base_salary_fen == 800000

    async def test_calculate_hourly(self, svc, mock_session):
        """时薪合同 → hours × hourly_rate_fen"""
        batch = _make_batch()
        asn = _make_assignment()
        contract = _make_contract(asn.id, {
            "type": "hourly",
            "hourly_rate_fen": 3000,  # 30元/小时
        })
        # 10天 × 480分钟 = 4800分钟 = 80小时
        att_rows = [_make_attendance(work_minutes=480) for _ in range(10)]

        self._setup_calculate_mocks(mock_session, batch, [asn], contract, att_rows)

        items = await svc.calculate(batch.id, mock_session)
        assert len(items) == 1
        assert items[0].base_salary_fen == 240000  # 80h * 3000分

    async def test_calculate_with_social_insurance(self, svc, mock_session):
        """社保计算结果 > 0"""
        batch = _make_batch()
        asn = _make_assignment()
        contract = _make_contract(asn.id, {"type": "fixed_monthly", "base_salary_fen": 600000})
        att_rows = [_make_attendance() for _ in range(22)]

        self._setup_calculate_mocks(mock_session, batch, [asn], contract, att_rows)

        items = await svc.calculate(batch.id, mock_session)
        assert items[0].social_insurance_fen > 0

    async def test_calculate_with_tax(self, svc, mock_session):
        """高于起征点的工资 → tax_fen > 0"""
        batch = _make_batch()
        asn = _make_assignment()
        # 15000元月薪，扣社保后约13500，减5000起征点=~8500应税
        contract = _make_contract(asn.id, {"type": "fixed_monthly", "base_salary_fen": 1500000})
        att_rows = [_make_attendance() for _ in range(22)]

        self._setup_calculate_mocks(mock_session, batch, [asn], contract, att_rows)

        items = await svc.calculate(batch.id, mock_session)
        assert items[0].tax_fen > 0
        # net = gross - si - tax
        assert items[0].net_fen == (
            items[0].gross_fen - items[0].social_insurance_fen - items[0].tax_fen
        )

    async def test_calculate_no_contract_uses_fallback(self, svc, mock_session):
        """无合同 → 使用回退默认值"""
        batch = _make_batch()
        asn = _make_assignment()
        att_rows = [_make_attendance() for _ in range(22)]

        # contract = None
        self._setup_calculate_mocks(
            mock_session, batch, [asn], None, att_rows
        )

        items = await svc.calculate(batch.id, mock_session)
        assert len(items) == 1
        # 回退到 _FALLBACK_BASE_FEN = 400000
        assert items[0].base_salary_fen == 400000


# ===========================================================================
# HRExportService test (1)
# ===========================================================================

class TestHRExportService:

    async def test_export_payroll_xlsx_has_3_sheets(self):
        """导出Excel包含3个sheet"""
        batch_id = uuid.uuid4()
        batch = MagicMock()
        batch.id = batch_id
        batch.period_year = 2026
        batch.period_month = 3
        batch.org_node_id = "store-001"
        batch.total_gross_fen = 1000000
        batch.total_net_fen = 800000

        item = MagicMock()
        item.assignment_id = uuid.uuid4()
        item.base_salary_fen = 600000
        item.overtime_fen = 10000
        item.deduction_late_fen = 5000
        item.deduction_absent_fen = 0
        item.gross_fen = 605000
        item.social_insurance_fen = 57000
        item.tax_fen = 3000
        item.net_fen = 545000

        session = AsyncMock()
        session.execute = AsyncMock(side_effect=[
            _make_scalar_result(batch),
            _make_scalars_result([item]),
        ])

        svc = HRExportService()
        buf = await svc.export_payroll_batch(batch_id, session)

        assert isinstance(buf, BytesIO)

        # Verify it's valid xlsx with 3 sheets
        from openpyxl import load_workbook
        wb = load_workbook(buf)
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames[0] == "月度汇总"
        assert wb.sheetnames[1] == "个人工资条"
        assert wb.sheetnames[2] == "部门成本"

        # Verify detail sheet has header + 1 data row
        ws_detail = wb["个人工资条"]
        assert ws_detail.max_row == 2  # header + 1 item
