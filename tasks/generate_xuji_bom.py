"""生成徐记海鲜BOM数据表 (.xlsx)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT = os.path.join(os.path.dirname(__file__), "徐记海鲜_BOM数据表_POC.xlsx")

# ── 样式定义 ──
BRAND_COLOR = "FF6B2C"
HEADER_FILL = PatternFill("solid", fgColor=BRAND_COLOR)
HEADER_FONT = Font(name="Noto Sans SC", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Noto Sans SC", bold=True, size=14, color=BRAND_COLOR)
BODY_FONT = Font(name="Noto Sans SC", size=10)
MONEY_FMT = '#,##0.00'
PCT_FMT = '0.0%'
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)
ALT_FILL = PatternFill("solid", fgColor="FFF5F0")


def style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        fill = ALT_FILL if (r - start_row) % 2 == 1 else PatternFill()
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')


def auto_width(ws, col_count, min_width=12):
    for c in range(1, col_count + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(min_width, 16)


# ── Sheet 1: BOM配方主表 ──
def create_sheet1(wb):
    ws = wb.active
    ws.title = "BOM配方主表"
    ws.merge_cells('A1:K1')
    ws['A1'] = "徐记海鲜 — BOM配方主表"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 36

    headers = [
        "配方ID", "菜品名称", "菜品编码", "版本", "生效日期",
        "是否激活", "出品率", "标准出品数", "售价（元）",
        "理论成本（元）", "理论成本率"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    dishes = [
        ("BOM-001", "蒜蓉波士顿龙虾", "XJ-D001", "v1.0", "2026-03-15", "是", 0.85, 1, 168.00, 52.80, 0.314),
        ("BOM-002", "剁椒鱼头", "XJ-D002", "v1.0", "2026-03-15", "是", 0.90, 1, 128.00, 35.20, 0.275),
        ("BOM-003", "清蒸石斑鱼", "XJ-D003", "v1.0", "2026-03-15", "是", 0.82, 1, 238.00, 78.50, 0.330),
        ("BOM-004", "蟹黄豆腐", "XJ-D004", "v1.0", "2026-03-15", "是", 0.95, 1, 68.00, 21.00, 0.309),
        ("BOM-005", "湘味小炒黄牛肉", "XJ-D005", "v1.0", "2026-03-15", "是", 0.88, 1, 78.00, 22.10, 0.283),
        ("BOM-006", "口味虾", "XJ-D006", "v1.0", "2026-03-15", "是", 0.75, 1, 128.00, 42.00, 0.328),
        ("BOM-007", "红烧甲鱼", "XJ-D007", "v1.0", "2026-03-15", "是", 0.80, 1, 198.00, 62.00, 0.313),
        ("BOM-008", "蒸蛋", "XJ-D008", "v1.0", "2026-03-15", "是", 0.98, 1, 28.00, 5.60, 0.200),
    ]
    for r, dish in enumerate(dishes, 4):
        for c, val in enumerate(dish, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 9 or c == 10:
                cell.number_format = MONEY_FMT
            elif c == 11 or c == 7:
                cell.number_format = PCT_FMT

    style_body(ws, 4, 4 + len(dishes) - 1, len(headers))
    auto_width(ws, len(headers))


# ── Sheet 2: BOM食材明细 ──
def create_sheet2(wb):
    ws = wb.create_sheet("BOM食材明细")
    ws.merge_cells('A1:J1')
    ws['A1'] = "徐记海鲜 — BOM食材明细"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 36

    headers = [
        "配方ID", "食材名称", "食材编码", "标准用量",
        "单位", "单位成本（元）", "损耗系数", "实际用量",
        "行项成本（元）", "成本占比"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    # 蒜蓉波士顿龙虾 BOM明细
    items = [
        ("BOM-001", "波士顿龙虾", "XJ-I001", 0.50, "kg", 180.00, 0.15, None, None, None),
        ("BOM-001", "蒜蓉酱", "XJ-I002", 0.05, "kg", 40.00, 0.05, None, None, None),
        ("BOM-001", "粉丝", "XJ-I003", 0.10, "kg", 12.00, 0.02, None, None, None),
        ("BOM-001", "葱姜", "XJ-I004", 0.03, "kg", 8.00, 0.10, None, None, None),
        ("BOM-002", "鳙鱼头", "XJ-I010", 1.20, "kg", 22.00, 0.12, None, None, None),
        ("BOM-002", "剁椒酱", "XJ-I011", 0.08, "kg", 35.00, 0.05, None, None, None),
        ("BOM-002", "豆豉", "XJ-I012", 0.02, "kg", 25.00, 0.03, None, None, None),
        ("BOM-003", "石斑鱼", "XJ-I020", 0.80, "kg", 85.00, 0.18, None, None, None),
        ("BOM-003", "蒸鱼豉油", "XJ-I021", 0.03, "L", 18.00, 0.02, None, None, None),
        ("BOM-004", "蟹黄", "XJ-I030", 0.15, "kg", 120.00, 0.08, None, None, None),
        ("BOM-004", "嫩豆腐", "XJ-I031", 0.30, "kg", 6.00, 0.05, None, None, None),
        ("BOM-005", "黄牛肉", "XJ-I040", 0.25, "kg", 65.00, 0.10, None, None, None),
        ("BOM-005", "青椒", "XJ-I041", 0.15, "kg", 8.00, 0.08, None, None, None),
        ("BOM-006", "小龙虾", "XJ-I050", 1.00, "kg", 35.00, 0.20, None, None, None),
        ("BOM-006", "香料包", "XJ-I051", 0.05, "kg", 60.00, 0.03, None, None, None),
        ("BOM-007", "甲鱼", "XJ-I060", 0.80, "kg", 68.00, 0.15, None, None, None),
        ("BOM-007", "料酒", "XJ-I061", 0.05, "L", 15.00, 0.02, None, None, None),
        ("BOM-008", "鸡蛋", "XJ-I070", 4.00, "个", 1.20, 0.05, None, None, None),
        ("BOM-008", "高汤", "XJ-I071", 0.10, "L", 10.00, 0.03, None, None, None),
    ]

    for r, item in enumerate(items, 4):
        row_data = list(item)
        std_qty = row_data[3]
        unit_cost = row_data[5]
        waste = row_data[6]
        actual_qty = round(std_qty * (1 + waste), 4)
        line_cost = round(actual_qty * unit_cost, 2)
        row_data[7] = actual_qty
        row_data[8] = line_cost
        row_data[9] = None  # 占比稍后算

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (6, 8, 9):
                cell.number_format = MONEY_FMT
            elif c == 7:
                cell.number_format = PCT_FMT

    # 计算成本占比（按配方分组）
    from collections import defaultdict
    bom_totals = defaultdict(float)
    for r in range(4, 4 + len(items)):
        bom_id = ws.cell(row=r, column=1).value
        line_cost = ws.cell(row=r, column=9).value or 0
        bom_totals[bom_id] += line_cost

    for r in range(4, 4 + len(items)):
        bom_id = ws.cell(row=r, column=1).value
        line_cost = ws.cell(row=r, column=9).value or 0
        total = bom_totals[bom_id]
        pct = line_cost / total if total > 0 else 0
        cell = ws.cell(row=r, column=10, value=pct)
        cell.number_format = PCT_FMT

    style_body(ws, 4, 4 + len(items) - 1, len(headers))
    auto_width(ws, len(headers))


# ── Sheet 3: 招牌菜成本汇总 ──
def create_sheet3(wb):
    ws = wb.create_sheet("招牌菜成本汇总")
    ws.merge_cells('A1:G1')
    ws['A1'] = "徐记海鲜 — 招牌菜成本对比分析"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 36

    headers = ["菜品", "售价¥", "理论成本¥", "成本率", "关键食材", "成本占比最大项", "优化建议"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    data = [
        ("蒜蓉波士顿龙虾", 168, 52.80, 0.314, "龙虾0.5kg", "波士顿龙虾 92%", "探索国产替代或季节性采购"),
        ("剁椒鱼头", 128, 35.20, 0.275, "鳙鱼头1.2kg", "鳙鱼头 89%", "本地养殖合作降价"),
        ("清蒸石斑鱼", 238, 78.50, 0.330, "石斑鱼0.8kg", "石斑鱼 96%", "预订制减少死鱼损耗"),
        ("蟹黄豆腐", 68, 21.00, 0.309, "蟹黄0.15kg", "蟹黄 93%", "蟹黄批量预处理降损耗"),
        ("湘味小炒黄牛肉", 78, 22.10, 0.283, "黄牛肉0.25kg", "黄牛肉 88%", "锁定月度供应价"),
        ("口味虾", 128, 42.00, 0.328, "小龙虾1.0kg", "小龙虾 86%", "旺季前锁价，淡季减频"),
        ("红烧甲鱼", 198, 62.00, 0.313, "甲鱼0.8kg", "甲鱼 92%", "控制单份克重标准化"),
        ("蒸蛋", 28, 5.60, 0.200, "鸡蛋4个", "鸡蛋 88%", "高毛利引流菜，维持现状"),
    ]
    for r, row in enumerate(data, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (2, 3):
                cell.number_format = MONEY_FMT
            elif c == 4:
                cell.number_format = PCT_FMT

    style_body(ws, 4, 4 + len(data) - 1, len(headers))
    auto_width(ws, len(headers))

    # 汇总行
    sum_row = 4 + len(data)
    ws.cell(row=sum_row, column=1, value="平均").font = Font(bold=True, name="Noto Sans SC")
    ws.cell(row=sum_row, column=2, value=round(sum(d[1] for d in data) / len(data), 2)).number_format = MONEY_FMT
    ws.cell(row=sum_row, column=3, value=round(sum(d[2] for d in data) / len(data), 2)).number_format = MONEY_FMT
    ws.cell(row=sum_row, column=4, value=round(sum(d[3] for d in data) / len(data), 3)).number_format = PCT_FMT


# ── Sheet 4: 月度采购汇总模板 ──
def create_sheet4(wb):
    ws = wb.create_sheet("月度采购汇总")
    ws.merge_cells('A1:H1')
    ws['A1'] = "徐记海鲜 — 月度采购汇总（模板）"
    ws['A1'].font = TITLE_FONT
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 36

    headers = ["食材", "编码", "单位", "月用量", "单价¥", "月总成本¥", "占总成本比", "同比变化"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, len(headers))

    # 模拟月度数据
    monthly = [
        ("波士顿龙虾", "XJ-I001", "kg", 150, 180.00, 27000.00, 0.22, 0.05),
        ("小龙虾", "XJ-I050", "kg", 300, 35.00, 10500.00, 0.086, -0.10),
        ("石斑鱼", "XJ-I020", "kg", 120, 85.00, 10200.00, 0.083, 0.08),
        ("鳙鱼头", "XJ-I010", "kg", 360, 22.00, 7920.00, 0.065, -0.03),
        ("甲鱼", "XJ-I060", "kg", 100, 68.00, 6800.00, 0.055, 0.12),
        ("黄牛肉", "XJ-I040", "kg", 75, 65.00, 4875.00, 0.040, 0.02),
        ("蟹黄", "XJ-I030", "kg", 30, 120.00, 3600.00, 0.029, 0.15),
        ("蒜蓉酱", "XJ-I002", "kg", 50, 40.00, 2000.00, 0.016, 0.00),
        ("香料包", "XJ-I051", "kg", 25, 60.00, 1500.00, 0.012, -0.05),
        ("鸡蛋", "XJ-I070", "个", 1200, 1.20, 1440.00, 0.012, 0.03),
    ]
    for r, row in enumerate(monthly, 4):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (5, 6):
                cell.number_format = MONEY_FMT
            elif c in (7, 8):
                cell.number_format = PCT_FMT

    style_body(ws, 4, 4 + len(monthly) - 1, len(headers))

    # 合计行
    sum_row = 4 + len(monthly)
    ws.cell(row=sum_row, column=1, value="合计").font = Font(bold=True, name="Noto Sans SC")
    total_cost = sum(d[5] for d in monthly)
    ws.cell(row=sum_row, column=6, value=total_cost).number_format = MONEY_FMT
    ws.cell(row=sum_row, column=6).font = Font(bold=True, name="Noto Sans SC")

    auto_width(ws, len(headers))


def main():
    wb = openpyxl.Workbook()
    create_sheet1(wb)
    create_sheet2(wb)
    create_sheet3(wb)
    create_sheet4(wb)
    wb.save(OUTPUT)
    print(f"✅ BOM数据表已生成: {OUTPUT}")


if __name__ == "__main__":
    main()
